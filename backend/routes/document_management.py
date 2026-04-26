"""
Document Management Routes
Dynamic template-based document generation with PDF/Excel export
"""
from flask import Blueprint, request, jsonify, send_file, render_template_string, abort
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db
from models.document_management import (
    DocumentTemplate, Document, DocumentRevision, 
    DocumentCategory, DocumentAttachment, DocumentLog
)
from models.user import User
from datetime import datetime
from utils import generate_number
import os
import json
from io import BytesIO
from utils.timezone import get_local_now, get_local_today

# For PDF generation
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError) as e:
    WEASYPRINT_AVAILABLE = False
    print(f"Warning: WeasyPrint not available. PDF generation will be limited. Error: {e}")

# For Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel generation disabled.")

document_bp = Blueprint('document', __name__, url_prefix='/api/documents')


@document_bp.route('/templates', methods=['GET'])
@jwt_required()
def get_templates():
    """Get all document templates"""
    try:
        document_type = request.args.get('document_type')
        
        query = DocumentTemplate.query.filter_by(is_active=True)
        
        if document_type:
            query = query.filter_by(document_type=document_type)
        
        templates = query.all()
        
        return jsonify({
            'templates': [{
                'id': t.id,
                'template_name': t.template_name,
                'template_code': t.template_code,
                'document_type': t.document_type,
                'paper_size': t.paper_size,
                'orientation': t.orientation,
                'is_default': t.is_default
            } for t in templates]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@document_bp.route('/templates', methods=['POST'])
@jwt_required()
def create_template():
    """Create new document template"""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        template = DocumentTemplate(
            template_name=data['template_name'],
            template_code=data['template_code'],
            document_type=data['document_type'],
            template_structure=data['template_structure'],
            paper_size=data.get('paper_size', 'A4'),
            orientation=data.get('orientation', 'portrait'),
            margins=data.get('margins', {'top': 20, 'right': 20, 'bottom': 20, 'left': 20}),
            header_template=data.get('header_template'),
            footer_template=data.get('footer_template'),
            available_fields=data.get('available_fields', []),
            required_fields=data.get('required_fields', []),
            font_family=data.get('font_family', 'Arial'),
            font_size=data.get('font_size', 10),
            custom_css=data.get('custom_css'),
            is_default=data.get('is_default', False),
            created_by=current_user_id
        )
        
        db.session.add(template)
        db.session.commit()
        
        return jsonify({
            'message': 'Template created successfully',
            'template_id': template.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@document_bp.route('/templates/<int:id>', methods=['GET'])
@jwt_required()
def get_template(id):
    """Get single template"""
    try:
        template = db.session.get(DocumentTemplate, id) or abort(404)
        
        return jsonify({
            'template': {
                'id': template.id,
                'template_name': template.template_name,
                'template_code': template.template_code,
                'document_type': template.document_type,
                'paper_size': template.paper_size,
                'orientation': template.orientation,
                'margins': template.margins,
                'font_family': template.font_family,
                'font_size': template.font_size,
                'header_template': template.header_template,
                'footer_template': template.footer_template,
                'template_structure': template.template_structure,
                'custom_css': template.custom_css,
                'is_default': template.is_default,
                'is_active': template.is_active
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@document_bp.route('/templates/<int:id>', methods=['PUT'])
@jwt_required()
def update_template(id):
    """Update document template"""
    try:
        template = db.session.get(DocumentTemplate, id) or abort(404)
        data = request.get_json()
        
        template.template_name = data.get('template_name', template.template_name)
        template.template_code = data.get('template_code', template.template_code)
        template.document_type = data.get('document_type', template.document_type)
        template.paper_size = data.get('paper_size', template.paper_size)
        template.orientation = data.get('orientation', template.orientation)
        template.margins = data.get('margins', template.margins)
        template.font_family = data.get('font_family', template.font_family)
        template.font_size = data.get('font_size', template.font_size)
        template.header_template = data.get('header_template', template.header_template)
        template.footer_template = data.get('footer_template', template.footer_template)
        template.template_structure = data.get('template_structure', template.template_structure)
        template.custom_css = data.get('custom_css', template.custom_css)
        template.is_default = data.get('is_default', template.is_default)
        template.is_active = data.get('is_active', template.is_active)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Template updated successfully',
            'template_id': template.id
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@document_bp.route('/preview-data/<document_type>', methods=['GET'])
@jwt_required()
def get_preview_data(document_type):
    """Get sample data for template preview based on document type"""
    try:
        from models.sales import SalesOrder, SalesOrderItem
        from models.production import WorkOrder
        from models.purchasing import PurchaseOrder
        from models.inventory import Product
        from models.crm import Customer
        
        # Get company info from settings or use default
        company_data = {
            'name': 'PT. CONTOH PERUSAHAAN',
            'address': 'Jl. Contoh No. 123, Jakarta 12345',
            'phone': '021-1234567',
            'email': 'info@contoh.com',
            'npwp': '01.234.567.8-901.000'
        }
        
        preview_data = {'company': company_data}
        
        # Get real sample data based on document type
        if document_type in ['perintah_kerja', 'penyelesaian_barang_jadi', 'formula_produksi']:
            work_order = WorkOrder.query.order_by(WorkOrder.id.desc()).first()
            if work_order:
                preview_data['workOrder'] = {
                    'number': work_order.work_order_number,
                    'transDate': work_order.start_date.strftime('%d/%m/%Y') if work_order.start_date else '-',
                    'dueDate': work_order.due_date.strftime('%d/%m/%Y') if work_order.due_date else '-',
                    'status': work_order.status,
                    'manufacturePlan': f'MP-{work_order.id:04d}',
                    'description': work_order.notes or '',
                    'item': {'name': work_order.product.name if work_order.product else '-', 'code': work_order.product.sku if work_order.product else '-'},
                    'quantity': work_order.quantity,
                    'itemUnit': {'name': work_order.product.unit if work_order.product else 'Pcs'},
                    'totalMaterial': f'Rp {work_order.material_cost or 0:,.0f}',
                    'totalExpense': f'Rp {work_order.overhead_cost or 0:,.0f}',
                    'totalAmount': f'Rp {work_order.total_cost or 0:,.0f}'
                }
                # Materials
                preview_data['workOrderMaterial'] = []
                if hasattr(work_order, 'materials'):
                    for mat in work_order.materials[:5]:
                        preview_data['workOrderMaterial'].append({
                            'item': {'code': mat.product.sku if mat.product else '-', 'name': mat.product.name if mat.product else '-'},
                            'quantity': mat.quantity_required,
                            'itemUnit': {'name': mat.product.unit if mat.product else 'Pcs'},
                            'unitPrice': f'Rp {mat.unit_cost or 0:,.0f}',
                            'amount': f'Rp {(mat.quantity_required or 0) * (mat.unit_cost or 0):,.0f}'
                        })
        
        elif document_type in ['pesanan_penjualan', 'faktur_penjualan', 'pengiriman_pesanan']:
            sales_order = SalesOrder.query.order_by(SalesOrder.id.desc()).first()
            if sales_order:
                preview_data['salesOrder'] = {
                    'number': sales_order.order_number,
                    'orderDate': sales_order.order_date.strftime('%d/%m/%Y') if sales_order.order_date else '-',
                    'requiredDate': sales_order.required_date.strftime('%d/%m/%Y') if sales_order.required_date else '-',
                    'customer': {
                        'name': sales_order.customer.name if sales_order.customer else '-',
                        'address': sales_order.customer.address if sales_order.customer else '-',
                        'phone': sales_order.customer.phone if sales_order.customer else '-'
                    },
                    'deliveryAddress': sales_order.delivery_address or '-',
                    'subtotal': f'Rp {sales_order.subtotal or 0:,.0f}',
                    'taxAmount': f'Rp {sales_order.tax_amount or 0:,.0f}',
                    'discountAmount': f'Rp {sales_order.discount_amount or 0:,.0f}',
                    'totalAmount': f'Rp {sales_order.total_amount or 0:,.0f}',
                    'notes': sales_order.notes or ''
                }
                # Items
                preview_data['salesOrderItem'] = []
                for i, item in enumerate(sales_order.items[:5], 1):
                    preview_data['salesOrderItem'].append({
                        'lineNo': i,
                        'product': {'code': item.product.sku if item.product else '-', 'name': item.product.name if item.product else '-'},
                        'quantity': item.quantity,
                        'unit': item.product.unit if item.product else 'Pcs',
                        'unitPrice': f'Rp {item.unit_price or 0:,.0f}',
                        'discount': f'Rp {item.discount_amount or 0:,.0f}',
                        'amount': f'Rp {item.total_price or 0:,.0f}'
                    })
        
        elif document_type in ['pesanan_pembelian', 'faktur_pembelian', 'penerimaan_barang']:
            purchase_order = PurchaseOrder.query.order_by(PurchaseOrder.id.desc()).first()
            if purchase_order:
                preview_data['purchaseOrder'] = {
                    'number': purchase_order.po_number,
                    'orderDate': purchase_order.order_date.strftime('%d/%m/%Y') if purchase_order.order_date else '-',
                    'supplier': {
                        'name': purchase_order.supplier.name if purchase_order.supplier else '-',
                        'address': purchase_order.supplier.address if purchase_order.supplier else '-'
                    },
                    'totalAmount': f'Rp {purchase_order.total_amount or 0:,.0f}'
                }
        
        return jsonify(preview_data), 200
        
    except Exception as e:
        # Return mock data if error
        return jsonify({
            'company': {'name': 'PT. CONTOH PERUSAHAAN', 'address': 'Jl. Contoh No. 123, Jakarta'},
            'workOrder': {'number': 'WO-SAMPLE-001', 'transDate': '28/11/2024', 'item': {'name': 'Produk Sample'}, 'quantity': 100, 'itemUnit': {'name': 'Pcs'}},
            'error': str(e)
        }), 200


@document_bp.route('/render-pdf', methods=['POST'])
@jwt_required()
def render_pdf_from_template():
    """Render PDF directly from template structure with data"""
    try:
        if not WEASYPRINT_AVAILABLE:
            return jsonify({'error': 'PDF generation not available. Install WeasyPrint.'}), 500
        
        data = request.get_json()
        template_structure = data.get('template_structure', {})
        document_data = data.get('document_data', {})
        paper_size = data.get('paper_size', 'A4')
        orientation = data.get('orientation', 'portrait')
        margins = data.get('margins', {'left': 12, 'right': 12, 'top': 14, 'bottom': 11})
        
        # Paper dimensions in mm
        paper_sizes = {
            'A4': {'width': 210, 'height': 297},
            'A5': {'width': 148, 'height': 210},
            'Letter': {'width': 216, 'height': 279},
            'Legal': {'width': 216, 'height': 356},
            'F4': {'width': 215, 'height': 330}
        }
        paper = paper_sizes.get(paper_size, paper_sizes['A4'])
        page_width = paper['height'] if orientation == 'landscape' else paper['width']
        page_height = paper['width'] if orientation == 'landscape' else paper['height']
        
        # Helper to get nested value from data
        def get_field_value(field_path, data):
            parts = field_path.split('.')
            value = data
            for part in parts:
                if isinstance(value, dict) and part in value:
                    value = value[part]
                else:
                    return ''
            return str(value) if value is not None else ''
        
        # Build HTML from template structure
        bands = template_structure.get('bands', [])
        content_html = ''
        
        for band in bands:
            if not band.get('visible', True):
                continue
            
            band_height = band.get('height', 20)
            band_html = f'<div style="position: relative; height: {band_height}mm; margin-left: {margins["left"]}mm; margin-right: {margins["right"]}mm;">'
            
            for el in band.get('elements', []):
                el_type = el.get('type', 'label')
                x = el.get('x', 0)
                y = el.get('y', 0)
                width = el.get('width', 50)
                height = el.get('height', 6)
                font_size = el.get('fontSize', 9)
                font_weight = el.get('fontWeight', 'normal')
                text_align = el.get('textAlign', 'left')
                color = el.get('color', '#000')
                bg_color = el.get('backgroundColor', 'transparent')
                border_width = el.get('borderWidth', 0)
                border_color = el.get('borderColor', '#000')
                
                style = f'''
                    position: absolute;
                    left: {x}mm; top: {y}mm;
                    width: {width}mm; height: {height}mm;
                    font-size: {font_size}pt;
                    font-weight: {font_weight};
                    text-align: {text_align};
                    color: {color};
                    background-color: {bg_color};
                    display: flex; align-items: center;
                    padding: 0 1mm;
                    overflow: hidden;
                    white-space: nowrap;
                '''
                
                if el_type == 'box':
                    style += f'border: {border_width}px solid {border_color};'
                elif el_type == 'line':
                    style += f'border-bottom: {border_width}px solid {border_color}; height: 0;'
                
                content = ''
                if el_type == 'label':
                    content = el.get('content', '')
                elif el_type == 'field':
                    field_path = el.get('fieldPath', '')
                    content = get_field_value(field_path, document_data)
                elif el_type == 'image':
                    img_src = el.get('imageSrc', '')
                    if img_src:
                        content = f'<img src="{img_src}" style="max-width: 100%; max-height: 100%; object-fit: contain;" />'
                
                band_html += f'<div style="{style}">{content}</div>'
            
            band_html += '</div>'
            content_html += band_html
        
        # Full HTML document
        html = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{
                    size: {page_width}mm {page_height}mm;
                    margin: {margins["top"]}mm {margins["right"]}mm {margins["bottom"]}mm {margins["left"]}mm;
                }}
                body {{
                    font-family: Arial, sans-serif;
                    margin: 0;
                    padding: 0;
                }}
            </style>
        </head>
        <body>
            {content_html}
        </body>
        </html>
        '''
        
        # Generate PDF
        pdf_buffer = BytesIO()
        HTML(string=html).write_pdf(pdf_buffer)
        pdf_buffer.seek(0)
        
        filename = data.get('filename', f'document-{get_local_now().strftime("%Y%m%d%H%M%S")}.pdf')
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@document_bp.route('/generate', methods=['POST'])
@jwt_required()
def generate_document():
    """Generate document from template"""
    try:
        current_user_id = get_jwt_identity()
        data = request.get_json()
        
        template = db.session.get(DocumentTemplate, data['template_id']) or abort(404)
        
        # Generate document number
        document_number = generate_number(
            data.get('number_prefix', 'DOC'),
            Document,
            'document_number'
        )
        
        # Create document
        document = Document(
            document_number=document_number,
            document_title=data.get('document_title'),
            document_type=template.document_type,
            template_id=template.id,
            document_data=data['document_data'],
            reference_type=data.get('reference_type'),
            reference_id=data.get('reference_id'),
            reference_number=data.get('reference_number'),
            signature_fields=data.get('signature_fields'),
            document_date=datetime.fromisoformat(data['document_date']) if data.get('document_date') else get_local_now(),
            created_by=current_user_id
        )
        
        # Generate HTML content for preview
        html_content = render_document_html(template, data['document_data'])
        document.html_content = html_content
        
        db.session.add(document)
        db.session.flush()
        
        # Log activity
        log = DocumentLog(
            document_id=document.id,
            activity_type='created',
            activity_description=f'Document {document_number} created',
            user_id=current_user_id
        )
        db.session.add(log)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Document generated successfully',
            'document_id': document.id,
            'document_number': document_number
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@document_bp.route('/<int:id>/preview', methods=['GET'])
@jwt_required()
def preview_document(id):
    """Get document HTML preview"""
    try:
        document = db.session.get(Document, id) or abort(404)
        
        return jsonify({
            'document': {
                'id': document.id,
                'document_number': document.document_number,
                'document_title': document.document_title,
                'html_content': document.html_content,
                'status': document.status
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@document_bp.route('/<int:id>/pdf', methods=['GET'])
@jwt_required()
def generate_pdf(id):
    """Generate and download PDF"""
    try:
        current_user_id = get_jwt_identity()
        document = db.session.get(Document, id) or abort(404)
        
        if not WEASYPRINT_AVAILABLE:
            return jsonify({'error': 'PDF generation not available. Install WeasyPrint.'}), 500
        
        # Generate PDF from HTML
        html_content = document.html_content
        pdf_buffer = BytesIO()
        
        HTML(string=html_content).write_pdf(pdf_buffer)
        pdf_buffer.seek(0)
        
        # Update print count
        document.print_count += 1
        document.last_printed_at = get_local_now()
        document.printed_by = current_user_id
        
        # Log activity
        log = DocumentLog(
            document_id=document.id,
            activity_type='printed',
            activity_description=f'Document {document.document_number} printed as PDF',
            user_id=current_user_id
        )
        db.session.add(log)
        db.session.commit()
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'{document.document_number}.pdf'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@document_bp.route('/<int:id>/excel', methods=['GET'])
@jwt_required()
def generate_excel(id):
    """Generate and download Excel"""
    try:
        current_user_id = get_jwt_identity()
        document = db.session.get(Document, id) or abort(404)
        
        if not OPENPYXL_AVAILABLE:
            return jsonify({'error': 'Excel generation not available. Install openpyxl.'}), 500
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = document.document_type.upper()
        
        # Styling
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add document data
        row = 1
        
        # Document header
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = document.document_title or document.document_type.upper()
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        # Document number and date
        ws[f'A{row}'] = 'Document Number:'
        ws[f'B{row}'] = document.document_number
        ws[f'A{row}'].font = header_font
        row += 1
        
        ws[f'A{row}'] = 'Date:'
        ws[f'B{row}'] = document.document_date.strftime('%Y-%m-%d')
        ws[f'A{row}'].font = header_font
        row += 2
        
        # Document data
        document_data = document.document_data
        
        for key, value in document_data.items():
            if isinstance(value, list):
                # Table data
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'A{row}'].font = header_font
                row += 1
                
                if value and isinstance(value[0], dict):
                    # Headers
                    col = 1
                    for header in value[0].keys():
                        cell = ws.cell(row=row, column=col)
                        cell.value = header.replace('_', ' ').title()
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border
                        col += 1
                    row += 1
                    
                    # Data rows
                    for item in value:
                        col = 1
                        for val in item.values():
                            cell = ws.cell(row=row, column=col)
                            cell.value = val
                            cell.border = border
                            col += 1
                        row += 1
                row += 1
            else:
                # Simple key-value
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = header_font
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Log activity
        log = DocumentLog(
            document_id=document.id,
            activity_type='exported',
            activity_description=f'Document {document.document_number} exported to Excel',
            user_id=current_user_id
        )
        db.session.add(log)
        db.session.commit()
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{document.document_number}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@document_bp.route('', methods=['GET'])
@jwt_required()
def get_documents():
    """Get all documents with filters"""
    try:
        document_type = request.args.get('document_type')
        status = request.args.get('status')
        reference_type = request.args.get('reference_type')
        
        query = Document.query
        
        if document_type:
            query = query.filter_by(document_type=document_type)
        if status:
            query = query.filter_by(status=status)
        if reference_type:
            query = query.filter_by(reference_type=reference_type)
        
        documents = query.order_by(Document.created_at.desc()).all()
        
        return jsonify({
            'documents': [{
                'id': d.id,
                'document_number': d.document_number,
                'document_title': d.document_title,
                'document_type': d.document_type,
                'reference_number': d.reference_number,
                'status': d.status,
                'document_date': d.document_date.isoformat(),
                'print_count': d.print_count,
                'created_at': d.created_at.isoformat()
            } for d in documents]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@document_bp.route('/<int:id>/print', methods=['POST'])
@jwt_required()
def record_print(id):
    """Record document print"""
    try:
        current_user_id = get_jwt_identity()
        document = db.session.get(Document, id) or abort(404)
        
        document.print_count += 1
        document.last_printed_at = get_local_now()
        document.printed_by = current_user_id
        document.status = 'printed'
        
        # Log activity
        log = DocumentLog(
            document_id=document.id,
            activity_type='printed',
            activity_description=f'Document {document.document_number} printed',
            user_id=current_user_id
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'message': 'Print recorded',
            'print_count': document.print_count
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@document_bp.route('/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_document(id):
    """Delete document"""
    try:
        document = db.session.get(Document, id) or abort(404)
        
        # Delete related logs
        DocumentLog.query.filter_by(document_id=id).delete()
        
        db.session.delete(document)
        db.session.commit()
        
        return jsonify({'message': 'Document deleted'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@document_bp.route('/templates/<int:id>/set-default', methods=['PUT'])
@jwt_required()
def set_default_template(id):
    """Set template as default for its document type"""
    try:
        template = db.session.get(DocumentTemplate, id) or abort(404)
        
        # Remove default from other templates of same type
        DocumentTemplate.query.filter_by(
            document_type=template.document_type,
            is_default=True
        ).update({'is_default': False})
        
        # Set this template as default
        template.is_default = True
        db.session.commit()
        
        return jsonify({
            'message': 'Default template updated',
            'template_id': template.id
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@document_bp.route('/templates/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_template(id):
    """Delete document template"""
    try:
        template = db.session.get(DocumentTemplate, id) or abort(404)
        
        # Check if template is in use
        doc_count = Document.query.filter_by(template_id=id).count()
        if doc_count > 0:
            return jsonify({
                'error': f'Template is used by {doc_count} documents. Cannot delete.'
            }), 400
        
        db.session.delete(template)
        db.session.commit()
        
        return jsonify({'message': 'Template deleted'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


def render_document_html(template, data):
    """Render document HTML from template and data"""
    
    # Get template values with defaults
    paper_size = template.paper_size or 'A4'
    margins = template.margins or {'top': 20, 'right': 20, 'bottom': 20, 'left': 20}
    font_family = template.font_family or 'Arial'
    font_size = template.font_size or 10
    custom_css = template.custom_css or ''
    
    # Basic HTML template
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: {paper_size};
                margin: {margins.get('top', 20)}mm {margins.get('right', 20)}mm 
                        {margins.get('bottom', 20)}mm {margins.get('left', 20)}mm;
            }}
            body {{
                font-family: {font_family}, sans-serif;
                font-size: {font_size}pt;
                line-height: 1.6;
            }}
            .header {{
                text-align: center;
                margin-bottom: 20px;
                border-bottom: 2px solid #333;
                padding-bottom: 10px;
            }}
            .content {{
                margin: 20px 0;
            }}
            .footer {{
                margin-top: 30px;
                border-top: 1px solid #ccc;
                padding-top: 10px;
                font-size: 9pt;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin: 15px 0;
            }}
            th, td {{
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }}
            th {{
                background-color: #f2f2f2;
                font-weight: bold;
            }}
            .field-label {{
                font-weight: bold;
                width: 200px;
            }}
            {custom_css}
        </style>
    </head>
    <body>
    """
    
    # Add header
    if template.header_template:
        html += "<div class='header'>"
        html += f"<h2>{template.header_template.get('company_name', 'Company Name')}</h2>"
        html += f"<p>{template.header_template.get('address', '')}</p>"
        html += "</div>"
    
    # Add content
    html += "<div class='content'>"
    html += f"<h3>{template.template_name}</h3>"
    
    # Render data fields
    for key, value in data.items():
        if isinstance(value, list):
            # Render as table
            html += f"<h4>{key.replace('_', ' ').title()}</h4>"
            if value and isinstance(value[0], dict):
                html += "<table>"
                html += "<tr>"
                for header in value[0].keys():
                    html += f"<th>{header.replace('_', ' ').title()}</th>"
                html += "</tr>"
                for item in value:
                    html += "<tr>"
                    for val in item.values():
                        html += f"<td>{val}</td>"
                    html += "</tr>"
                html += "</table>"
        else:
            # Render as key-value
            html += f"<p><span class='field-label'>{key.replace('_', ' ').title()}:</span> {value}</p>"
    
    html += "</div>"
    
    # Add footer
    if template.footer_template:
        html += "<div class='footer'>"
        html += f"<p>{template.footer_template.get('notes', '')}</p>"
        html += "</div>"
    
    html += "</body></html>"
    
    return html


@document_bp.route('/categories', methods=['GET'])
@jwt_required()
def get_categories():
    """Get document categories"""
    try:
        categories = DocumentCategory.query.filter_by(is_active=True).all()
        
        return jsonify({
            'categories': [{
                'id': c.id,
                'category_name': c.category_name,
                'category_code': c.category_code,
                'number_prefix': c.number_prefix,
                'number_format': c.number_format
            } for c in categories]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@document_bp.route('/generate-from-sales-order/<int:sales_order_id>', methods=['POST'])
@jwt_required()
def generate_from_sales_order(sales_order_id):
    """Auto-generate Surat Jalan from Sales Order (like Accurate)"""
    try:
        current_user_id = get_jwt_identity()
        
        from utils.document_generator import generate_surat_jalan_from_sales_order
        document = generate_surat_jalan_from_sales_order(sales_order_id, current_user_id)
        
        return jsonify({
            'message': 'Surat Jalan generated successfully',
            'document_id': document.id,
            'document_number': document.document_number
        }), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@document_bp.route('/generate-from-work-order/<int:work_order_id>', methods=['POST'])
@jwt_required()
def generate_from_work_order(work_order_id):
    """Auto-generate SPK from Work Order (like Accurate)"""
    try:
        current_user_id = get_jwt_identity()
        
        from utils.document_generator import generate_spk_from_work_order
        document = generate_spk_from_work_order(work_order_id, current_user_id)
        
        return jsonify({
            'message': 'SPK generated successfully',
            'document_id': document.id,
            'document_number': document.document_number
        }), 201
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@document_bp.route('/dashboard', methods=['GET'])
@jwt_required()
def get_dashboard():
    """Get document management dashboard stats"""
    try:
        from sqlalchemy import func
        
        total_documents = Document.query.count()
        total_templates = DocumentTemplate.query.filter_by(is_active=True).count()
        
        # Documents by type
        by_type = db.session.query(
            Document.document_type,
            func.count(Document.id)
        ).group_by(Document.document_type).all()
        
        # Recent documents
        recent = Document.query.order_by(Document.created_at.desc()).limit(10).all()
        
        return jsonify({
            'statistics': {
                'total_documents': total_documents,
                'total_templates': total_templates,
                'by_type': [{'type': t[0], 'count': t[1]} for t in by_type]
            },
            'recent_documents': [{
                'id': d.id,
                'document_number': d.document_number,
                'document_type': d.document_type,
                'created_at': d.created_at.isoformat()
            } for d in recent]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
