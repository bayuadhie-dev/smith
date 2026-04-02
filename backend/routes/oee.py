from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db, Machine, MaintenanceRecord, MaintenanceSchedule, User
from utils.i18n import success_response, error_response, get_message
from models.oee import OEERecord, OEEDowntimeRecord, OEETarget, OEEAlert, MaintenanceImpact, OEEAnalytics, QualityDefect, MachineMonthlyTarget, DowntimeRootCause
from models.product_excel_schema import ProductNew
from utils import generate_number
from utils.helpers import detect_downtime_category
from datetime import datetime, date, timedelta
from sqlalchemy import func, and_, or_, desc
import json
import re
import io
from utils.timezone import get_local_now, get_local_today

oee_bp = Blueprint('oee', __name__)

@oee_bp.route('/records', methods=['GET'])
@oee_bp.route('/records/', methods=['GET'])
@jwt_required()
def get_records():
    """Get OEE records from both OEERecord and ShiftProduction"""
    try:
        from models.production import ShiftProduction
        
        machine_id = request.args.get('machine_id', type=int)
        limit = request.args.get('limit', 100, type=int)
        
        all_records = []
        
        # Get from OEERecord
        oee_query = OEERecord.query
        if machine_id:
            oee_query = oee_query.filter(OEERecord.machine_id == machine_id)
        
        oee_records = oee_query.order_by(OEERecord.record_date.desc()).limit(limit).all()
        for r in oee_records:
            all_records.append({
                'id': r.id,
                'source': 'oee_record',
                'record_number': r.record_number,
                'machine_id': r.machine_id,
                'machine_name': r.machine.name if r.machine else None,
                'record_date': r.record_date.isoformat() if r.record_date else None,
                'availability': float(r.availability) if r.availability else 0,
                'performance': float(r.performance) if r.performance else 0,
                'quality': float(r.quality) if r.quality else 0,
                'oee_percentage': float(r.oee) if r.oee else 0
            })
        
        # Get from ShiftProduction
        shift_query = ShiftProduction.query
        if machine_id:
            shift_query = shift_query.filter(ShiftProduction.machine_id == machine_id)
        
        shift_records = shift_query.order_by(ShiftProduction.production_date.desc()).limit(limit).all()
        for sp in shift_records:
            # Get product name from work order or product directly
            product_name = None
            target_qty = int(sp.target_quantity) if sp.target_quantity else 0
            actual_qty = int(sp.actual_quantity) if sp.actual_quantity else 0
            pack_per_karton = 50  # Default value
            
            # Try to get product name from product or work_order
            if sp.product:
                product_name = sp.product.name
            elif sp.work_order and sp.work_order.product:
                product_name = sp.work_order.product.name
            
            # Get pack_per_karton from ProductNew if product name available
            if product_name:
                try:
                    # Remove "WIP " prefix for matching
                    search_name = product_name.replace('WIP ', '').strip()
                    product_new = ProductNew.query.filter(
                        ProductNew.name.ilike(f'%{search_name}%')
                    ).first()
                    if product_new and product_new.pack_per_karton:
                        pack_per_karton = int(product_new.pack_per_karton)
                except Exception as e:
                    # If query fails, use default
                    pack_per_karton = 50
            
            # Get target from work order if not set
            if not target_qty and sp.work_order:
                target_qty = int(sp.work_order.quantity) if sp.work_order.quantity else 0
            
            # Parse ALL downtime entries from issues field
            # Format: "60 menit - Produk bocor (endseal kotor) [others]; 20 menit - Kain keluar jalur [others]; ..."
            downtime_breakdown = []
            total_downtime_minutes = 0  # Sum of NON-IDLE downtime entries only
            idle_time_minutes = 0  # Separate from downtime
            
            # IDLE TIME keywords - sinkron dengan daily controller
            idle_keywords = [
                # Tunggu kain
                'tunggu kain', 'ambil kain', 'menunggu kain', 'nunggu kain', 'kain belum datang',
                # Tunggu obat/tinta
                'tunggu obat', 'menunggu obat', 'nunggu obat', 'obat belum datang',
                'tunggu tinta', 'menunggu tinta', 'nunggu tinta',
                # Tunggu ingredient
                'tunggu ingredient', 'ingredient habis', 'tunggu bahan kimia',
                # Tunggu stiker
                'tunggu stiker', 'menunggu stiker', 'nunggu stiker', 'stiker belum datang',
                # Tunggu packing/packaging
                'tunggu packing', 'tunggu packaging', 'menunggu packing', 'nunggu packing',
                'packaging belum datang', 'box belum datang',
                # Tunggu mixing
                'tunggu mixing', 'menunggu mixing', 'nunggu mixing', 'mixing belum siap',
                # Tunggu label/karton/box
                'tunggu label', 'tunggu box', 'tunggu karton', 'tunggu lem',
                # Tunggu produk (dari mesin lain)
                'tunggu produk',
                # Tunggu temperatur
                'tunggu temperatur stabil', 'tunggu temperatur',
                'tunggu temperature stabil', 'tunggu temperature',
                # General tunggu
                'tunggu bahan', 'tunggu material', 'tunggu order', 'tunggu instruksi',
                'tunggu approval', 'tunggu qc', 'tunggu hasil qc',
                # xxx habis (idle karena kehabisan)
                'kain habis', 'stiker habis', 'packing habis', 'packaging habis',
                'label habis', 'karton habis', 'box habis', 'lem habis', 'tinta habis',
                'bahan habis', 'material habis',
                # English
                'waiting for', 'standby material', 'waiting material', 'no material',
                # Idle lainnya
                'idle', 'standby', 'menganggur', 'tidak ada order', 'no order', 'menhabiskan order'
            ]
            
            if sp.issues:
                # Split by semicolon
                issue_parts = sp.issues.split(';')
                for part in issue_parts:
                    part = part.strip()
                    if not part:
                        continue
                    # Parse: "60 menit - Produk bocor (endseal kotor) [others]"
                    # More flexible regex - captures number and reason
                    match = re.match(r'(\d+)\s*menit\s*-\s*(.+)', part, re.IGNORECASE)
                    if match:
                        duration = int(match.group(1))
                        reason_with_tag = match.group(2).strip()
                        
                        # Check for [idle] category tag BEFORE removing it
                        has_idle_tag = bool(re.search(r'\[idle\]', reason_with_tag, re.IGNORECASE))
                        
                        # Remove trailing category bracket if present [category]
                        reason = re.sub(r'\s*\[\w+\]\s*$', '', reason_with_tag).strip()
                        reason_lower = reason.lower()
                        
                        # Check if idle: either has [idle] tag OR matches idle keywords
                        is_idle = has_idle_tag or any(kw in reason_lower for kw in idle_keywords)
                        if is_idle:
                            # Idle time is SEPARATE from downtime
                            idle_time_minutes += duration
                        else:
                            # Only non-idle goes to downtime
                            total_downtime_minutes += duration
                        
                        # Add to breakdown for Top 3 display (with frequency tracking)
                        existing_dt = next((d for d in downtime_breakdown if d['reason'] == reason), None)
                        if existing_dt:
                            existing_dt['duration_minutes'] += duration
                            existing_dt['frequency'] = existing_dt.get('frequency', 1) + 1
                        else:
                            downtime_breakdown.append({'reason': reason, 'duration_minutes': duration, 'is_idle': is_idle, 'frequency': 1})
            
            # Fallback: use sum of category fields if issues parsing yielded 0
            if total_downtime_minutes == 0:
                # Try downtime_minutes field first
                if sp.downtime_minutes:
                    total_downtime_minutes = int(sp.downtime_minutes)
                else:
                    # Sum from individual category fields
                    total_downtime_minutes = (
                        (int(sp.downtime_mesin) if sp.downtime_mesin else 0) +
                        (int(sp.downtime_operator) if sp.downtime_operator else 0) +
                        (int(sp.downtime_material) if sp.downtime_material else 0) +
                        (int(sp.downtime_design) if sp.downtime_design else 0) +
                        (int(sp.downtime_others) if sp.downtime_others else 0)
                    )
            
            # Sort by duration descending and take top 3
            downtime_breakdown.sort(key=lambda x: x['duration_minutes'], reverse=True)
            top_3_downtime = downtime_breakdown[:3]
            
            # Extract shift number from shift field (e.g., "shift_1" -> 1)
            shift_num = 1
            if sp.shift:
                shift_match = re.search(r'(\d+)', str(sp.shift))
                if shift_match:
                    shift_num = int(shift_match.group(1))
            
            # Get planned runtime per shift (Shift 1 & 2 = 510, Shift 1 Friday = 540, Shift 3 = 450 minutes)
            # Friday (weekday() == 4) has longer shift due to Jumatan prayer
            is_friday = sp.production_date and sp.production_date.weekday() == 4
            if shift_num == 1:
                default_planned = 540 if is_friday else 510  # 06:30-15:00 = 8.5 hours (9 hours on Friday)
            elif shift_num == 2:
                default_planned = 510  # 15:00-23:00 = 8.5 hours
            else:
                default_planned = 450  # 23:00-06:30 = 7.5 hours
            
            planned_runtime = int(sp.planned_runtime) if sp.planned_runtime else default_planned
            
            # Runtime = Planned - Downtime - Idle (calculated, not from input)
            runtime_minutes = max(0, planned_runtime - total_downtime_minutes - idle_time_minutes)
            
            # Shift data with Grade A, B, C and Runtime/Downtime/Idle metrics
            shift_data = {
                'shift': shift_num,
                'grade_a': int(sp.good_quantity) if sp.good_quantity else 0,
                'grade_b': int(sp.rework_quantity) if sp.rework_quantity else 0,
                'grade_c': int(sp.reject_quantity) if sp.reject_quantity else 0,
                'total': int(sp.actual_quantity) if sp.actual_quantity else 0,
                'runtime_minutes': runtime_minutes,
                'downtime_minutes': total_downtime_minutes,
                'idle_time_minutes': idle_time_minutes,
                'planned_runtime': planned_runtime
            }
            
            all_records.append({
                'id': sp.id,
                'source': 'shift_production',
                'record_number': f"SP-{sp.id}",
                'machine_id': sp.machine_id,
                'machine_name': sp.machine.name if sp.machine else None,
                'record_date': sp.production_date.isoformat() if sp.production_date else None,
                'availability': float(sp.efficiency_rate) if sp.efficiency_rate else 0,
                'performance': 100.0,  # Default performance
                'quality': float(sp.quality_rate) if sp.quality_rate else 0,
                'oee_percentage': float(sp.oee_score) if sp.oee_score else 0,
                'product_name': product_name,
                'target_quantity': target_qty,
                'actual_quantity': actual_qty,
                'top_3_downtime': top_3_downtime,
                'shift_data': shift_data,
                'pack_per_karton': pack_per_karton,
                # New fields for runtime/downtime tracking
                'runtime_minutes': runtime_minutes,
                'total_downtime_minutes': total_downtime_minutes,
                'idle_time_minutes': idle_time_minutes,
                'planned_runtime_minutes': planned_runtime
            })
        
        # Sort by date and limit
        all_records.sort(key=lambda x: x['record_date'] or '', reverse=True)
        all_records = all_records[:limit]
        
        return jsonify({
            'records': all_records
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/export-excel', methods=['GET'])
@jwt_required()
def export_controller_excel():
    """Export Controller report to Excel - per tanggal per mesin per work order"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from models.production import ShiftProduction
        from collections import Counter
        
        machine_id = request.args.get('machine_id', type=int)
        period = request.args.get('period', 'day')  # 'day', 'week' or 'month'
        selected_date = request.args.get('date')  # Optional: specific date for 'day' period
        
        # Calculate date range based on period
        if period == 'day':
            if selected_date:
                # Use the selected date
                start_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
                end_date = start_date
            else:
                # Find the latest production date for this machine
                latest_query = ShiftProduction.query
                if machine_id:
                    latest_query = latest_query.filter(ShiftProduction.machine_id == machine_id)
                latest_record = latest_query.order_by(ShiftProduction.production_date.desc()).first()
                
                if latest_record:
                    start_date = latest_record.production_date
                    end_date = latest_record.production_date
                else:
                    start_date = get_local_now().date()
                    end_date = get_local_now().date()
        else:
            end_date = get_local_now().date()
            if period == 'month':
                start_date = end_date - timedelta(days=30)
            else:  # week
                start_date = end_date - timedelta(days=7)
        
        # Get shift production records within date range
        query = ShiftProduction.query.filter(
            ShiftProduction.production_date >= start_date,
            ShiftProduction.production_date <= end_date
        )
        if machine_id:
            query = query.filter(ShiftProduction.machine_id == machine_id)
        
        records = query.order_by(
            ShiftProduction.production_date.desc(),
            ShiftProduction.machine_id,
            ShiftProduction.work_order_id
        ).all()
        
        # Keywords for human-related downtime (excluded from top 3)
        HUMAN_DOWNTIME_KEYWORDS = [
            'istirahat', 'makan', 'sholat', 'toilet', 'wc', 'break',
            'pulang', 'datang', 'terlambat', 'absen', 'cuti', 'sakit',
            'meeting', 'rapat', 'briefing', 'training', 'pelatihan',
            'sanitasi', 'persiapan & sanitasi', 'sterilisasi',
            'setting packaging', 'setting kemasan', 'setting mesin', 'setting mc'
        ]
        
        def is_human_downtime(reason):
            """Check if downtime reason is human-related"""
            reason_lower = reason.lower()
            return any(keyword in reason_lower for keyword in HUMAN_DOWNTIME_KEYWORDS)
        
        # Collect all issues across all records to find most frequent
        all_downtime_issues = []
        for sp in records:
            if sp.issues:
                parts = [p.strip() for p in sp.issues.split(';') if p.strip()]
                for part in parts:
                    match = re.match(r'(\d+)\s*menit\s*-\s*(.+?)(?:\s*\[.+\])?$', part, re.IGNORECASE)
                    if match:
                        reason = match.group(2).strip()
                        reason = re.sub(r'\s*\[.+\]\s*$', '', reason).strip()
                        duration = int(match.group(1))
                        all_downtime_issues.append({
                            'reason': reason,
                            'duration': duration,
                            'is_human': is_human_downtime(reason)
                        })
        
        # Count frequency of each issue (excluding human downtime for top ranking)
        non_human_issues = [i['reason'] for i in all_downtime_issues if not i['is_human']]
        issue_frequency = Counter(non_human_issues)
        
        # Get top 3 most frequent non-human issues
        top_3_issues = [issue for issue, count in issue_frequency.most_common(3)]
        print(f"[DEBUG] Top 3 non-human issues: {top_3_issues}")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Controller Report"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Styles for downtime
        red_font = Font(color="FF0000", bold=True)  # Red for top 3 issues
        black_font = Font(color="000000")  # Black for other issues
        
        # Headers - single Issue column, issues go down as rows
        headers = [
            "Tanggal", "Mesin", "Shift", "Produk", 
            "Target (pcs)", "Aktual (pcs)", "Target (karton)", "Aktual (karton)",
            "Grade A", "Grade B", "Grade C",
            "Availability (%)", "Performance (%)", "Quality (%)", "OEE (%)",
            "Downtime (menit)", "Issue"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        row_num = 2
        for sp in records:
            # Get product name
            product_name = None
            if sp.product:
                product_name = sp.product.name
            elif sp.work_order and sp.work_order.product:
                product_name = sp.work_order.product.name
            
            # Get pack_per_karton
            pack_per_karton = 50
            if product_name:
                try:
                    search_name = product_name.replace('WIP ', '').strip()
                    product_new = ProductNew.query.filter(
                        ProductNew.name.ilike(f'%{search_name}%')
                    ).first()
                    if product_new and product_new.pack_per_karton:
                        pack_per_karton = int(product_new.pack_per_karton)
                except:
                    pass
            
            # Get shift number
            shift_num = 1
            if sp.shift:
                shift_match = re.search(r'(\d+)', str(sp.shift))
                if shift_match:
                    shift_num = int(shift_match.group(1))
            
            # Parse all issues from issues field
            record_issues = []
            if sp.issues:
                parts = [p.strip() for p in sp.issues.split(';') if p.strip()]
                for part in parts:
                    # Parse format: "XX menit - reason [category]"
                    match = re.match(r'(\d+)\s*menit\s*-\s*(.+?)(?:\s*\[.+\])?$', part, re.IGNORECASE)
                    if match:
                        duration = int(match.group(1))
                        reason = match.group(2).strip()
                        reason = re.sub(r'\s*\[.+\]\s*$', '', reason).strip()
                        is_top3 = reason in top_3_issues  # Check if in global top 3
                        record_issues.append({
                            'duration': duration, 
                            'reason': reason,
                            'is_top3': is_top3,
                            'is_human': is_human_downtime(reason)
                        })
                    else:
                        # Fallback: just use the text
                        record_issues.append({
                            'duration': 0, 
                            'reason': part,
                            'is_top3': False,
                            'is_human': is_human_downtime(part)
                        })
                
                # Sort: top 3 issues first (by duration), then others (by duration)
                record_issues.sort(key=lambda x: (not x['is_top3'], -x['duration']))
            
            target_qty = int(sp.target_quantity) if sp.target_quantity else 0
            actual_qty = int(sp.actual_quantity) if sp.actual_quantity else 0
            
            # Base row data (without issues)
            base_row_data = [
                sp.production_date.strftime('%Y-%m-%d') if sp.production_date else '',
                sp.machine.name if sp.machine else '',
                f"Shift {shift_num}",
                product_name or '',
                target_qty,
                actual_qty,
                round(target_qty / pack_per_karton) if pack_per_karton else 0,
                round(actual_qty / pack_per_karton) if pack_per_karton else 0,
                int(sp.good_quantity) if sp.good_quantity else 0,
                int(sp.rework_quantity) if sp.rework_quantity else 0,
                int(sp.reject_quantity) if sp.reject_quantity else 0,
                round(float(sp.efficiency_rate), 1) if sp.efficiency_rate else 0,
                100.0,  # Default performance
                round(float(sp.quality_rate), 1) if sp.quality_rate else 0,
                round(float(sp.oee_score), 1) if sp.oee_score else 0,
                int(sp.downtime_minutes) if sp.downtime_minutes else 0
            ]
            
            # If no issues, write single row with empty issue
            if not record_issues:
                for col, value in enumerate(base_row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = thin_border
                    if col >= 5:
                        cell.alignment = Alignment(horizontal="right")
                # Empty issue cell
                cell = ws.cell(row=row_num, column=17, value="")
                cell.border = thin_border
                row_num += 1
            else:
                # Write multiple rows - one per issue
                # First row has all data, subsequent rows only have issue
                for i, issue in enumerate(record_issues):
                    if i == 0:
                        # First row: write all base data + first issue
                        for col, value in enumerate(base_row_data, 1):
                            cell = ws.cell(row=row_num, column=col, value=value)
                            cell.border = thin_border
                            if col >= 5:
                                cell.alignment = Alignment(horizontal="right")
                    else:
                        # Subsequent rows: empty cells for base data (or merge later)
                        for col in range(1, 17):
                            cell = ws.cell(row=row_num, column=col, value="")
                            cell.border = thin_border
                    
                    # Write issue in column 17
                    issue_text = f"{issue['duration']} menit - {issue['reason']}"
                    cell = ws.cell(row=row_num, column=17, value=issue_text)
                    cell.border = thin_border
                    
                    # Top 3 global issues (non-human) in red, others in black
                    # Human downtime never gets red even if frequent
                    if issue['is_top3'] and not issue['is_human']:
                        cell.font = red_font
                    else:
                        cell.font = black_font
                    
                    row_num += 1
        
        # Set fixed column widths for better readability
        column_widths = {
            1: 12,   # Tanggal
            2: 15,   # Mesin
            3: 8,    # Shift
            4: 25,   # Produk
            5: 12,   # Target (pcs)
            6: 12,   # Aktual (pcs)
            7: 10,   # Target (karton)
            8: 10,   # Aktual (karton)
            9: 10,   # Grade A
            10: 10,  # Grade B
            11: 10,  # Grade C
            12: 12,  # Availability
            13: 12,  # Performance
            14: 10,  # Quality
            15: 8,   # OEE
            16: 12,  # Downtime
            17: 40   # Issue
        }
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"controller_report_{get_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/records', methods=['POST'])
@jwt_required()
def create_record():
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        
        record_number = generate_number('OEE', OEERecord, 'record_number')
        
        # Calculate OEE metrics
        availability = ((data['planned_production_time'] - data['downtime']) / data['planned_production_time']) * 100
        performance = ((data['total_pieces_produced'] * data['ideal_cycle_time']) / data['actual_production_time']) * 100 if data['actual_production_time'] > 0 else 0
        quality = (data['good_pieces'] / data['total_pieces_produced']) * 100 if data['total_pieces_produced'] > 0 else 0
        oee = (availability * performance * quality) / 10000
        
        record = OEERecord(
            record_number=record_number,
            machine_id=data['machine_id'],
            work_order_id=data.get('work_order_id'),
            record_date=datetime.fromisoformat(data['record_date']),
            shift=data.get('shift'),
            planned_production_time=data['planned_production_time'],
            downtime=data['downtime'],
            actual_production_time=data['actual_production_time'],
            ideal_cycle_time=data['ideal_cycle_time'],
            total_pieces_produced=data['total_pieces_produced'],
            good_pieces=data['good_pieces'],
            rejected_pieces=data['rejected_pieces'],
            availability=availability,
            performance=performance,
            quality=quality,
            oee=oee,
            recorded_by=user_id
        )
        
        db.session.add(record)
        db.session.commit()
        return jsonify({'message': 'OEE record created', 'record_id': record.id, 'oee': oee}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/downtime', methods=['GET'])
@jwt_required()
def get_downtime():
    """Get downtime records from both OEEDowntimeRecord and ShiftProduction"""
    try:
        from models.production import ShiftProduction
        
        machine_id = request.args.get('machine_id', type=int)
        limit = request.args.get('limit', 50, type=int)
        
        all_records = []
        
        # Get from OEEDowntimeRecord
        query = OEEDowntimeRecord.query
        if machine_id:
            query = query.filter(OEEDowntimeRecord.machine_id == machine_id)
        
        oee_downtimes = query.order_by(OEEDowntimeRecord.start_time.desc()).limit(limit).all()
        for r in oee_downtimes:
            all_records.append({
                'id': r.id,
                'source': 'oee_downtime',
                'machine_id': r.machine_id,
                'reason': r.reason or r.downtime_category or 'Unknown',
                'duration_minutes': r.duration_minutes or 0,
                'start_time': r.start_time.isoformat() if r.start_time else None,
                'end_time': r.end_time.isoformat() if r.end_time else None,
                'downtime_category': r.downtime_category
            })
        
        # Get from ShiftProduction - create downtime entries per category
        shift_query = ShiftProduction.query
        if machine_id:
            shift_query = shift_query.filter(ShiftProduction.machine_id == machine_id)
        
        shift_records = shift_query.order_by(ShiftProduction.production_date.desc()).limit(limit).all()
        for sp in shift_records:
            base_time = datetime.combine(sp.production_date, sp.shift_start) if sp.production_date and sp.shift_start else get_local_now()
            
            # Add downtime entries for each category that has minutes
            if sp.downtime_mesin and sp.downtime_mesin > 0:
                all_records.append({
                    'id': f"sp-{sp.id}-mesin",
                    'source': 'shift_production',
                    'machine_id': sp.machine_id,
                    'reason': 'Downtime Mesin',
                    'duration_minutes': sp.downtime_mesin,
                    'start_time': base_time.isoformat(),
                    'end_time': None,
                    'downtime_category': 'mesin'
                })
            
            if sp.downtime_operator and sp.downtime_operator > 0:
                all_records.append({
                    'id': f"sp-{sp.id}-operator",
                    'source': 'shift_production',
                    'machine_id': sp.machine_id,
                    'reason': 'Downtime Operator',
                    'duration_minutes': sp.downtime_operator,
                    'start_time': base_time.isoformat(),
                    'end_time': None,
                    'downtime_category': 'operator'
                })
            
            if sp.downtime_material and sp.downtime_material > 0:
                all_records.append({
                    'id': f"sp-{sp.id}-material",
                    'source': 'shift_production',
                    'machine_id': sp.machine_id,
                    'reason': 'Downtime Material',
                    'duration_minutes': sp.downtime_material,
                    'start_time': base_time.isoformat(),
                    'end_time': None,
                    'downtime_category': 'material'
                })
            
            if sp.downtime_design and sp.downtime_design > 0:
                all_records.append({
                    'id': f"sp-{sp.id}-design",
                    'source': 'shift_production',
                    'machine_id': sp.machine_id,
                    'reason': 'Downtime Design Change',
                    'duration_minutes': sp.downtime_design,
                    'start_time': base_time.isoformat(),
                    'end_time': None,
                    'downtime_category': 'design'
                })
            
            if sp.downtime_others and sp.downtime_others > 0:
                all_records.append({
                    'id': f"sp-{sp.id}-others",
                    'source': 'shift_production',
                    'machine_id': sp.machine_id,
                    'reason': sp.issues or 'Downtime Lainnya',
                    'duration_minutes': sp.downtime_others,
                    'start_time': base_time.isoformat(),
                    'end_time': None,
                    'downtime_category': 'others'
                })
        
        # Sort by start_time and limit
        all_records.sort(key=lambda x: x['start_time'] or '', reverse=True)
        all_records = all_records[:limit]
        
        return jsonify({
            'records': all_records
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/downtime', methods=['POST'])
@jwt_required()
def create_downtime():
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        
        downtime = OEEDowntimeRecord(
            oee_record_id=data.get('oee_record_id'),
            machine_id=data['machine_id'],
            start_time=datetime.fromisoformat(data['start_time']),
            end_time=datetime.fromisoformat(data['end_time']) if data.get('end_time') else None,
            duration_minutes=data.get('duration_minutes'),
            downtime_category=data['downtime_category'],
            reason=data.get('reason'),
            recorded_by=user_id
        )
        
        db.session.add(downtime)
        db.session.commit()
        return jsonify(success_response('api.success')), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/shift-production/<int:id>/downtime', methods=['PUT'])
@jwt_required()
def update_shift_production_downtime(id):
    """Update downtime breakdown for a ShiftProduction record"""
    try:
        from models.production import ShiftProduction
        
        sp = ShiftProduction.query.get(id)
        if not sp:
            return jsonify({'error': 'ShiftProduction not found'}), 404
        
        data = request.get_json()
        
        # Update downtime by category
        if 'downtime_mesin' in data:
            sp.downtime_mesin = int(data['downtime_mesin'])
        if 'downtime_operator' in data:
            sp.downtime_operator = int(data['downtime_operator'])
        if 'downtime_material' in data:
            sp.downtime_material = int(data['downtime_material'])
        if 'downtime_design' in data:
            sp.downtime_design = int(data['downtime_design'])
        if 'downtime_others' in data:
            sp.downtime_others = int(data['downtime_others'])
        
        # Recalculate total downtime
        sp.downtime_minutes = (sp.downtime_mesin or 0) + (sp.downtime_operator or 0) + \
                              (sp.downtime_material or 0) + (sp.downtime_design or 0) + \
                              (sp.downtime_others or 0)
        
        # Recalculate loss percentages
        planned_runtime = sp.planned_runtime or 510
        sp.loss_mesin = round((sp.downtime_mesin or 0) / planned_runtime * 100, 2) if planned_runtime > 0 else 0
        sp.loss_operator = round((sp.downtime_operator or 0) / planned_runtime * 100, 2) if planned_runtime > 0 else 0
        sp.loss_material = round((sp.downtime_material or 0) / planned_runtime * 100, 2) if planned_runtime > 0 else 0
        sp.loss_design = round((sp.downtime_design or 0) / planned_runtime * 100, 2) if planned_runtime > 0 else 0
        sp.loss_others = round((sp.downtime_others or 0) / planned_runtime * 100, 2) if planned_runtime > 0 else 0
        
        # Recalculate efficiency and OEE
        sp.actual_runtime = planned_runtime - sp.downtime_minutes
        sp.efficiency_rate = round((sp.actual_runtime / planned_runtime * 100) if planned_runtime > 0 else 100, 2)
        sp.oee_score = round((sp.efficiency_rate * float(sp.quality_rate or 100)) / 100, 2)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Downtime updated successfully',
            'data': {
                'id': sp.id,
                'downtime_mesin': sp.downtime_mesin,
                'downtime_operator': sp.downtime_operator,
                'downtime_material': sp.downtime_material,
                'downtime_design': sp.downtime_design,
                'downtime_others': sp.downtime_others,
                'downtime_minutes': sp.downtime_minutes,
                'efficiency_rate': float(sp.efficiency_rate),
                'oee_score': float(sp.oee_score)
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/shift-production', methods=['GET'])
@jwt_required()
def get_shift_productions():
    """Get ShiftProduction records for editing"""
    try:
        from models.production import ShiftProduction
        
        machine_id = request.args.get('machine_id', type=int)
        limit = request.args.get('limit', 50, type=int)
        
        query = ShiftProduction.query
        if machine_id:
            query = query.filter(ShiftProduction.machine_id == machine_id)
        
        records = query.order_by(ShiftProduction.production_date.desc()).limit(limit).all()
        
        return jsonify({
            'records': [{
                'id': sp.id,
                'production_date': sp.production_date.isoformat() if sp.production_date else None,
                'shift': sp.shift,
                'machine_id': sp.machine_id,
                'machine_name': sp.machine.name if sp.machine else None,
                'product_name': sp.product.name if sp.product else None,
                'actual_quantity': float(sp.actual_quantity) if sp.actual_quantity else 0,
                'downtime_minutes': sp.downtime_minutes or 0,
                'downtime_mesin': sp.downtime_mesin or 0,
                'downtime_operator': sp.downtime_operator or 0,
                'downtime_material': sp.downtime_material or 0,
                'downtime_design': sp.downtime_design or 0,
                'downtime_others': sp.downtime_others or 0,
                'efficiency_rate': float(sp.efficiency_rate) if sp.efficiency_rate else 0,
                'quality_rate': float(sp.quality_rate) if sp.quality_rate else 0,
                'oee_score': float(sp.oee_score) if sp.oee_score else 0,
                'issues': sp.issues
            } for sp in records]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===============================
# ENHANCED OEE ENDPOINTS
# ===============================

@oee_bp.route('/dashboard', methods=['GET'])
@jwt_required()
def get_oee_dashboard():
    """Get comprehensive OEE dashboard data from both OEERecord and ShiftProduction"""
    try:
        from models.production import ShiftProduction
        
        # Get query parameters
        machine_id_param = request.args.get('machine_id')
        machine_id = None
        if machine_id_param and machine_id_param != 'null' and machine_id_param != '':
            try:
                machine_id = int(machine_id_param)
            except (ValueError, TypeError):
                machine_id = None
        
        days = request.args.get('days', 30, type=int)
        
        # Date range
        end_date = get_local_today()
        start_date = end_date - timedelta(days=days)
        
        # Get OEERecord data
        oee_query = OEERecord.query.filter(
            OEERecord.record_date.between(start_date, end_date)
        )
        if machine_id:
            oee_query = oee_query.filter(OEERecord.machine_id == machine_id)
        oee_records = oee_query.all()
        
        # Get ShiftProduction data (from WorkOrder production input)
        shift_query = ShiftProduction.query.filter(
            ShiftProduction.production_date.between(start_date, end_date)
        )
        if machine_id:
            shift_query = shift_query.filter(ShiftProduction.machine_id == machine_id)
        shift_records = shift_query.all()
        
        # Combine records for calculations
        records = oee_records  # Keep original for compatibility
        
        # Calculate overall metrics combining both sources
        all_oee_values = []
        all_availability_values = []
        all_performance_values = []
        all_quality_values = []
        
        # From OEERecord
        for r in oee_records:
            all_oee_values.append(float(r.oee) if r.oee else 0)
            all_availability_values.append(float(r.availability) if r.availability else 0)
            all_performance_values.append(float(r.performance) if r.performance else 0)
            all_quality_values.append(float(r.quality) if r.quality else 0)
        
        # From ShiftProduction - group by unique machine+date+shift to avoid multi-product over-counting
        sp_by_shift = {}
        for sp in shift_records:
            shift_key = f"{sp.machine_id}_{sp.production_date}_{sp.shift}"
            if shift_key not in sp_by_shift:
                sp_by_shift[shift_key] = []
            sp_by_shift[shift_key].append(sp)
        
        for shift_key, sp_group in sp_by_shift.items():
            # Average OEE values within the same shift (multi-product)
            oee_vals = [float(sp.oee_score) for sp in sp_group if sp.oee_score]
            eff_vals = [float(sp.efficiency_rate) for sp in sp_group if sp.efficiency_rate]
            qual_vals = [float(sp.quality_rate) for sp in sp_group if sp.quality_rate]
            all_oee_values.append(sum(oee_vals) / len(oee_vals) if oee_vals else 0)
            all_availability_values.append(sum(eff_vals) / len(eff_vals) if eff_vals else 0)
            all_performance_values.append(100.0)  # Default performance
            all_quality_values.append(sum(qual_vals) / len(qual_vals) if qual_vals else 0)
        
        if all_oee_values:
            avg_oee = sum(all_oee_values) / len(all_oee_values)
            avg_availability = sum(all_availability_values) / len(all_availability_values)
            avg_performance = sum(all_performance_values) / len(all_performance_values)
            avg_quality = sum(all_quality_values) / len(all_quality_values)
            best_oee = max(all_oee_values)
            worst_oee = min(all_oee_values)
        else:
            avg_oee = avg_availability = avg_performance = avg_quality = 0
            best_oee = worst_oee = 0
        
        # Get machine performance data
        machines_query = Machine.query.filter(Machine.is_active == True)
        if machine_id:
            machines_query = machines_query.filter(Machine.id == machine_id)
        
        machines = machines_query.all()
        machine_performance = []
        
        for machine in machines:
            # From OEERecord
            machine_oee_records = [r for r in oee_records if r.machine_id == machine.id]
            # From ShiftProduction - group by unique shift
            machine_shift_records = [sp for sp in shift_records if sp.machine_id == machine.id]
            
            # Combine OEE values
            machine_oee_values = []
            machine_downtime = 0
            machine_production = 0
            
            for r in machine_oee_records:
                machine_oee_values.append(float(r.oee) if r.oee else 0)
                machine_downtime += r.downtime or 0
                machine_production += r.total_pieces_produced or 0
            
            # Group SP by unique shift to avoid multi-product over-counting OEE
            m_sp_by_shift = {}
            for sp in machine_shift_records:
                sk = f"{sp.production_date}_{sp.shift}"
                if sk not in m_sp_by_shift:
                    m_sp_by_shift[sk] = []
                m_sp_by_shift[sk].append(sp)
                # Production and downtime: sum across all records (each record is unique output)
                machine_downtime += sp.downtime_minutes or 0
                machine_production += float(sp.actual_quantity) if sp.actual_quantity else 0
            
            for sk, sp_grp in m_sp_by_shift.items():
                oee_vals = [float(sp.oee_score) for sp in sp_grp if sp.oee_score]
                machine_oee_values.append(sum(oee_vals) / len(oee_vals) if oee_vals else 0)
            
            if machine_oee_values:
                machine_avg_oee = sum(machine_oee_values) / len(machine_oee_values)
            else:
                machine_avg_oee = 0
            
            # Get maintenance info
            next_maintenance = MaintenanceSchedule.query.filter(
                MaintenanceSchedule.machine_id == machine.id,
                MaintenanceSchedule.is_active == True,
                MaintenanceSchedule.next_maintenance_date >= get_local_today()
            ).order_by(MaintenanceSchedule.next_maintenance_date).first()
            
            # Get recent alerts
            recent_alerts = OEEAlert.query.filter(
                OEEAlert.machine_id == machine.id,
                OEEAlert.status == 'active'
            ).count()
            
            machine_performance.append({
                'machine_id': machine.id,
                'machine_name': machine.name,
                'machine_code': machine.code,
                'status': machine.status,
                'avg_oee': round(machine_avg_oee, 2),
                'total_downtime': machine_downtime,
                'total_production': machine_production,
                'next_maintenance': next_maintenance.next_maintenance_date.isoformat() if next_maintenance else None,
                'active_alerts': recent_alerts,
                'efficiency': float(machine.efficiency) if machine.efficiency else 100,
                'availability': float(machine.availability) if machine.availability else 100
            })
        
        # Get trend data (last 7 days) - combining both sources
        trend_data = []
        for i in range(7):
            trend_date = end_date - timedelta(days=i)
            day_oee_values = []
            
            # From OEERecord
            for r in oee_records:
                if r.record_date == trend_date:
                    day_oee_values.append(float(r.oee) if r.oee else 0)
            
            # From ShiftProduction - group by unique shift per day
            day_sp_by_shift = {}
            for sp in shift_records:
                if sp.production_date == trend_date:
                    sk = f"{sp.machine_id}_{sp.shift}"
                    if sk not in day_sp_by_shift:
                        day_sp_by_shift[sk] = []
                    day_sp_by_shift[sk].append(sp)
            for sk, sp_grp in day_sp_by_shift.items():
                oee_vals = [float(sp.oee_score) for sp in sp_grp if sp.oee_score]
                day_oee_values.append(sum(oee_vals) / len(oee_vals) if oee_vals else 0)
            
            if day_oee_values:
                day_avg_oee = sum(day_oee_values) / len(day_oee_values)
            else:
                day_avg_oee = 0
            
            trend_data.append({
                'date': trend_date.isoformat(),
                'oee': round(day_avg_oee, 2)
            })
        
        trend_data.reverse()  # Show oldest to newest
        
        # Get active alerts
        alerts_query = OEEAlert.query.filter(OEEAlert.status == 'active')
        if machine_id:
            alerts_query = alerts_query.filter(OEEAlert.machine_id == machine_id)
        
        active_alerts = alerts_query.order_by(desc(OEEAlert.alert_date)).limit(10).all()
        
        # Get downtime analysis - combining both sources
        downtime_by_category = {}
        
        # From OEEDowntimeRecord
        downtime_records = OEEDowntimeRecord.query.join(OEERecord).filter(
            OEERecord.record_date.between(start_date, end_date)
        )
        if machine_id:
            downtime_records = downtime_records.filter(OEEDowntimeRecord.machine_id == machine_id)
        
        for downtime in downtime_records.all():
            category = downtime.downtime_category or 'others'
            if category not in downtime_by_category:
                downtime_by_category[category] = 0
            downtime_by_category[category] += downtime.duration_minutes or 0
        
        # From ShiftProduction - add downtime by category
        for sp in shift_records:
            if sp.downtime_mesin and sp.downtime_mesin > 0:
                downtime_by_category['mesin'] = downtime_by_category.get('mesin', 0) + sp.downtime_mesin
            if sp.downtime_operator and sp.downtime_operator > 0:
                downtime_by_category['operator'] = downtime_by_category.get('operator', 0) + sp.downtime_operator
            if sp.downtime_material and sp.downtime_material > 0:
                downtime_by_category['material'] = downtime_by_category.get('material', 0) + sp.downtime_material
            if sp.downtime_design and sp.downtime_design > 0:
                downtime_by_category['design'] = downtime_by_category.get('design', 0) + sp.downtime_design
            if sp.downtime_others and sp.downtime_others > 0:
                downtime_by_category['others'] = downtime_by_category.get('others', 0) + sp.downtime_others
        
        return jsonify({
            'summary': {
                'avg_oee': round(avg_oee, 2),
                'avg_availability': round(avg_availability, 2),
                'avg_performance': round(avg_performance, 2),
                'avg_quality': round(avg_quality, 2),
                'best_oee': round(best_oee, 2),
                'worst_oee': round(worst_oee, 2),
                'total_records': len(oee_records) + len(sp_by_shift),
                'total_production_days': len(set(
                    [r.record_date for r in oee_records if r.record_date] +
                    [sp.production_date for sp in shift_records if sp.production_date]
                )),
                'total_machine_days': len(set(
                    [(r.machine_id, r.record_date) for r in oee_records if r.record_date] +
                    [(sp.machine_id, sp.production_date) for sp in shift_records if sp.production_date]
                )),
                'date_range': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                }
            },
            'machine_performance': machine_performance,
            'trend_data': trend_data,
            'active_alerts': [{
                'id': alert.id,
                'machine_name': alert.machine.name,
                'alert_type': alert.alert_type,
                'severity': alert.severity,
                'title': alert.title,
                'message': alert.message,
                'alert_date': alert.alert_date.isoformat(),
                'threshold_value': float(alert.threshold_value) if alert.threshold_value else None,
                'actual_value': float(alert.actual_value) if alert.actual_value else None
            } for alert in active_alerts],
            'downtime_analysis': [
                {'category': category, 'minutes': minutes}
                for category, minutes in downtime_by_category.items()
            ]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/alerts', methods=['GET'])
@jwt_required()
def get_alerts():
    """Get OEE alerts"""
    try:
        status = request.args.get('status', 'active')
        
        # Handle machine_id parameter properly
        machine_id_param = request.args.get('machine_id')
        machine_id = None
        if machine_id_param and machine_id_param != 'null' and machine_id_param != '':
            try:
                machine_id = int(machine_id_param)
            except (ValueError, TypeError):
                machine_id = None
        
        severity = request.args.get('severity')
        
        query = OEEAlert.query
        
        if status:
            query = query.filter(OEEAlert.status == status)
        if machine_id:
            query = query.filter(OEEAlert.machine_id == machine_id)
        if severity:
            query = query.filter(OEEAlert.severity == severity)
        
        alerts = query.order_by(desc(OEEAlert.alert_date)).all()
        
        return jsonify({
            'alerts': [{
                'id': alert.id,
                'machine_id': alert.machine_id,
                'machine_name': alert.machine.name,
                'alert_type': alert.alert_type,
                'severity': alert.severity,
                'title': alert.title,
                'message': alert.message,
                'threshold_value': float(alert.threshold_value) if alert.threshold_value else None,
                'actual_value': float(alert.actual_value) if alert.actual_value else None,
                'alert_date': alert.alert_date.isoformat(),
                'status': alert.status,
                'acknowledged_by': alert.acknowledged_by_user.username if alert.acknowledged_by_user else None,
                'acknowledged_at': alert.acknowledged_at.isoformat() if alert.acknowledged_at else None,
                'resolved_by': alert.resolved_by_user.username if alert.resolved_by_user else None,
                'resolved_at': alert.resolved_at.isoformat() if alert.resolved_at else None
            } for alert in alerts]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/alerts/<int:alert_id>/acknowledge', methods=['PUT'])
@jwt_required()
def acknowledge_alert(alert_id):
    """Acknowledge an OEE alert"""
    try:
        user_id = get_jwt_identity()
        alert = OEEAlert.query.get_or_404(alert_id)
        
        alert.status = 'acknowledged'
        alert.acknowledged_by = user_id
        alert.acknowledged_at = get_local_now()
        
        db.session.commit()
        
        return jsonify(success_response('api.success')), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/alerts/<int:alert_id>/resolve', methods=['PUT'])
@jwt_required()
def resolve_alert(alert_id):
    """Resolve an OEE alert"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json()
        alert = OEEAlert.query.get_or_404(alert_id)
        
        alert.status = 'resolved'
        alert.resolved_by = user_id
        alert.resolved_at = get_local_now()
        alert.resolution_notes = data.get('resolution_notes', '')
        
        db.session.commit()
        
        return jsonify(success_response('api.success')), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/maintenance-impact', methods=['POST'])
@jwt_required()
def create_maintenance_impact():
    """Create maintenance impact record"""
    try:
        data = request.get_json()
        
        impact = MaintenanceImpact(
            maintenance_record_id=data['maintenance_record_id'],
            machine_id=data['machine_id'],
            impact_date=datetime.strptime(data['impact_date'], '%Y-%m-%d').date(),
            planned_downtime_hours=data.get('planned_downtime_hours', 0),
            actual_downtime_hours=data.get('actual_downtime_hours', 0),
            production_loss_units=data.get('production_loss_units', 0),
            revenue_impact=data.get('revenue_impact', 0),
            oee_before_maintenance=data.get('oee_before_maintenance'),
            oee_after_maintenance=data.get('oee_after_maintenance'),
            notes=data.get('notes')
        )
        
        # Calculate improvement percentage
        if impact.oee_before_maintenance and impact.oee_after_maintenance:
            impact.improvement_percentage = float(impact.oee_after_maintenance) - float(impact.oee_before_maintenance)
        
        db.session.add(impact)
        db.session.commit()
        
        return jsonify({
            'message': 'Maintenance impact recorded successfully',
            'impact_id': impact.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@oee_bp.route('/machines/<int:machine_id>/analytics', methods=['GET'])
@jwt_required()
def get_machine_analytics(machine_id):
    """Get detailed analytics for a specific machine"""
    try:
        machine = Machine.query.get_or_404(machine_id)
        
        # Get query parameters
        period = request.args.get('period', 'monthly')  # daily, weekly, monthly
        months = request.args.get('months', 6, type=int)
        
        # Get analytics data
        analytics = OEEAnalytics.query.filter(
            OEEAnalytics.machine_id == machine_id,
            OEEAnalytics.period_type == period
        ).order_by(desc(OEEAnalytics.analysis_date)).limit(months).all()
        
        # Get maintenance impact data
        maintenance_impacts = MaintenanceImpact.query.filter(
            MaintenanceImpact.machine_id == machine_id
        ).order_by(desc(MaintenanceImpact.impact_date)).limit(10).all()
        
        # Get recent OEE records for detailed view
        recent_records = OEERecord.query.filter(
            OEERecord.machine_id == machine_id
        ).order_by(desc(OEERecord.record_date)).limit(30).all()
        
        # Calculate trends
        if len(analytics) >= 2:
            latest = analytics[0]
            previous = analytics[1]
            oee_trend = float(latest.avg_oee) - float(previous.avg_oee)
            availability_trend = float(latest.avg_availability) - float(previous.avg_availability)
            performance_trend = float(latest.avg_performance) - float(previous.avg_performance)
            quality_trend = float(latest.avg_quality) - float(previous.avg_quality)
        else:
            oee_trend = availability_trend = performance_trend = quality_trend = 0
        
        return jsonify({
            'machine': {
                'id': machine.id,
                'name': machine.name,
                'code': machine.code,
                'type': machine.machine_type,
                'status': machine.status,
                'capacity_per_hour': float(machine.capacity_per_hour) if machine.capacity_per_hour else None,
                'last_maintenance': machine.last_maintenance.isoformat() if machine.last_maintenance else None,
                'next_maintenance': machine.next_maintenance.isoformat() if machine.next_maintenance else None
            },
            'trends': {
                'oee_trend': round(oee_trend, 2),
                'availability_trend': round(availability_trend, 2),
                'performance_trend': round(performance_trend, 2),
                'quality_trend': round(quality_trend, 2)
            },
            'analytics': [{
                'date': a.analysis_date.isoformat(),
                'period_type': a.period_type,
                'avg_oee': float(a.avg_oee),
                'avg_availability': float(a.avg_availability),
                'avg_performance': float(a.avg_performance),
                'avg_quality': float(a.avg_quality),
                'total_downtime_hours': float(a.total_downtime_hours),
                'total_production_hours': float(a.total_production_hours),
                'total_units_produced': float(a.total_units_produced),
                'defect_rate': float(a.defect_rate),
                'maintenance_hours': float(a.maintenance_hours),
                'breakdown_count': a.breakdown_count
            } for a in analytics],
            'maintenance_impacts': [{
                'date': mi.impact_date.isoformat(),
                'planned_downtime': float(mi.planned_downtime_hours),
                'actual_downtime': float(mi.actual_downtime_hours),
                'production_loss': float(mi.production_loss_units),
                'revenue_impact': float(mi.revenue_impact),
                'oee_before': float(mi.oee_before_maintenance) if mi.oee_before_maintenance else None,
                'oee_after': float(mi.oee_after_maintenance) if mi.oee_after_maintenance else None,
                'improvement': float(mi.improvement_percentage) if mi.improvement_percentage else None
            } for mi in maintenance_impacts],
            'recent_records': [{
                'date': r.record_date.isoformat(),
                'shift': r.shift,
                'oee': float(r.oee),
                'availability': float(r.availability),
                'performance': float(r.performance),
                'quality': float(r.quality),
                'downtime': r.downtime,
                'total_pieces': r.total_pieces_produced,
                'good_pieces': r.good_pieces
            } for r in recent_records]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/daily-controller', methods=['GET'])
@jwt_required()
def get_daily_controller():
    """Get all machines' production data for a specific date - Daily Controller view"""
    try:
        from models.production import ShiftProduction
        
        selected_date = request.args.get('date')
        if selected_date:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        else:
            target_date = get_local_today()
        
        # Get all shift productions for this date with eager loading
        from sqlalchemy.orm import joinedload
        shift_records = ShiftProduction.query.options(
            joinedload(ShiftProduction.work_order),
            joinedload(ShiftProduction.machine),
            joinedload(ShiftProduction.product)
        ).filter(
            ShiftProduction.production_date == target_date
        ).order_by(ShiftProduction.machine_id, ShiftProduction.shift).all()
        
        # Group by machine
        machines_data = {}
        
        # IDLE TIME keywords - sinkron dengan frontend
        idle_keywords = [
            # Tunggu kain
            'tunggu kain', 'ambil kain', 'menunggu kain', 'nunggu kain', 'kain belum datang',
            # Tunggu obat/tinta
            'tunggu obat', 'menunggu obat', 'nunggu obat', 'obat belum datang',
            'tunggu tinta', 'menunggu tinta', 'nunggu tinta',
            # Tunggu ingredient
            'tunggu ingredient', 'ingredient habis', 'tunggu bahan kimia',
            # Tunggu stiker
            'tunggu stiker', 'menunggu stiker', 'nunggu stiker', 'stiker belum datang',
            # Tunggu packing/packaging
            'tunggu packing', 'tunggu packaging', 'menunggu packing', 'nunggu packing',
            'packaging belum datang', 'box belum datang',
            # Tunggu mixing
            'tunggu mixing', 'menunggu mixing', 'nunggu mixing', 'mixing belum siap',
            # Tunggu label/karton/box
            'tunggu label', 'tunggu box', 'tunggu karton', 'tunggu lem',
            # Tunggu produk (dari mesin lain)
            'tunggu produk',
            # Tunggu temperatur
            'tunggu temperatur stabil', 'tunggu temperatur',
            'tunggu temperature stabil', 'tunggu temperature',
            # General tunggu
            'tunggu bahan', 'tunggu material', 'tunggu order', 'tunggu instruksi',
            'tunggu approval', 'tunggu qc', 'tunggu hasil qc',
            # xxx habis (idle karena kehabisan)
            'kain habis', 'stiker habis', 'packing habis', 'packaging habis',
            'label habis', 'karton habis', 'box habis', 'lem habis', 'tinta habis',
            'bahan habis', 'material habis',
            # English
            'waiting for', 'standby material', 'waiting material', 'no material',
            # Idle lainnya
            'idle', 'standby', 'menganggur', 'tidak ada order', 'no order', 'menhabiskan order',
            # Susun produk
            'susun produk'
        ]
        
        for sp in shift_records:
            machine_id = sp.machine_id
            
            if machine_id not in machines_data:
                machines_data[machine_id] = {
                    'machine_id': machine_id,
                    'machine_name': sp.machine.name if sp.machine else f'Machine {machine_id}',
                    'machine_code': sp.machine.code if sp.machine else None,
                    'target_efficiency': int(sp.machine.target_efficiency) if sp.machine and sp.machine.target_efficiency else 60,
                    'date': target_date.isoformat(),
                    'shifts': [],
                    'total_planned': 0,
                    'total_runtime': 0,
                    'total_downtime': 0,
                    'total_idle': 0,
                    'total_output': 0,
                    'total_grade_a': 0,
                    'total_grade_b': 0,
                    'total_grade_c': 0,
                    'total_machine_speed': 0,
                    'total_target': 0,
                    'total_average_time': 0,
                    'products': [],
                    'shifts_seen': {},
                    'idle_breakdown': [],
                    'top_3_downtime': [],
                    'per_shift_downtime': {},  # {shift_num: [downtime_items]}
                    'efficiency': 0,
                    'machine_efficiency': 0,
                    'quality': 0,
                    'average_time': 0,
                    'runtime': 0,
                    'waktu_tercatat': 0,
                    'waktu_tidak_tercatat': 0
                }
            
            # Parse downtime from issues - IDLE is SEPARATE from downtime
            downtime_breakdown = []
            shift_downtime = 0  # Non-idle downtime only
            shift_idle = 0  # Idle time separate
            
            if sp.issues:
                issue_parts = sp.issues.split(';')
                for part in issue_parts:
                    part = part.strip()
                    if not part:
                        continue
                    match = re.match(r'(\d+)\s*menit\s*-\s*(.+)', part, re.IGNORECASE)
                    if match:
                        duration = int(match.group(1))
                        reason_with_tag = match.group(2).strip()
                        
                        # Check for [idle] category tag BEFORE removing it
                        has_idle_tag = bool(re.search(r'\[idle\]', reason_with_tag, re.IGNORECASE))
                        
                        reason = re.sub(r'\s*\[\w+\]\s*$', '', reason_with_tag).strip()
                        reason_lower = reason.lower()
                        
                        # Check if idle: either has [idle] tag OR matches idle keywords
                        is_idle = has_idle_tag or any(kw in reason_lower for kw in idle_keywords)
                        if is_idle:
                            shift_idle += duration
                        else:
                            shift_downtime += duration
                        
                        # Track is_idle status for filtering later
                        downtime_breakdown.append({'reason': reason, 'duration_minutes': duration, 'is_idle': is_idle})
            
            # Fallback for downtime (only if parsing yielded 0)
            if shift_downtime == 0:
                if sp.downtime_minutes:
                    shift_downtime = int(sp.downtime_minutes)
                else:
                    shift_downtime = (
                        (int(sp.downtime_mesin) if sp.downtime_mesin else 0) +
                        (int(sp.downtime_operator) if sp.downtime_operator else 0) +
                        (int(sp.downtime_material) if sp.downtime_material else 0) +
                        (int(sp.downtime_design) if sp.downtime_design else 0) +
                        (int(sp.downtime_others) if sp.downtime_others else 0)
                    )
            
            # Use idle_time from database field if available (parsed idle is additional)
            # Idle time is DIFFERENT from Waktu Tidak Tercatat
            if shift_idle == 0 and sp.idle_time:
                shift_idle = int(sp.idle_time)
            
            # Extract shift number
            shift_num = 1
            if sp.shift:
                shift_match = re.search(r'(\d+)', str(sp.shift))
                if shift_match:
                    shift_num = int(shift_match.group(1))
            
            # Get planned runtime per shift (540 on Friday for Shift 1)
            # Friday (weekday() == 4) has longer shift due to Jumatan prayer
            is_friday = sp.production_date and sp.production_date.weekday() == 4
            if shift_num == 1:
                default_planned = 540 if is_friday else 510  # 06:30-15:00 = 8.5 hours (9 hours on Friday)
            elif shift_num == 2:
                default_planned = 510  # 15:00-23:00 = 8.5 hours
            else:
                default_planned = 450  # 23:00-06:30 = 7.5 hours
            
            planned_runtime = int(sp.planned_runtime) if sp.planned_runtime else default_planned
            
            # Runtime = Planned - Downtime - Idle (calculated)
            shift_runtime = max(0, planned_runtime - shift_downtime - shift_idle)
            
            # Get product name and packs_per_karton
            product_name = None
            product_code = None
            pack_per_carton = 0
            if sp.product:
                product_name = sp.product.name
                product_code = sp.product.code
            elif sp.work_order and sp.work_order.product:
                product_name = sp.work_order.product.name
                product_code = sp.work_order.product.code
            
            # PRIORITY 1: Use pack_per_carton from ShiftProduction record (per shift input)
            if hasattr(sp, 'pack_per_carton') and sp.pack_per_carton and int(sp.pack_per_carton) > 0:
                pack_per_carton = int(sp.pack_per_carton)
            
            # PRIORITY 2: Use pack_per_carton from Work Order
            if pack_per_carton == 0 and sp.work_order and hasattr(sp.work_order, 'pack_per_carton') and sp.work_order.pack_per_carton:
                pack_per_carton = int(sp.work_order.pack_per_carton)
            
            # PRIORITY 3: Use pack_per_carton from BOM
            if pack_per_carton == 0 and sp.work_order and sp.work_order.bom:
                bom_ppc = sp.work_order.bom.pack_per_carton
                if bom_ppc and bom_ppc > 1:
                    pack_per_carton = int(bom_ppc)
            
            # PRIORITY 3: Fallback to products_new table
            if pack_per_carton == 0 and product_code:
                product_new_data = db.session.execute(
                    db.text('SELECT pack_per_karton FROM products WHERE code = :code'),
                    {'code': product_code}
                ).fetchone()
                if product_new_data and product_new_data[0]:
                    pack_per_carton = int(product_new_data[0])
            
            # PRIORITY 4: Fallback by product name
            if pack_per_carton == 0 and product_name:
                search_name = product_name
                if search_name.upper().startswith('WIP '):
                    search_name = search_name[4:]
                product_new_data = db.session.execute(
                    db.text("SELECT pack_per_karton FROM products WHERE name LIKE :name ORDER BY CASE WHEN name LIKE '%@24' THEN 0 WHEN name LIKE '%@24%' THEN 1 ELSE 2 END, pack_per_karton DESC LIMIT 1"),
                    {'name': f'%{search_name}%'}
                ).fetchone()
                if product_new_data and product_new_data[0]:
                    pack_per_carton = int(product_new_data[0])
            
            grade_a = int(sp.good_quantity) if sp.good_quantity else 0
            import math
            shift_data = {
                'shift': shift_num,
                'sub_shift': sp.sub_shift,  # 'a', 'b', 'c' or None for legacy
                'shift_production_id': sp.id,  # For ordering legacy data
                'planned_runtime': planned_runtime,  # User input average time
                'runtime_minutes': shift_runtime,
                'downtime_minutes': shift_downtime,
                'idle_time_minutes': shift_idle,
                'waktu_tidak_tercatat': int(sp.waktu_tidak_tercatat) if hasattr(sp, 'waktu_tidak_tercatat') and sp.waktu_tidak_tercatat is not None else 0,
                'grade_a': grade_a,
                'grade_a_carton': math.floor(grade_a / pack_per_carton) if pack_per_carton > 0 else 0,
                'grade_b': int(sp.rework_quantity) if sp.rework_quantity else 0,
                'grade_c': int(sp.reject_quantity) if sp.reject_quantity else 0,
                'setting_sticker': int(sp.setting_sticker) if hasattr(sp, 'setting_sticker') and sp.setting_sticker else 0,
                'setting_packaging': int(sp.setting_packaging) if hasattr(sp, 'setting_packaging') and sp.setting_packaging else 0,
                'total': int(sp.actual_quantity) if sp.actual_quantity else 0,
                'product_name': product_name,
                'pack_per_carton': pack_per_carton,
                'wo_number': sp.work_order.wo_number if sp.work_order else None,
                # Early stop info - convert to bool explicitly for safety
                'early_stop': bool(sp.early_stop) if sp.early_stop else False,
                'early_stop_time': (sp.early_stop_time.strftime('%H:%M') if hasattr(sp.early_stop_time, 'strftime') else str(sp.early_stop_time)[:5]) if sp.early_stop_time else None,
                'early_stop_reason': sp.early_stop_reason or None,
                'early_stop_notes': sp.early_stop_notes or None,
                'operator_reassigned': bool(sp.operator_reassigned) if sp.operator_reassigned else False,
                'reassignment_task': sp.reassignment_task or None
            }
            
            # Get machine_speed from ShiftProduction
            shift_machine_speed = int(sp.machine_speed) if sp.machine_speed else 0
            
            # Calculate per sub-shift efficiency: runtime_subshift / planned_runtime * 100
            subshift_runtime = 0
            if shift_machine_speed > 0:
                subshift_runtime = int(math.floor(grade_a / shift_machine_speed + 0.5))
            subshift_efficiency = round((subshift_runtime / planned_runtime) * 100, 1) if planned_runtime > 0 else 0
            
            shift_data['machine_speed'] = shift_machine_speed
            shift_data['runtime'] = subshift_runtime
            shift_data['efficiency'] = subshift_efficiency
            
            machines_data[machine_id]['shifts'].append(shift_data)
            
            # Track unique shifts: use default shift duration for average_time
            # and track machine_speed per unique shift for proper averaging
            if shift_num not in machines_data[machine_id]['shifts_seen']:
                machines_data[machine_id]['shifts_seen'][shift_num] = {
                    'default_planned': planned_runtime if planned_runtime > 0 else default_planned,
                    'machine_speed': shift_machine_speed
                }
            else:
                # For multi-product in same shift, keep the larger planned_runtime
                # (the first record usually has the full shift duration)
                existing = machines_data[machine_id]['shifts_seen'][shift_num]
                # Update machine_speed if this record has a higher speed
                if shift_machine_speed > existing['machine_speed']:
                    existing['machine_speed'] = shift_machine_speed
            
            machines_data[machine_id]['total_planned'] += planned_runtime
            machines_data[machine_id]['total_runtime'] += shift_runtime
            machines_data[machine_id]['total_downtime'] += shift_downtime
            machines_data[machine_id]['total_idle'] += shift_idle
            machines_data[machine_id]['total_output'] += shift_data['total']
            machines_data[machine_id]['total_grade_a'] += shift_data['grade_a']
            machines_data[machine_id]['total_grade_b'] += shift_data['grade_b']
            machines_data[machine_id]['total_grade_c'] += shift_data['grade_c']
            machines_data[machine_id]['total_machine_speed'] += shift_machine_speed
            # Note: total_target is now calculated at machine level using formula:
            # Target = Machine Speed × Average Time × Target Efficiency (60%)
            
            if product_name and product_name not in machines_data[machine_id]['products']:
                machines_data[machine_id]['products'].append(product_name)
            
            # Accumulate downtime breakdown with frequency
            # Also track per-shift downtime breakdown
            if shift_num not in machines_data[machine_id]['per_shift_downtime']:
                machines_data[machine_id]['per_shift_downtime'][shift_num] = []
            
            for dt in downtime_breakdown:
                if dt.get('is_idle'):
                    # Track idle breakdown separately
                    existing_idle = next((d for d in machines_data[machine_id]['idle_breakdown'] if d['reason'] == dt['reason']), None)
                    if existing_idle:
                        existing_idle['duration_minutes'] += dt['duration_minutes']
                    else:
                        machines_data[machine_id]['idle_breakdown'].append({'reason': dt['reason'], 'duration_minutes': dt['duration_minutes']})
                
                # Machine-level accumulation
                existing = next((d for d in machines_data[machine_id]['top_3_downtime'] if d['reason'] == dt['reason']), None)
                if existing:
                    existing['duration_minutes'] += dt['duration_minutes']
                    existing['frequency'] = existing.get('frequency', 1) + dt.get('frequency', 1)
                else:
                    dt_copy = dt.copy()
                    dt_copy['frequency'] = dt.get('frequency', 1)
                    machines_data[machine_id]['top_3_downtime'].append(dt_copy)
                
                # Per-shift accumulation
                shift_dt_list = machines_data[machine_id]['per_shift_downtime'][shift_num]
                existing_shift = next((d for d in shift_dt_list if d['reason'] == dt['reason']), None)
                if existing_shift:
                    existing_shift['duration_minutes'] += dt['duration_minutes']
                    existing_shift['frequency'] = existing_shift.get('frequency', 1) + dt.get('frequency', 1)
                else:
                    dt_copy2 = dt.copy()
                    dt_copy2['frequency'] = dt.get('frequency', 1)
                    shift_dt_list.append(dt_copy2)
        
        # Keywords to exclude from Top 3 Downtime (biological/personal breaks + idle time + design change + operator setting)
        excluded_keywords = [
            # Biological/personal breaks
            'istirahat', 'sholat', 'solat', 'makan', 'minum', 'toilet', 
            'wc', 'buang air', 'pipis', 'bab', 'break', 'rest', 'pray',
            'ibadah', 'jumatan', 'jumat', 'dhuhur', 'ashar', 'maghrib',
            # IDLE TIME - exclude from Top 3 downtime (tracked separately)
            'tunggu kain', 'ambil kain', 'tunggu obat', 'tunggu tinta', 'tunggu ingredient',
            'tunggu stiker', 'tunggu packing', 'tunggu packaging', 'tunggu mixing',
            'tunggu label', 'tunggu box', 'tunggu karton', 'tunggu lem',
            'tunggu bahan', 'tunggu material', 'tunggu order', 'tunggu instruksi',
            'menunggu kain', 'menunggu obat', 'menunggu stiker', 'menunggu packing',
            'nunggu kain', 'nunggu obat', 'nunggu stiker', 'nunggu packing',
            'kain habis', 'stiker habis', 'packing habis', 'packaging habis',
            'label habis', 'karton habis', 'box habis', 'lem habis', 'tinta habis',
            'bahan habis', 'material habis', 'ingredient habis', 'obat habis',
            'waiting for', 'standby', 'idle', 'menganggur', 'no order', 'menghabiskan order',
            # DESIGN CHANGE - exclude from Top 3 (treated as separate category)
            'design change', 'design baru', 'ganti design', 'ubah design',
            'repack', 'repacking', 're-pack', 're packing',
            'sanitasi', 'persiapan & sanitasi', 'sterilisasi',
            'setting packaging', 'setting kemasan', 'setting mesin', 'setting mc', 'sanitasi dan setting',
            # OPERATOR SETTING - exclude from Top 3 (not actual repair)
            'setting', 'seting', 'setel', 'adjust', 'adjustment', 'kalibrasi',
            'calibration', 'setup', 'set up', 'konfigurasi', 'konfigurasi ulang',
            'setting parameter', 'setting awal'
        ]
        
        # Only include MACHINE and OPERATOR REPAIR in Top 3
        machine_keywords = [
            'mesin', 'machine', 'rusak', 'error', 'breakdown', 'maintenance',
            'perbaikan', 'repair', 'overhaul', 'service', 'ganti sparepart',
            'spare part', 'komponen', 'sensor', 'motor', 'pneumatic',
            'hydraulic', 'electrical', 'mekanik', 'tooling', 'mould', 'die',
            # Common machine issues
            'conveyor', 'vanbelt', 'belt', 'inkjet', 'printer', 'nozzle',
            'infeeding', 'putus', 'patah', 'bocor', 'macet', 'mampet', 'stuck',
            'tidak keluar', 'tidak jalan', 'tidak nyala', 'mati',
            'kain keluar', 'kain putus', 'kain sobek', 'roll', 'roller', 'bearing', 'gear', 'rantai', 'chain',
            'pompa', 'pump', 'valve', 'seal', 'packing', 'gasket',
            'cutter', 'blade', 'pisau', 'heater', 'pemanas', 'cooler',
            'menggulung', 'slip', 'geser', 'longgar', 'kendor', 'aus',
            # Pressure/air issues
            'tekanan', 'angin', 'drop', 'pressure', 'compressor', 'kompresor',
            'vacuum', 'vakum', 'hisap', 'tiup', 'blower',
            # Temperature issues
            'suhu', 'temperatur', 'panas', 'dingin', 'overheat',
            # Quality/output issues  
            'tidak maksimal', 'kurang', 'lemah', 'jelek', 'cacat', 'defect',
            'lipatan', 'folding', 'lipat', 'tidak rata', 'miring', 'bengkok'
        ]
        
        operator_repair_keywords = [
            'perbaikan', 'operator repair', 'service',
            'ganti part', 'maintenance'
        ]
        
        # Sort top 3 downtime and calculate efficiency for each machine
        for machine_id in machines_data:
            # Filter to only include MACHINE and OPERATOR REPAIR issues
            filtered_downtime = []
            for dt in machines_data[machine_id]['top_3_downtime']:
                # Skip items marked as idle time
                if dt.get('is_idle', False):
                    continue
                reason_lower = dt['reason'].lower()
                category = dt.get('category', '').lower()
                
                # PRIORITY 1: Use category from user input if available
                # If category is 'mesin' or 'operator', include in Top 3
                if category in ['mesin', 'machine', 'operator']:
                    # Still exclude biological breaks even if categorized as mesin/operator
                    biological_breaks = ['istirahat', 'sholat', 'solat', 'makan', 'minum', 
                                        'toilet', 'wc', 'buang air', 'jumatan', 'jumat']
                    if not any(kw in reason_lower for kw in biological_breaks):
                        filtered_downtime.append(dt)
                    continue
                
                # PRIORITY 2: Check excluded keywords (skip if matched)
                if any(kw in reason_lower for kw in excluded_keywords):
                    continue
                    
                # PRIORITY 3: Check machine or operator repair keywords
                if (any(kw in reason_lower for kw in machine_keywords) or
                    any(kw in reason_lower for kw in operator_repair_keywords)):
                    filtered_downtime.append(dt)
            
            filtered_downtime.sort(key=lambda x: x['duration_minutes'], reverse=True)
            machines_data[machine_id]['top_3_downtime'] = filtered_downtime[:3]
            # Sort shifts by shift number
            machines_data[machine_id]['shifts'].sort(key=lambda x: x['shift'])
            
            # Consolidate early_stop to LAST sub-shift per shift number
            # e.g., if shift 1a has early_stop, move it to shift 1c (last entry)
            shifts_list = machines_data[machine_id]['shifts']
            shift_groups = {}
            for idx, s in enumerate(shifts_list):
                sn = s['shift']
                if sn not in shift_groups:
                    shift_groups[sn] = []
                shift_groups[sn].append(idx)
            
            for sn, indices in shift_groups.items():
                if len(indices) > 1:
                    # Collect early_stop from all entries in this shift group
                    early_stop_data = None
                    for idx in indices:
                        if shifts_list[idx].get('early_stop'):
                            early_stop_data = {
                                'early_stop': True,
                                'early_stop_time': shifts_list[idx].get('early_stop_time'),
                                'early_stop_reason': shifts_list[idx].get('early_stop_reason'),
                                'early_stop_notes': shifts_list[idx].get('early_stop_notes'),
                                'operator_reassigned': shifts_list[idx].get('operator_reassigned'),
                                'reassignment_task': shifts_list[idx].get('reassignment_task'),
                            }
                            # Clear from this entry
                            shifts_list[idx]['early_stop'] = False
                            shifts_list[idx]['early_stop_time'] = None
                            shifts_list[idx]['early_stop_reason'] = None
                            shifts_list[idx]['early_stop_notes'] = None
                    
                    # Assign to last entry
                    if early_stop_data:
                        last_idx = indices[-1]
                        shifts_list[last_idx].update(early_stop_data)
            
            # Calculate Runtime = Grade A / machine speed
            total_grade_a = machines_data[machine_id]['total_grade_a']
            total_downtime = machines_data[machine_id]['total_downtime']
            total_idle = machines_data[machine_id]['total_idle']
            
            # Get machine speed: average by UNIQUE shifts (not total records)
            # e.g., 3 records in shift_1 all with speed=25 → average = 25, not 25*3/3
            shifts_seen = machines_data[machine_id].get('shifts_seen', {})
            unique_shift_count = len(shifts_seen)
            shift_count = unique_shift_count if unique_shift_count > 0 else 1
            
            # Sum machine_speed per unique shift for averaging
            machine_speed_total = sum(s['machine_speed'] for s in shifts_seen.values())
            machine_speed_per_minute = machine_speed_total / shift_count if shift_count > 0 else 0
            
            # Runtime = Grade A / Speed (in minutes)
            # Use math.floor(x + 0.5) for standard rounding (Python's round uses banker's rounding)
            import math
            runtime = 0
            if machine_speed_per_minute > 0:
                runtime = int(math.floor(total_grade_a / machine_speed_per_minute + 0.5))
            
            # Average Time = total available time based on UNIQUE shifts
            # For multi-product in same shift, use default shift duration (not sum of per-record planned_runtime)
            is_friday = target_date.weekday() == 4
            average_time = 0
            for sn, sdata in shifts_seen.items():
                if sn == 1:
                    average_time += 540 if is_friday else 510
                elif sn == 2:
                    average_time += 510
                else:
                    average_time += 450
            
            # Fallback
            if average_time == 0:
                default_minutes_per_shift = 540 if is_friday else 510
                average_time = default_minutes_per_shift
            
            # Get user-input waktu_tidak_tercatat from shift data (sum of all shifts)
            user_waktu_tidak_tercatat = sum(shift.get('waktu_tidak_tercatat', 0) for shift in machines_data[machine_id]['shifts'])
            
            # Waktu Tercatat = Runtime + Downtime + Idle (Idle IS recorded time, just different category)
            waktu_tercatat = runtime + total_downtime + total_idle
            
            # Waktu Tidak Tercatat = Always calculate dynamically from average_time
            # (stored DB value is stale because runtime is recalculated dynamically)
            waktu_tidak_tercatat = max(0, average_time - waktu_tercatat)
            
            # Store calculated values
            machines_data[machine_id]['runtime'] = runtime
            machines_data[machine_id]['average_time'] = average_time
            # IMPORTANT: Override total_planned with capped average_time
            # This prevents inflation when multiple sub-shifts exist in same shift number
            # e.g., Shift 1a (300m) + Shift 1b (210m) should still total 510m, not 510+510
            machines_data[machine_id]['total_planned'] = average_time
            machines_data[machine_id]['waktu_tercatat'] = waktu_tercatat
            machines_data[machine_id]['waktu_tidak_tercatat'] = waktu_tidak_tercatat
            machines_data[machine_id]['mrt'] = runtime  # For backward compatibility
            machines_data[machine_id]['total_time'] = average_time
            
            # Efficiency = Runtime / Average Time * 100 (this IS the machine efficiency now)
            if average_time > 0:
                machines_data[machine_id]['efficiency'] = round((runtime / average_time) * 100, 1)
                machines_data[machine_id]['machine_efficiency'] = machines_data[machine_id]['efficiency']
            machines_data[machine_id]['machine_speed'] = int(round(machine_speed_per_minute))  # pcs/menit
            
            # Calculate Quality = (Grade A / Total Output) * 100
            total_output = machines_data[machine_id]['total_output']
            if total_output > 0:
                machines_data[machine_id]['quality'] = round((total_grade_a / total_output) * 100, 1)
            
            # Calculate Target = Machine Speed × Average Time × Target Efficiency (60%)
            target_efficiency_pct = machines_data[machine_id].get('target_efficiency', 60) / 100
            calculated_target = int(machine_speed_per_minute * average_time * target_efficiency_pct)
            machines_data[machine_id]['total_target'] = calculated_target
            
            # ========== PER-SHIFT EFFICIENCY ==========
            # Group shift entries by shift number and calculate efficiency per shift
            shift_summaries = {}
            for s in machines_data[machine_id]['shifts']:
                sn = s['shift']
                if sn not in shift_summaries:
                    shift_summaries[sn] = {
                        'shift': sn,
                        'grade_a': 0,
                        'grade_b': 0,
                        'grade_c': 0,
                        'total_output': 0,
                        'downtime': 0,
                        'idle': 0,
                        'products': [],
                    }
                shift_summaries[sn]['grade_a'] += s['grade_a']
                shift_summaries[sn]['grade_b'] += s['grade_b']
                shift_summaries[sn]['grade_c'] += s['grade_c']
                shift_summaries[sn]['total_output'] += s['total']
                shift_summaries[sn]['downtime'] += s['downtime_minutes']
                shift_summaries[sn]['idle'] += s['idle_time_minutes']
                if s.get('product_name') and s['product_name'] not in shift_summaries[sn]['products']:
                    shift_summaries[sn]['products'].append(s['product_name'])
            
            for sn, ss in shift_summaries.items():
                # Get machine speed for this shift from shifts_seen
                shift_speed = shifts_seen.get(sn, {}).get('machine_speed', 0)
                
                # Available time for this shift number
                if sn == 1:
                    ss['average_time'] = 540 if is_friday else 510
                elif sn == 2:
                    ss['average_time'] = 510
                else:
                    ss['average_time'] = 450
                
                # Runtime = Grade A / Speed
                ss['machine_speed'] = shift_speed
                ss['runtime'] = 0
                if shift_speed > 0:
                    ss['runtime'] = int(math.floor(ss['grade_a'] / shift_speed + 0.5))
                
                # Waktu tercatat & tidak tercatat
                ss['waktu_tercatat'] = ss['runtime'] + ss['downtime'] + ss['idle']
                ss['waktu_tidak_tercatat'] = max(0, ss['average_time'] - ss['waktu_tercatat'])
                
                # Efficiency = Runtime / Available Time * 100 (this IS the machine efficiency)
                ss['efficiency'] = 0
                if ss['average_time'] > 0:
                    ss['efficiency'] = round((ss['runtime'] / ss['average_time']) * 100, 1)
                ss['machine_efficiency'] = ss['efficiency']
                
                # Quality = Grade A / Total Output * 100
                ss['quality'] = 0
                if ss['total_output'] > 0:
                    ss['quality'] = round((ss['grade_a'] / ss['total_output']) * 100, 1)
                
                # Target = Speed * Available Time * Target Efficiency%
                ss['target'] = int(shift_speed * ss['average_time'] * target_efficiency_pct)
                
                # Per-shift Top 3 Downtime (same filter logic as machine-level)
                per_shift_dt = machines_data[machine_id].get('per_shift_downtime', {}).get(sn, [])
                filtered_shift_dt = []
                for dt in per_shift_dt:
                    if dt.get('is_idle', False):
                        continue
                    reason_lower = dt['reason'].lower()
                    category = dt.get('category', '').lower()
                    if category in ['mesin', 'machine', 'operator']:
                        biological_breaks = ['istirahat', 'sholat', 'solat', 'makan', 'minum',
                                            'toilet', 'wc', 'buang air', 'jumatan', 'jumat']
                        if not any(kw in reason_lower for kw in biological_breaks):
                            filtered_shift_dt.append(dt)
                        continue
                    if any(kw in reason_lower for kw in excluded_keywords):
                        continue
                    if (any(kw in reason_lower for kw in machine_keywords) or
                        any(kw in reason_lower for kw in operator_repair_keywords)):
                        filtered_shift_dt.append(dt)
                filtered_shift_dt.sort(key=lambda x: x['duration_minutes'], reverse=True)
                ss['top_3_downtime'] = [{'reason': d['reason'], 'duration_minutes': d['duration_minutes'], 'frequency': d.get('frequency', 1)} for d in filtered_shift_dt[:3]]
            
            # Store as list sorted by shift number
            machines_data[machine_id]['shift_summaries'] = [
                shift_summaries[sn] for sn in sorted(shift_summaries.keys())
            ]
        
        # Convert to list and sort by machine name
        result = list(machines_data.values())
        result.sort(key=lambda x: x['machine_name'])
        
        return jsonify({
            'date': target_date.isoformat(),
            'machines': result,
            'total_machines': len(result)
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/daily-controller-detail', methods=['GET'])
@jwt_required()
def get_daily_controller_detail():
    """Get detailed production data for a specific date - includes all downtime, timeline, and WO details"""
    try:
        from models.production import ShiftProduction, WorkOrder
        
        selected_date = request.args.get('date')
        if selected_date:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        else:
            target_date = get_local_today()
        
        # Get all shift productions for this date with eager loading
        from sqlalchemy.orm import joinedload
        shift_records = ShiftProduction.query.options(
            joinedload(ShiftProduction.work_order),
            joinedload(ShiftProduction.machine),
            joinedload(ShiftProduction.product),
            joinedload(ShiftProduction.operator),
            joinedload(ShiftProduction.supervisor)
        ).filter(
            ShiftProduction.production_date == target_date
        ).order_by(ShiftProduction.machine_id, ShiftProduction.shift).all()
        
        # IDLE TIME keywords
        idle_keywords = [
            'tunggu kain', 'ambil kain', 'menunggu kain', 'nunggu kain', 'kain belum datang',
            'tunggu obat', 'menunggu obat', 'nunggu obat', 'obat belum datang',
            'tunggu tinta', 'menunggu tinta', 'nunggu tinta',
            'tunggu ingredient', 'ingredient habis', 'tunggu bahan kimia',
            'tunggu stiker', 'menunggu stiker', 'nunggu stiker', 'stiker belum datang',
            'tunggu packing', 'tunggu packaging', 'menunggu packing', 'nunggu packing',
            'packaging belum datang', 'box belum datang',
            'tunggu mixing', 'menunggu mixing', 'nunggu mixing', 'mixing belum siap',
            'tunggu label', 'tunggu box', 'tunggu karton', 'tunggu lem',
            'tunggu produk', 'tunggu temperatur stabil', 'tunggu temperatur',
            'tunggu temperature stabil', 'tunggu temperature',
            'tunggu bahan', 'tunggu material', 'tunggu order', 'tunggu instruksi',
            'tunggu approval', 'tunggu qc', 'tunggu hasil qc',
            'kain habis', 'stiker habis', 'packing habis', 'packaging habis',
            'label habis', 'karton habis', 'box habis', 'lem habis', 'tinta habis',
            'bahan habis', 'material habis',
            'waiting for', 'standby material', 'waiting material', 'no material',
            'idle', 'standby', 'menganggur', 'tidak ada order', 'no order', 'menhabiskan order',
            'susun produk'
        ]
        
        # Downtime category mapping - use unified function from helpers
        def get_downtime_category(reason_lower):
            return detect_downtime_category(reason_lower)
        
        # PIC mapping berdasarkan kategori downtime
        PIC_BY_CATEGORY = {
            'mesin': 'MTC',
            'operator': 'Supervisor',
            'material': 'Warehouse',
            'design': 'Supervisor',
            'idle': 'Supervisor',
            'istirahat': 'Supervisor',
            'others': 'Supervisor',
            'lainnya': 'Supervisor'
        }
        
        # Group data by machine
        machines_data = {}
        all_work_orders = {}
        timeline_data = []
        
        for sp in shift_records:
            machine_id = sp.machine_id
            
            if machine_id not in machines_data:
                machines_data[machine_id] = {
                    'machine_id': machine_id,
                    'machine_name': sp.machine.name if sp.machine else f'Machine {machine_id}',
                    'machine_code': sp.machine.code if sp.machine else None,
                    'target_efficiency': int(sp.machine.target_efficiency) if sp.machine and sp.machine.target_efficiency else 60,
                    'date': target_date.isoformat(),
                    'all_downtime': [],
                    'work_orders': [],
                    'timeline': [],
                    'operators': [],
                    'supervisors': [],
                    'total_output': 0,
                    'total_grade_a': 0,
                    'total_grade_b': 0,
                    'total_grade_c': 0,
                    'total_downtime': 0,
                    'total_idle': 0,
                    'total_runtime': 0
                }
            
            # Get operator and supervisor info (PIC)
            operator_name = None
            supervisor_name = None
            if sp.operator:
                operator_name = sp.operator.full_name if hasattr(sp.operator, 'full_name') else f"Employee {sp.operator_id}"
                if operator_name not in machines_data[machine_id]['operators']:
                    machines_data[machine_id]['operators'].append(operator_name)
            if sp.supervisor:
                supervisor_name = sp.supervisor.full_name if hasattr(sp.supervisor, 'full_name') else f"Employee {sp.supervisor_id}"
                if supervisor_name not in machines_data[machine_id]['supervisors']:
                    machines_data[machine_id]['supervisors'].append(supervisor_name)
            
            # Extract shift number
            shift_num = 1
            if sp.shift:
                shift_match = re.search(r'(\d+)', str(sp.shift))
                if shift_match:
                    shift_num = int(shift_match.group(1))
            
            # Parse ALL downtime from issues field
            if sp.issues:
                issue_parts = sp.issues.split(';')
                for part in issue_parts:
                    part = part.strip()
                    if not part:
                        continue
                    match = re.match(r'(\d+)\s*menit\s*-\s*(.+)', part, re.IGNORECASE)
                    if match:
                        duration = int(match.group(1))
                        reason_with_tag = match.group(2).strip()
                        
                        # Check for category tag
                        category_match = re.search(r'\[(\w+)\]', reason_with_tag, re.IGNORECASE)
                        explicit_category = category_match.group(1).lower() if category_match else None
                        
                        reason = re.sub(r'\s*\[\w+\]\s*$', '', reason_with_tag).strip()
                        reason_lower = reason.lower()
                        
                        # Determine category
                        if explicit_category == 'idle':
                            category = 'idle'
                        elif explicit_category:
                            category = explicit_category
                        else:
                            category = get_downtime_category(reason_lower)
                        
                        # Always re-check: if auto-detect says 'idle', override explicit tag
                        auto_cat = get_downtime_category(reason_lower)
                        if auto_cat == 'idle':
                            category = 'idle'
                        
                        is_idle = category == 'idle'
                        
                        downtime_entry = {
                            'reason': reason,
                            'duration_minutes': duration,
                            'category': category,
                            'is_idle': is_idle,
                            'shift': shift_num,
                            'pic': PIC_BY_CATEGORY.get(category, '-'),
                            'product_name': sp.product.name if sp.product else None,
                            'wo_number': sp.work_order.wo_number if sp.work_order else None
                        }
                        machines_data[machine_id]['all_downtime'].append(downtime_entry)
                        
                        if is_idle:
                            machines_data[machine_id]['total_idle'] += duration
                        else:
                            machines_data[machine_id]['total_downtime'] += duration
            
            # Also include downtime_records from database (if available)
            try:
                for dr in sp.downtime_records:
                    resolved_by_name = None
                    try:
                        if dr.resolved_by_employee:
                            resolved_by_name = getattr(dr.resolved_by_employee, 'full_name', None)
                    except:
                        pass
                    
                    dr_category = dr.downtime_category or 'others'
                    # Override: if auto-detect says 'idle', use that
                    dr_reason = (dr.downtime_reason or '').lower()
                    if get_downtime_category(dr_reason) == 'idle':
                        dr_category = 'idle'
                    downtime_entry = {
                        'reason': dr.downtime_reason or 'Unknown',
                        'duration_minutes': dr.duration_minutes or 0,
                        'category': dr_category,
                        'is_idle': dr_category in ['idle', 'material_shortage'],
                        'shift': shift_num,
                        'pic': PIC_BY_CATEGORY.get(dr_category, '-'),
                        'root_cause': dr.root_cause,
                        'action_taken': dr.action_taken,
                        'status': dr.status,
                        'product_name': sp.product.name if sp.product else None,
                        'wo_number': sp.work_order.wo_number if sp.work_order else None
                    }
                    # Avoid duplicates
                    existing = next((d for d in machines_data[machine_id]['all_downtime'] 
                                   if d['reason'] == dr.downtime_reason and d['shift'] == shift_num), None)
                    if not existing:
                        machines_data[machine_id]['all_downtime'].append(downtime_entry)
            except Exception as dr_error:
                pass  # Skip downtime_records if there's an error
            
            # Work Order details
            if sp.work_order:
                try:
                    wo = sp.work_order
                    wo_key = f"{machine_id}_{wo.id}"
                    
                    # Get product info safely
                    wo_product_name = 'Unknown'
                    wo_product_code = None
                    try:
                        if wo.product:
                            wo_product_name = wo.product.name
                            wo_product_code = wo.product.code
                        elif sp.product:
                            wo_product_name = sp.product.name
                            wo_product_code = sp.product.code
                    except:
                        pass
                    
                    if wo_key not in all_work_orders:
                        all_work_orders[wo_key] = {
                            'wo_id': wo.id,
                            'wo_number': wo.wo_number,
                            'machine_id': machine_id,
                            'machine_name': machines_data[machine_id]['machine_name'],
                            'product_name': wo_product_name,
                            'product_code': wo_product_code,
                            'target_quantity': float(wo.target_quantity) if wo.target_quantity else 0,
                            'actual_quantity': 0,
                            'good_quantity': 0,
                            'reject_quantity': 0,
                            'rework_quantity': 0,
                            'efficiency': 0,
                            'quality_rate': 0,
                            'shifts_worked': [],
                            'operators': [],
                            'supervisors': [],
                            'status': wo.status,
                            'start_date': wo.start_date.isoformat() if wo.start_date else None,
                            'due_date': wo.due_date.isoformat() if wo.due_date else None
                        }
                    
                    # Accumulate quantities
                    all_work_orders[wo_key]['actual_quantity'] += float(sp.actual_quantity) if sp.actual_quantity else 0
                    all_work_orders[wo_key]['good_quantity'] += float(sp.good_quantity) if sp.good_quantity else 0
                    all_work_orders[wo_key]['reject_quantity'] += float(sp.reject_quantity) if sp.reject_quantity else 0
                    all_work_orders[wo_key]['rework_quantity'] += float(sp.rework_quantity) if sp.rework_quantity else 0
                    
                    if shift_num not in all_work_orders[wo_key]['shifts_worked']:
                        all_work_orders[wo_key]['shifts_worked'].append(shift_num)
                    if operator_name and operator_name not in all_work_orders[wo_key]['operators']:
                        all_work_orders[wo_key]['operators'].append(operator_name)
                    if supervisor_name and supervisor_name not in all_work_orders[wo_key]['supervisors']:
                        all_work_orders[wo_key]['supervisors'].append(supervisor_name)
                except Exception as wo_error:
                    pass  # Skip work order if there's an error
            
            # Timeline entry
            shift_times = {1: ('06:30', '15:00'), 2: ('15:00', '23:00'), 3: ('23:00', '06:30')}
            start_time, end_time = shift_times.get(shift_num, ('00:00', '00:00'))
            
            timeline_entry = {
                'machine_id': machine_id,
                'machine_name': machines_data[machine_id]['machine_name'],
                'shift': shift_num,
                'start_time': start_time,
                'end_time': end_time,
                'product_name': sp.product.name if sp.product else 'Unknown',
                'wo_number': sp.work_order.wo_number if sp.work_order else None,
                'output': float(sp.actual_quantity) if sp.actual_quantity else 0,
                'grade_a': float(sp.good_quantity) if sp.good_quantity else 0,
                'downtime': int(sp.downtime_mesin or 0) + int(sp.downtime_operator or 0) + int(sp.downtime_material or 0) + int(sp.downtime_design or 0) + int(sp.downtime_others or 0),
                'idle': int(sp.idle_time) if sp.idle_time else 0,
                'operator': operator_name,
                'supervisor': supervisor_name,
                'efficiency_rate': float(sp.efficiency_rate) if sp.efficiency_rate else 0,
                'quality_rate': float(sp.quality_rate) if sp.quality_rate else 0
            }
            machines_data[machine_id]['timeline'].append(timeline_entry)
            
            # Accumulate totals
            machines_data[machine_id]['total_output'] += float(sp.actual_quantity) if sp.actual_quantity else 0
            machines_data[machine_id]['total_grade_a'] += float(sp.good_quantity) if sp.good_quantity else 0
            machines_data[machine_id]['total_grade_b'] += float(sp.rework_quantity) if sp.rework_quantity else 0
            machines_data[machine_id]['total_grade_c'] += float(sp.reject_quantity) if sp.reject_quantity else 0
        
        # Calculate WO efficiency and quality
        for wo_key, wo_data in all_work_orders.items():
            if wo_data['target_quantity'] > 0:
                wo_data['efficiency'] = round((wo_data['good_quantity'] / wo_data['target_quantity']) * 100, 1)
            if wo_data['actual_quantity'] > 0:
                wo_data['quality_rate'] = round((wo_data['good_quantity'] / wo_data['actual_quantity']) * 100, 1)
        
        # Attach WO data to machines
        for wo_key, wo_data in all_work_orders.items():
            machine_id = wo_data['machine_id']
            if machine_id in machines_data:
                machines_data[machine_id]['work_orders'].append(wo_data)
        
        # Sort downtime by duration (descending)
        for machine_id in machines_data:
            machines_data[machine_id]['all_downtime'].sort(key=lambda x: x['duration_minutes'], reverse=True)
            machines_data[machine_id]['timeline'].sort(key=lambda x: x['shift'])
        
        # Convert to list
        result = list(machines_data.values())
        result.sort(key=lambda x: x['machine_name'])
        
        # Calculate summary
        wo_list = list(all_work_orders.values())
        total_downtime = sum(m['total_downtime'] for m in result)
        total_idle = sum(m['total_idle'] for m in result)
        total_output = sum(m['total_output'] for m in result)
        total_grade_a = sum(m['total_grade_a'] for m in result)
        
        # Calculate average efficiency from work orders
        avg_efficiency = 0
        if wo_list:
            valid_efficiencies = [wo['efficiency'] for wo in wo_list if wo['efficiency'] > 0]
            if valid_efficiencies:
                avg_efficiency = round(sum(valid_efficiencies) / len(valid_efficiencies), 1)
        
        # If no WO efficiency, calculate from output
        if avg_efficiency == 0 and total_output > 0:
            total_target = sum(wo['target_quantity'] for wo in wo_list) if wo_list else 0
            if total_target > 0:
                avg_efficiency = round((total_grade_a / total_target) * 100, 1)
        
        return jsonify({
            'date': target_date.isoformat(),
            'machines': result,
            'work_orders': wo_list,
            'total_machines': len(result),
            'summary': {
                'total_downtime': total_downtime,
                'total_idle': total_idle,
                'total_output': total_output,
                'total_grade_a': total_grade_a,
                'total_work_orders': len(wo_list),
                'avg_efficiency': avg_efficiency
            }
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/efficiency-alerts', methods=['GET'])
@jwt_required()
def get_efficiency_alerts():
    """Get machines with efficiency below target for today"""
    try:
        from models.production import ShiftProduction, Machine
        
        # Get today's date
        today = get_local_now().date()
        
        # Get all shift productions for today
        shift_records = ShiftProduction.query.filter(
            ShiftProduction.production_date == today
        ).all()
        
        # Group by machine and calculate efficiency
        machines_data = {}
        
        for sp in shift_records:
            machine_id = sp.machine_id
            if not machine_id:
                continue
                
            if machine_id not in machines_data:
                machines_data[machine_id] = {
                    'machine_id': machine_id,
                    'machine_name': sp.machine.name if sp.machine else f'Machine {machine_id}',
                    'target_efficiency': int(sp.machine.target_efficiency) if sp.machine and sp.machine.target_efficiency else 60,
                    'total_grade_a': 0,
                    'total_downtime': 0,
                    'total_idle': 0,
                    'total_machine_speed': 0,
                    'shift_count': 0
                }
            
            grade_a = int(sp.good_quantity) if sp.good_quantity else 0
            downtime = int(sp.downtime_mesin or 0) + int(sp.downtime_operator or 0) + int(sp.downtime_material or 0) + int(sp.downtime_design or 0) + int(sp.downtime_others or 0)
            idle = int(sp.idle_time) if sp.idle_time else 0
            speed = int(sp.machine_speed) if sp.machine_speed else 0
            
            machines_data[machine_id]['total_grade_a'] += grade_a
            machines_data[machine_id]['total_downtime'] += downtime
            machines_data[machine_id]['total_idle'] += idle
            machines_data[machine_id]['total_machine_speed'] += speed
            machines_data[machine_id]['shift_count'] += 1
        
        # Calculate efficiency and filter alerts
        alerts = []
        for machine_id, data in machines_data.items():
            speed_per_min = data['total_machine_speed'] / data['shift_count'] if data['shift_count'] > 0 else 0
            mrt = data['total_grade_a'] / speed_per_min if speed_per_min > 0 else 0
            total_time = mrt + data['total_downtime'] + data['total_idle']
            efficiency = round((mrt / total_time) * 100, 1) if total_time > 0 else 0
            
            target = data['target_efficiency']
            if efficiency < target:
                alert_level = 'critical' if efficiency < target * 0.7 else 'warning'
                alerts.append({
                    'machine_id': machine_id,
                    'machine_name': data['machine_name'],
                    'efficiency': efficiency,
                    'target': target,
                    'gap': round(target - efficiency, 1),
                    'level': alert_level
                })
        
        # Sort by gap (largest gap first)
        alerts.sort(key=lambda x: x['gap'], reverse=True)
        
        return jsonify({
            'date': today.isoformat(),
            'alerts': alerts,
            'total_alerts': len(alerts),
            'critical_count': len([a for a in alerts if a['level'] == 'critical']),
            'warning_count': len([a for a in alerts if a['level'] == 'warning'])
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/weekly-controller', methods=['GET'])
@jwt_required()
def get_weekly_controller():
    """Get weekly efficiency summary for all machines"""
    try:
        from models.production import ShiftProduction, Machine
        
        # Get week start date (Monday) - default to current week
        date_str = request.args.get('week_start')
        if date_str:
            week_start = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            today = get_local_now().date()
            week_start = today - timedelta(days=today.weekday())  # Monday
        
        week_end = week_start + timedelta(days=6)  # Sunday
        
        # Get all shift productions for the week
        shift_records = ShiftProduction.query.filter(
            ShiftProduction.production_date >= week_start,
            ShiftProduction.production_date <= week_end
        ).all()
        
        # Group by machine
        machines_data = {}
        
        for sp in shift_records:
            machine_id = sp.machine_id
            if not machine_id:
                continue
                
            if machine_id not in machines_data:
                machines_data[machine_id] = {
                    'machine_id': machine_id,
                    'machine_name': sp.machine.name if sp.machine else f'Machine {machine_id}',
                    'machine_code': sp.machine.code if sp.machine else None,
                    'target_efficiency': int(sp.machine.target_efficiency) if sp.machine and sp.machine.target_efficiency else 60,
                    'daily_data': {},
                    'total_grade_a': 0,
                    'total_output': 0,
                    'total_downtime': 0,
                    'total_idle': 0,
                    'total_machine_speed': 0,
                    'shift_count': 0
                }
            
            # Group by date
            date_key = sp.production_date.isoformat()
            if date_key not in machines_data[machine_id]['daily_data']:
                machines_data[machine_id]['daily_data'][date_key] = {
                    'grade_a': 0,
                    'output': 0,
                    'downtime': 0,
                    'idle': 0,
                    'machine_speed': 0,
                    'shift_count': 0,
                    'shifts_seen': {}
                }
            
            grade_a = int(sp.good_quantity) if sp.good_quantity else 0
            output = int(sp.actual_quantity) if sp.actual_quantity else 0
            downtime = int(sp.downtime_mesin or 0) + int(sp.downtime_operator or 0) + int(sp.downtime_material or 0) + int(sp.downtime_design or 0) + int(sp.downtime_others or 0)
            idle = int(sp.idle_time) if sp.idle_time else 0
            speed = int(sp.machine_speed) if sp.machine_speed else 0
            
            machines_data[machine_id]['daily_data'][date_key]['grade_a'] += grade_a
            machines_data[machine_id]['daily_data'][date_key]['output'] += output
            machines_data[machine_id]['daily_data'][date_key]['downtime'] += downtime
            machines_data[machine_id]['daily_data'][date_key]['idle'] += idle
            # Track unique shifts per day for proper speed averaging
            import re as _re
            shift_num = 1
            if sp.shift:
                _m = _re.search(r'(\d+)', str(sp.shift))
                if _m:
                    shift_num = int(_m.group(1))
            day_shifts_seen = machines_data[machine_id]['daily_data'][date_key]['shifts_seen']
            if shift_num not in day_shifts_seen:
                day_shifts_seen[shift_num] = speed
            elif speed > day_shifts_seen[shift_num]:
                day_shifts_seen[shift_num] = speed
            machines_data[machine_id]['daily_data'][date_key]['shift_count'] += 1
            
            machines_data[machine_id]['total_grade_a'] += grade_a
            machines_data[machine_id]['total_output'] += output
            machines_data[machine_id]['total_downtime'] += downtime
            machines_data[machine_id]['total_idle'] += idle
            machines_data[machine_id]['total_machine_speed'] += speed
            machines_data[machine_id]['shift_count'] += 1
        
        # Calculate weekly efficiency for each machine
        for machine_id in machines_data:
            data = machines_data[machine_id]
            
            # Calculate daily efficiencies
            for date_key in data['daily_data']:
                day_data = data['daily_data'][date_key]
                # Average speed by unique shifts, not total records
                day_shifts = day_data.get('shifts_seen', {})
                speed_per_min = sum(day_shifts.values()) / len(day_shifts) if day_shifts else 0
                mrt = day_data['grade_a'] / speed_per_min if speed_per_min > 0 else 0
                total_time = mrt + day_data['downtime'] + day_data['idle']
                day_data['efficiency'] = round((mrt / total_time) * 100, 1) if total_time > 0 else 0
                day_data['mrt'] = round(mrt, 1)
            
            # Calculate weekly average efficiency - use unique shifts across all days
            all_shifts = {}
            for dk in data['daily_data']:
                for sn, spd in data['daily_data'][dk].get('shifts_seen', {}).items():
                    day_shift_key = f"{dk}_{sn}"
                    all_shifts[day_shift_key] = spd
            speed_per_min = sum(all_shifts.values()) / len(all_shifts) if all_shifts else 0
            mrt = data['total_grade_a'] / speed_per_min if speed_per_min > 0 else 0
            total_time = mrt + data['total_downtime'] + data['total_idle']
            data['avg_efficiency'] = round((mrt / total_time) * 100, 1) if total_time > 0 else 0
            data['mrt'] = round(mrt, 1)
            data['total_time'] = round(total_time, 1)
            data['quality'] = round((data['total_grade_a'] / data['total_output']) * 100, 1) if data['total_output'] > 0 else 0
        
        result = list(machines_data.values())
        result.sort(key=lambda x: x['machine_name'])
        
        # === NEW: Collect products produced and downtime data ===
        products_produced = {}  # {product_name: {'quantity': X, 'grade_a': Y, 'machines': set()}}
        all_downtime = {}  # {reason: {'duration': X, 'frequency': Y}}
        daily_summary = {}  # {date: {'output': X, 'downtime': Y, 'efficiency': Z}}
        
        for sp in shift_records:
            # Products produced
            product_name = None
            if sp.product:
                product_name = sp.product.name
            elif sp.work_order and sp.work_order.product:
                product_name = sp.work_order.product.name
            
            if product_name:
                if product_name not in products_produced:
                    products_produced[product_name] = {
                        'product_name': product_name,
                        'quantity': 0,
                        'grade_a': 0,
                        'grade_b': 0,
                        'grade_c': 0,
                        'machines': set()
                    }
                products_produced[product_name]['quantity'] += int(sp.actual_quantity) if sp.actual_quantity else 0
                products_produced[product_name]['grade_a'] += int(sp.good_quantity) if sp.good_quantity else 0
                products_produced[product_name]['grade_b'] += int(sp.rework_quantity) if sp.rework_quantity else 0
                products_produced[product_name]['grade_c'] += int(sp.reject_quantity) if sp.reject_quantity else 0
                if sp.machine:
                    products_produced[product_name]['machines'].add(sp.machine.name)
            
            # Parse downtime from issues field
            if sp.issues:
                issue_parts = sp.issues.split(';')
                for part in issue_parts:
                    part = part.strip()
                    if not part:
                        continue
                    match = re.match(r'(\d+)\s*menit\s*-\s*(.+)', part, re.IGNORECASE)
                    if match:
                        duration = int(match.group(1))
                        reason = match.group(2).strip()
                        reason = re.sub(r'\s*\[\w+\]\s*$', '', reason).strip()
                        
                        if reason not in all_downtime:
                            all_downtime[reason] = {'reason': reason, 'duration': 0, 'frequency': 0, 'category': detect_downtime_category(reason.lower())}
                        all_downtime[reason]['duration'] += duration
                        all_downtime[reason]['frequency'] += 1
            
            # Daily summary for chart
            date_key = sp.production_date.isoformat()
            if date_key not in daily_summary:
                daily_summary[date_key] = {
                    'date': date_key,
                    'output': 0,
                    'grade_a': 0,
                    'downtime': 0,
                    'idle': 0,
                    'machine_speed': 0,
                    'shift_count': 0
                }
            daily_summary[date_key]['output'] += int(sp.actual_quantity) if sp.actual_quantity else 0
            daily_summary[date_key]['grade_a'] += int(sp.good_quantity) if sp.good_quantity else 0
            daily_summary[date_key]['downtime'] += int(sp.downtime_mesin or 0) + int(sp.downtime_operator or 0) + int(sp.downtime_material or 0) + int(sp.downtime_design or 0) + int(sp.downtime_others or 0)
            daily_summary[date_key]['idle'] += int(sp.idle_time) if sp.idle_time else 0
            daily_summary[date_key]['machine_speed'] += int(sp.machine_speed) if sp.machine_speed else 0
            daily_summary[date_key]['shift_count'] += 1
        
        # Calculate efficiency for daily summary
        for date_key in daily_summary:
            day = daily_summary[date_key]
            speed_per_min = day['machine_speed'] / day['shift_count'] if day['shift_count'] > 0 else 0
            mrt = day['grade_a'] / speed_per_min if speed_per_min > 0 else 0
            total_time = mrt + day['downtime'] + day['idle']
            day['efficiency'] = round((mrt / total_time) * 100, 1) if total_time > 0 else 0
        
        # Convert products to list and sort
        products_list = []
        for p in products_produced.values():
            p['machines'] = list(p['machines'])
            products_list.append(p)
        products_list.sort(key=lambda x: x['quantity'], reverse=True)
        
        # Sort downtime by duration and get top 10
        downtime_list = list(all_downtime.values())
        downtime_list.sort(key=lambda x: x['duration'], reverse=True)
        top_10_downtime = downtime_list[:10]
        
        # Daily chart data sorted by date
        chart_data = list(daily_summary.values())
        chart_data.sort(key=lambda x: x['date'])
        
        # Pie chart data - downtime distribution
        pie_chart_data = [{'name': d['reason'][:20], 'value': d['duration']} for d in top_10_downtime]
        
        return jsonify({
            'week_start': week_start.isoformat(),
            'week_end': week_end.isoformat(),
            'machines': result,
            'total_machines': len(result),
            'products_produced': products_list,
            'all_downtime': downtime_list,
            'top_10_downtime': top_10_downtime,
            'chart_data': chart_data,
            'pie_chart_data': pie_chart_data
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/monthly-controller', methods=['GET'])
@jwt_required()
def get_monthly_controller():
    """Get monthly efficiency summary for all machines"""
    try:
        from models.production import ShiftProduction, Machine
        from calendar import monthrange
        
        # Get month - default to current month
        year = request.args.get('year', get_local_now().year, type=int)
        month = request.args.get('month', get_local_now().month, type=int)
        
        month_start = date(year, month, 1)
        _, last_day = monthrange(year, month)
        month_end = date(year, month, last_day)
        
        # Get all shift productions for the month
        shift_records = ShiftProduction.query.filter(
            ShiftProduction.production_date >= month_start,
            ShiftProduction.production_date <= month_end
        ).all()
        
        # Group by machine
        machines_data = {}
        
        for sp in shift_records:
            machine_id = sp.machine_id
            if not machine_id:
                continue
                
            if machine_id not in machines_data:
                machines_data[machine_id] = {
                    'machine_id': machine_id,
                    'machine_name': sp.machine.name if sp.machine else f'Machine {machine_id}',
                    'machine_code': sp.machine.code if sp.machine else None,
                    'target_efficiency': int(sp.machine.target_efficiency) if sp.machine and sp.machine.target_efficiency else 60,
                    'weekly_data': {},
                    'total_grade_a': 0,
                    'total_output': 0,
                    'total_downtime': 0,
                    'total_idle': 0,
                    'total_machine_speed': 0,
                    'shift_count': 0
                }
            
            # Group by week number
            week_num = sp.production_date.isocalendar()[1]
            week_key = f'W{week_num}'
            if week_key not in machines_data[machine_id]['weekly_data']:
                machines_data[machine_id]['weekly_data'][week_key] = {
                    'grade_a': 0,
                    'output': 0,
                    'downtime': 0,
                    'idle': 0,
                    'machine_speed': 0,
                    'shift_count': 0,
                    'shifts_seen': {}
                }
            
            grade_a = int(sp.good_quantity) if sp.good_quantity else 0
            output = int(sp.actual_quantity) if sp.actual_quantity else 0
            downtime = int(sp.downtime_mesin or 0) + int(sp.downtime_operator or 0) + int(sp.downtime_material or 0) + int(sp.downtime_design or 0) + int(sp.downtime_others or 0)
            idle = int(sp.idle_time) if sp.idle_time else 0
            speed = int(sp.machine_speed) if sp.machine_speed else 0
            
            machines_data[machine_id]['weekly_data'][week_key]['grade_a'] += grade_a
            machines_data[machine_id]['weekly_data'][week_key]['output'] += output
            machines_data[machine_id]['weekly_data'][week_key]['downtime'] += downtime
            machines_data[machine_id]['weekly_data'][week_key]['idle'] += idle
            # Track unique shifts per date for proper speed averaging
            import re as _re
            _shift_num = 1
            if sp.shift:
                _sm = _re.search(r'(\d+)', str(sp.shift))
                if _sm:
                    _shift_num = int(_sm.group(1))
            _date_shift_key = f"{sp.production_date.isoformat()}_{_shift_num}"
            wk_shifts = machines_data[machine_id]['weekly_data'][week_key]['shifts_seen']
            if _date_shift_key not in wk_shifts:
                wk_shifts[_date_shift_key] = speed
            elif speed > wk_shifts[_date_shift_key]:
                wk_shifts[_date_shift_key] = speed
            machines_data[machine_id]['weekly_data'][week_key]['shift_count'] += 1
            
            machines_data[machine_id]['total_grade_a'] += grade_a
            machines_data[machine_id]['total_output'] += output
            machines_data[machine_id]['total_downtime'] += downtime
            machines_data[machine_id]['total_idle'] += idle
            machines_data[machine_id]['total_machine_speed'] += speed
            machines_data[machine_id]['shift_count'] += 1
        
        # Calculate monthly efficiency for each machine
        for machine_id in machines_data:
            data = machines_data[machine_id]
            
            # Calculate weekly efficiencies
            for week_key in data['weekly_data']:
                week_data = data['weekly_data'][week_key]
                # Average speed by unique shifts, not total records
                wk_shifts = week_data.get('shifts_seen', {})
                speed_per_min = sum(wk_shifts.values()) / len(wk_shifts) if wk_shifts else 0
                mrt = week_data['grade_a'] / speed_per_min if speed_per_min > 0 else 0
                total_time = mrt + week_data['downtime'] + week_data['idle']
                week_data['efficiency'] = round((mrt / total_time) * 100, 1) if total_time > 0 else 0
            
            # Calculate monthly average efficiency - use unique shifts across all weeks
            all_shifts = {}
            for wk in data['weekly_data']:
                for sk, spd in data['weekly_data'][wk].get('shifts_seen', {}).items():
                    all_shifts[sk] = spd
            speed_per_min = sum(all_shifts.values()) / len(all_shifts) if all_shifts else 0
            mrt = data['total_grade_a'] / speed_per_min if speed_per_min > 0 else 0
            total_time = mrt + data['total_downtime'] + data['total_idle']
            data['avg_efficiency'] = round((mrt / total_time) * 100, 1) if total_time > 0 else 0
            data['mrt'] = round(mrt, 1)
            data['total_time'] = round(total_time, 1)
            data['quality'] = round((data['total_grade_a'] / data['total_output']) * 100, 1) if data['total_output'] > 0 else 0
        
        result = list(machines_data.values())
        result.sort(key=lambda x: x['machine_name'])
        
        # === NEW: Collect products produced and downtime data ===
        products_produced = {}
        products_per_machine = {}  # {machine_name: {product_name: {quantity, grade_a, ...}}}
        all_downtime = {}
        weekly_summary = {}  # {week: {'output': X, 'downtime': Y, 'efficiency': Z}}
        
        for sp in shift_records:
            # Products produced
            product_name = None
            if sp.product:
                product_name = sp.product.name
            elif sp.work_order and sp.work_order.product:
                product_name = sp.work_order.product.name
            
            if product_name:
                if product_name not in products_produced:
                    products_produced[product_name] = {
                        'product_name': product_name,
                        'quantity': 0,
                        'grade_a': 0,
                        'grade_b': 0,
                        'grade_c': 0,
                        'machines': set()
                    }
                products_produced[product_name]['quantity'] += int(sp.actual_quantity) if sp.actual_quantity else 0
                products_produced[product_name]['grade_a'] += int(sp.good_quantity) if sp.good_quantity else 0
                products_produced[product_name]['grade_b'] += int(sp.rework_quantity) if sp.rework_quantity else 0
                products_produced[product_name]['grade_c'] += int(sp.reject_quantity) if sp.reject_quantity else 0
                if sp.machine:
                    products_produced[product_name]['machines'].add(sp.machine.name)
            
            # Parse downtime from issues field
            if sp.issues:
                issue_parts = sp.issues.split(';')
                for part in issue_parts:
                    part = part.strip()
                    if not part:
                        continue
                    match = re.match(r'(\d+)\s*menit\s*-\s*(.+)', part, re.IGNORECASE)
                    if match:
                        duration = int(match.group(1))
                        reason = match.group(2).strip()
                        reason = re.sub(r'\s*\[\w+\]\s*$', '', reason).strip()
                        
                        if reason not in all_downtime:
                            all_downtime[reason] = {'reason': reason, 'duration': 0, 'frequency': 0, 'category': detect_downtime_category(reason.lower())}
                        all_downtime[reason]['duration'] += duration
                        all_downtime[reason]['frequency'] += 1
                        
                # Track products per machine
                if sp.machine and product_name:
                    machine_name = sp.machine.name
                    if machine_name not in products_per_machine:
                        products_per_machine[machine_name] = {}
                    if product_name not in products_per_machine[machine_name]:
                        products_per_machine[machine_name][product_name] = {
                            'product_name': product_name,
                            'quantity': 0,
                            'grade_a': 0,
                            'grade_b': 0,
                            'grade_c': 0
                        }
                    products_per_machine[machine_name][product_name]['quantity'] += int(sp.actual_quantity) if sp.actual_quantity else 0
                    products_per_machine[machine_name][product_name]['grade_a'] += int(sp.good_quantity) if sp.good_quantity else 0
                    products_per_machine[machine_name][product_name]['grade_b'] += int(sp.rework_quantity) if sp.rework_quantity else 0
                    products_per_machine[machine_name][product_name]['grade_c'] += int(sp.reject_quantity) if sp.reject_quantity else 0
            
            # Weekly summary for chart
            week_num = sp.production_date.isocalendar()[1]
            week_key = f'W{week_num}'
            if week_key not in weekly_summary:
                weekly_summary[week_key] = {
                    'week': week_key,
                    'output': 0,
                    'grade_a': 0,
                    'downtime': 0,
                    'idle': 0,
                    'machine_speed': 0,
                    'shift_count': 0
                }
            weekly_summary[week_key]['output'] += int(sp.actual_quantity) if sp.actual_quantity else 0
            weekly_summary[week_key]['grade_a'] += int(sp.good_quantity) if sp.good_quantity else 0
            weekly_summary[week_key]['downtime'] += int(sp.downtime_mesin or 0) + int(sp.downtime_operator or 0) + int(sp.downtime_material or 0) + int(sp.downtime_design or 0) + int(sp.downtime_others or 0)
            weekly_summary[week_key]['idle'] += int(sp.idle_time) if sp.idle_time else 0
            weekly_summary[week_key]['machine_speed'] += int(sp.machine_speed) if sp.machine_speed else 0
            weekly_summary[week_key]['shift_count'] += 1
        
        # Calculate efficiency for weekly summary
        for week_key in weekly_summary:
            week = weekly_summary[week_key]
            speed_per_min = week['machine_speed'] / week['shift_count'] if week['shift_count'] > 0 else 0
            mrt = week['grade_a'] / speed_per_min if speed_per_min > 0 else 0
            total_time = mrt + week['downtime'] + week['idle']
            week['efficiency'] = round((mrt / total_time) * 100, 1) if total_time > 0 else 0
        
        # Convert products to list and sort
        products_list = []
        for p in products_produced.values():
            p['machines'] = list(p['machines'])
            products_list.append(p)
        products_list.sort(key=lambda x: x['quantity'], reverse=True)
        
        # Sort downtime by duration and get top 10
        downtime_list = list(all_downtime.values())
        downtime_list.sort(key=lambda x: x['duration'], reverse=True)
        top_10_downtime = downtime_list[:10]
        
        # Weekly chart data sorted by week
        chart_data = list(weekly_summary.values())
        chart_data.sort(key=lambda x: x['week'])
        
        # Pie chart data - downtime distribution
        pie_chart_data = [{'name': d['reason'][:20], 'value': d['duration']} for d in top_10_downtime]
        
        # Convert products_per_machine to list format
        products_per_machine_list = []
        for machine_name, products in products_per_machine.items():
            machine_products = list(products.values())
            machine_products.sort(key=lambda x: x['quantity'], reverse=True)
            products_per_machine_list.append({
                'machine_name': machine_name,
                'products': machine_products,
                'total_output': sum(p['quantity'] for p in machine_products),
                'total_grade_a': sum(p['grade_a'] for p in machine_products)
            })
        products_per_machine_list.sort(key=lambda x: x['machine_name'])
        
        # Month names in Indonesian
        month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 
                       'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        
        return jsonify({
            'year': year,
            'month': month,
            'month_name': month_names[month],
            'month_start': month_start.isoformat(),
            'month_end': month_end.isoformat(),
            'machines': result,
            'total_machines': len(result),
            'products_produced': products_list,
            'products_per_machine': products_per_machine_list,
            'all_downtime': downtime_list,
            'top_10_downtime': top_10_downtime,
            'chart_data': chart_data,
            'pie_chart_data': pie_chart_data
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/quality-objectives/production', methods=['GET'])
@jwt_required()
def get_production_quality_objectives():
    """Get production quality objectives - target vs actual per machine per month"""
    try:
        from models.production import ShiftProduction
        
        # Get year and month from params (default to current month)
        year = request.args.get('year', type=int) or get_local_today().year
        month = request.args.get('month', type=int) or get_local_today().month
        
        # Calculate month range
        month_start = date(year, month, 1)
        if month == 12:
            month_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(year, month + 1, 1) - timedelta(days=1)
        
        # Get machines that have production in this month (only active machines)
        active_machine_ids = db.session.query(ShiftProduction.machine_id).filter(
            ShiftProduction.production_date >= month_start,
            ShiftProduction.production_date <= month_end
        ).distinct().all()
        active_machine_ids = [m[0] for m in active_machine_ids]
        
        # Get machines sorted by name
        machines = Machine.query.filter(
            Machine.id.in_(active_machine_ids),
            Machine.is_active == True
        ).order_by(Machine.name).all()
        
        # Get manual monthly targets
        manual_targets = {t.machine_id: t.target_quantity for t in MachineMonthlyTarget.query.filter(
            MachineMonthlyTarget.year == year,
            MachineMonthlyTarget.month == month
        ).all()}
        
        result = []
        total_target = 0
        total_actual = 0
        machines_achieved = 0
        
        for machine in machines:
            # Get manual monthly target first, fallback to default calculation
            monthly_target = manual_targets.get(machine.id, 0)
            
            # If no manual target set, use default (0 means no target set)
            if monthly_target == 0:
                # Default: leave as 0 to indicate target not set
                monthly_target = 0
            
            # Get actual production from ShiftProduction
            shift_records = ShiftProduction.query.filter(
                ShiftProduction.machine_id == machine.id,
                ShiftProduction.production_date >= month_start,
                ShiftProduction.production_date <= month_end
            ).all()
            
            # Calculate totals
            actual_output = sum(int(sp.good_quantity or 0) for sp in shift_records)
            total_produced = sum(int(sp.actual_quantity or 0) for sp in shift_records)
            total_reject = sum(int(sp.reject_quantity or 0) for sp in shift_records)
            total_rework = sum(int(sp.rework_quantity or 0) for sp in shift_records)
            working_days_actual = len(set(sp.production_date for sp in shift_records))
            
            # Calculate achievement percentage
            achievement_pct = round((actual_output / monthly_target * 100), 1) if monthly_target > 0 else 0
            
            # Quality rate
            quality_rate = round((actual_output / total_produced * 100), 1) if total_produced > 0 else 0
            
            # Status
            is_achieved = achievement_pct >= 100
            status = 'Tercapai' if is_achieved else 'Tidak Tercapai'
            
            if is_achieved:
                machines_achieved += 1
            
            total_target += monthly_target
            total_actual += actual_output
            
            # Get average efficiency from shift records
            avg_efficiency = 0
            if shift_records:
                efficiencies = [float(sp.efficiency_rate or 0) for sp in shift_records if sp.efficiency_rate]
                avg_efficiency = round(sum(efficiencies) / len(efficiencies), 1) if efficiencies else 0
            
            result.append({
                'machine_id': machine.id,
                'machine_name': machine.name,
                'machine_code': machine.code if hasattr(machine, 'code') else None,
                'target_monthly': int(monthly_target),
                'actual_output': actual_output,
                'total_produced': total_produced,
                'total_reject': total_reject,
                'total_rework': total_rework,
                'achievement_pct': achievement_pct,
                'quality_rate': quality_rate,
                'avg_efficiency': avg_efficiency,
                'working_days': working_days_actual,
                'is_achieved': is_achieved,
                'status': status,
                'gap': actual_output - int(monthly_target)
            })
        
        # Sort by achievement percentage descending
        result.sort(key=lambda x: x['achievement_pct'], reverse=True)
        
        # Calculate overall achievement
        overall_achievement = round((total_actual / total_target * 100), 1) if total_target > 0 else 0
        
        # Month names in Indonesian
        month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 
                       'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        
        return jsonify({
            'department': 'Produksi',
            'year': year,
            'month': month,
            'month_name': month_names[month],
            'period': f'{month_names[month]} {year}',
            'summary': {
                'total_machines': len(machines),
                'machines_achieved': machines_achieved,
                'machines_not_achieved': len(machines) - machines_achieved,
                'achievement_rate': round((machines_achieved / len(machines) * 100), 1) if machines else 0,
                'total_target': int(total_target),
                'total_actual': total_actual,
                'overall_achievement_pct': overall_achievement,
                'overall_status': 'Tercapai' if overall_achievement >= 100 else 'Tidak Tercapai'
            },
            'machines': result
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/machine-monthly-targets', methods=['GET'])
@jwt_required()
def get_machine_monthly_targets():
    """Get all monthly targets for a specific period"""
    try:
        year = request.args.get('year', type=int) or get_local_today().year
        month = request.args.get('month', type=int) or get_local_today().month
        
        targets = MachineMonthlyTarget.query.filter(
            MachineMonthlyTarget.year == year,
            MachineMonthlyTarget.month == month
        ).all()
        
        result = []
        for t in targets:
            result.append({
                'id': t.id,
                'machine_id': t.machine_id,
                'machine_name': t.machine.name if t.machine else None,
                'year': t.year,
                'month': t.month,
                'target_quantity': t.target_quantity,
                'notes': t.notes
            })
        
        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/machine-monthly-targets', methods=['POST'])
@jwt_required()
def set_machine_monthly_target():
    """Set or update monthly target for a machine"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json()
        
        machine_id = data.get('machine_id')
        year = data.get('year')
        month = data.get('month')
        target_quantity = data.get('target_quantity', 0)
        notes = data.get('notes')
        
        # Check if target exists
        existing = MachineMonthlyTarget.query.filter(
            MachineMonthlyTarget.machine_id == machine_id,
            MachineMonthlyTarget.year == year,
            MachineMonthlyTarget.month == month
        ).first()
        
        if existing:
            existing.target_quantity = target_quantity
            existing.notes = notes
            existing.updated_at = get_local_now()
        else:
            new_target = MachineMonthlyTarget(
                machine_id=machine_id,
                year=year,
                month=month,
                target_quantity=target_quantity,
                notes=notes,
                created_by=user_id
            )
            db.session.add(new_target)
        
        db.session.commit()
        return jsonify({'message': 'Target saved successfully'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/machine-monthly-targets/bulk', methods=['POST'])
@jwt_required()
def set_bulk_machine_monthly_targets():
    """Set multiple monthly targets at once"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json()
        targets = data.get('targets', [])
        
        for t in targets:
            machine_id = t.get('machine_id')
            year = t.get('year')
            month = t.get('month')
            target_quantity = t.get('target_quantity', 0)
            
            existing = MachineMonthlyTarget.query.filter(
                MachineMonthlyTarget.machine_id == machine_id,
                MachineMonthlyTarget.year == year,
                MachineMonthlyTarget.month == month
            ).first()
            
            if existing:
                existing.target_quantity = target_quantity
                existing.updated_at = get_local_now()
            else:
                new_target = MachineMonthlyTarget(
                    machine_id=machine_id,
                    year=year,
                    month=month,
                    target_quantity=target_quantity,
                    created_by=user_id
                )
                db.session.add(new_target)
        
        db.session.commit()
        return jsonify({'message': f'{len(targets)} targets saved successfully'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/machine-downtime-analysis', methods=['GET'])
@jwt_required()
def get_machine_downtime_analysis():
    """Get downtime analysis per machine with top downtime reasons"""
    try:
        from models.production import ShiftProduction
        
        year = request.args.get('year', type=int) or get_local_today().year
        month = request.args.get('month', type=int) or get_local_today().month
        machine_id = request.args.get('machine_id', type=int)
        
        # Calculate month range
        month_start = date(year, month, 1)
        if month == 12:
            month_end = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(year, month + 1, 1) - timedelta(days=1)
        
        # Build query
        query = ShiftProduction.query.filter(
            ShiftProduction.production_date >= month_start,
            ShiftProduction.production_date <= month_end
        )
        
        if machine_id:
            query = query.filter(ShiftProduction.machine_id == machine_id)
        
        shift_records = query.all()
        
        # Group by machine
        machines_data = {}
        
        for sp in shift_records:
            mid = sp.machine_id
            if mid not in machines_data:
                machines_data[mid] = {
                    'machine_id': mid,
                    'machine_name': sp.machine.name if sp.machine else f'Machine {mid}',
                    'downtime_reasons': {},
                    'downtime_by_category': {
                        'mesin': 0, 'operator': 0, 'material': 0, 
                        'design': 0, 'idle': 0, 'others': 0
                    },
                    'total_downtime': 0,
                    'total_runtime': 0
                }
            
            # Parse issues/downtime from ShiftProduction
            if sp.issues:
                # Split by semicolon or comma
                import re
                issues_list = re.split(r'[;,]', sp.issues)
                issues_list = [i.strip() for i in issues_list if i.strip()]
                for issue in issues_list:
                    # Parse format: "Xm - reason [category]" or "reason (Xm)" or just "reason"
                    # Format 1: "15 menit - Sanitasi kecil [design]"
                    match1 = re.match(r'(\d+)\s*menit?\s*-\s*(.+?)(?:\s*\[.+?\])?\s*$', issue)
                    # Format 2: "reason (Xm)"
                    match2 = re.match(r'(.+?)\s*\((\d+)m\)', issue)
                    
                    if match1:
                        minutes = int(match1.group(1))
                        reason = match1.group(2).strip()
                    elif match2:
                        reason = match2.group(1).strip()
                        minutes = int(match2.group(2))
                    else:
                        reason = issue.strip()
                        # Try to extract category tag
                        reason = re.sub(r'\s*\[.+?\]\s*$', '', reason)
                        minutes = 0
                    
                    if reason and len(reason) > 2:
                        if reason in machines_data[mid]['downtime_reasons']:
                            machines_data[mid]['downtime_reasons'][reason]['count'] += 1
                            machines_data[mid]['downtime_reasons'][reason]['minutes'] += minutes
                        else:
                            machines_data[mid]['downtime_reasons'][reason] = {
                                'count': 1,
                                'minutes': minutes
                            }
            
            # Add category totals
            machines_data[mid]['downtime_by_category']['mesin'] += int(sp.downtime_mesin or 0)
            machines_data[mid]['downtime_by_category']['operator'] += int(sp.downtime_operator or 0)
            machines_data[mid]['downtime_by_category']['material'] += int(sp.downtime_material or 0)
            machines_data[mid]['downtime_by_category']['design'] += int(sp.downtime_design or 0)
            machines_data[mid]['downtime_by_category']['idle'] += int(sp.idle_time or 0)
            machines_data[mid]['downtime_by_category']['others'] += int(sp.downtime_others or 0)
            machines_data[mid]['total_downtime'] += int(sp.downtime_minutes or 0)
            machines_data[mid]['total_runtime'] += int(sp.actual_runtime or 0)
        
        # Keywords to exclude from Top 3 (design change, idle, biological breaks)
        excluded_keywords = [
            # Design change - not real downtime
            'ganti stiker', 'ganti packaging', 'ganti label', 'ganti karton',
            'repack', 'repacking', 'ganti kemasan', 'change over', 'changeover',
            'ganti produk', 'ganti order', 'setting mc', 'setting mesin', 'sanitasi dan setting',
            # Idle time - tracked separately
            'tunggu', 'menunggu', 'nunggu', 'habis', 'waiting', 'standby', 'idle',
            # Biological breaks
            'istirahat', 'sholat', 'makan', 'toilet'
        ]
        
        # Process and sort downtime reasons
        result = []
        for mid, data in machines_data.items():
            # Filter out excluded keywords
            filtered_reasons = {
                reason: stats for reason, stats in data['downtime_reasons'].items()
                if not any(kw in reason.lower() for kw in excluded_keywords)
            }
            
            # Sort by count (frequency) descending
            sorted_reasons = sorted(
                filtered_reasons.items(),
                key=lambda x: (x[1]['count'], x[1]['minutes']),
                reverse=True
            )
            
            top_3 = []
            for reason, stats in sorted_reasons[:3]:
                pct = round((stats['minutes'] / data['total_downtime'] * 100), 1) if data['total_downtime'] > 0 else 0
                top_3.append({
                    'reason': reason,
                    'count': stats['count'],
                    'minutes': stats['minutes'],
                    'percentage': pct
                })
            
            # Calculate category percentages
            category_chart = []
            for cat, mins in data['downtime_by_category'].items():
                pct = round((mins / data['total_downtime'] * 100), 1) if data['total_downtime'] > 0 else 0
                if mins > 0:
                    category_chart.append({
                        'category': cat,
                        'minutes': mins,
                        'percentage': pct
                    })
            
            result.append({
                'machine_id': data['machine_id'],
                'machine_name': data['machine_name'],
                'top_3_downtime': top_3,
                'downtime_by_category': category_chart,
                'total_downtime': data['total_downtime'],
                'total_runtime': data['total_runtime']
            })
        
        # Sort by machine name
        result.sort(key=lambda x: x['machine_name'])
        
        month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 
                       'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        
        return jsonify({
            'year': year,
            'month': month,
            'month_name': month_names[month],
            'period': f'{month_names[month]} {year}',
            'machines': result
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/downtime-root-causes', methods=['GET'])
@jwt_required()
def get_downtime_root_causes():
    """Get root cause analysis records"""
    try:
        year = request.args.get('year', type=int) or get_local_today().year
        month = request.args.get('month', type=int) or get_local_today().month
        machine_id = request.args.get('machine_id', type=int)
        
        query = DowntimeRootCause.query.filter(
            DowntimeRootCause.year == year,
            DowntimeRootCause.month == month
        )
        
        if machine_id:
            query = query.filter(DowntimeRootCause.machine_id == machine_id)
        
        records = query.order_by(DowntimeRootCause.total_minutes.desc()).all()
        
        result = []
        for r in records:
            result.append({
                'id': r.id,
                'machine_id': r.machine_id,
                'machine_name': r.machine.name if r.machine else None,
                'problem': r.problem,
                'category': r.category,
                'occurrence_count': r.occurrence_count,
                'total_minutes': r.total_minutes,
                'percentage': float(r.percentage) if r.percentage else 0,
                'root_cause': r.root_cause,
                'corrective_action': r.corrective_action,
                'preventive_action': r.preventive_action,
                'status': r.status
            })
        
        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/downtime-root-causes', methods=['POST'])
@jwt_required()
def create_downtime_root_cause():
    """Create or update root cause analysis"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json()
        
        record_id = data.get('id')
        
        if record_id:
            # Update existing
            record = DowntimeRootCause.query.get(record_id)
            if not record:
                return jsonify({'error': 'Record not found'}), 404
        else:
            # Create new
            record = DowntimeRootCause(
                machine_id=data.get('machine_id'),
                year=data.get('year'),
                month=data.get('month'),
                created_by=user_id
            )
            db.session.add(record)
        
        record.problem = data.get('problem')
        record.category = data.get('category')
        record.occurrence_count = data.get('occurrence_count', 1)
        record.total_minutes = data.get('total_minutes', 0)
        record.percentage = data.get('percentage', 0)
        record.root_cause = data.get('root_cause')
        record.corrective_action = data.get('corrective_action')
        record.preventive_action = data.get('preventive_action')
        record.status = data.get('status', 'open')
        record.updated_at = get_local_now()
        
        db.session.commit()
        return jsonify({'message': 'Root cause analysis saved', 'id': record.id}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@oee_bp.route('/downtime-root-causes/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_downtime_root_cause(id):
    """Delete root cause analysis"""
    try:
        record = DowntimeRootCause.query.get(id)
        if not record:
            return jsonify({'error': 'Record not found'}), 404
        
        db.session.delete(record)
        db.session.commit()
        return jsonify({'message': 'Record deleted'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
