from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from models import db, SalesOrder, WorkOrder, Inventory, Product, Customer, WasteRecord, MaintenanceRecord, QualityTest
from models.hr import Employee, Attendance
from models.hr_extended import PayrollRecord
from models.finance import Invoice
from models.production import ProductionRecord
from utils.i18n import success_response, error_response, get_message
from sqlalchemy import func, extract
from datetime import datetime, timedelta
from calendar import monthrange
import io
from utils.timezone import get_local_now, get_local_today

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/sales', methods=['GET'])
@jwt_required()
def sales_report():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = SalesOrder.query
        
        if start_date:
            query = query.filter(SalesOrder.order_date >= datetime.fromisoformat(start_date))
        if end_date:
            query = query.filter(SalesOrder.order_date <= datetime.fromisoformat(end_date))
        
        orders = query.all()
        
        total_orders = len(orders)
        total_amount = sum(float(o.total_amount) for o in orders)
        
        return jsonify({
            'total_orders': total_orders,
            'total_amount': total_amount,
            'orders': [{
                'order_number': o.order_number,
                'customer_name': o.customer.company_name,
                'order_date': o.order_date.isoformat(),
                'total_amount': float(o.total_amount),
                'status': o.status
            } for o in orders]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/production', methods=['GET'])
@jwt_required()
def production_report():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = WorkOrder.query.filter_by(status='completed')
        
        if start_date:
            query = query.filter(WorkOrder.actual_end_date >= datetime.fromisoformat(start_date))
        if end_date:
            query = query.filter(WorkOrder.actual_end_date <= datetime.fromisoformat(end_date))
        
        work_orders = query.all()
        
        total_quantity = sum(float(wo.quantity_produced) for wo in work_orders)
        total_good = sum(float(wo.quantity_good) for wo in work_orders)
        total_scrap = sum(float(wo.quantity_scrap) for wo in work_orders)
        
        return jsonify({
            'total_work_orders': len(work_orders),
            'total_quantity_produced': total_quantity,
            'total_good': total_good,
            'total_scrap': total_scrap,
            'efficiency': (total_good / total_quantity * 100) if total_quantity > 0 else 0,
            'work_orders': [{
                'wo_number': wo.wo_number,
                'product_name': wo.product.name,
                'quantity_produced': float(wo.quantity_produced),
                'quantity_good': float(wo.quantity_good),
                'status': wo.status
            } for wo in work_orders]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/inventory', methods=['GET'])
@jwt_required()
def inventory_report():
    try:
        results = db.session.query(
            Product.id,
            Product.code,
            Product.name,
            Product.primary_uom,
            func.sum(Inventory.quantity).label('total_quantity'),
            func.sum(Inventory.available_quantity).label('available_quantity')
        ).join(Inventory).group_by(Product.id, Product.code, Product.name, Product.primary_uom).all()
        
        return jsonify({
            'inventory': [{
                'product_code': r.code,
                'product_name': r.name,
                'total_quantity': float(r.total_quantity or 0),
                'available_quantity': float(r.available_quantity or 0),
                'uom': r.primary_uom
            } for r in results]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/waste', methods=['GET'])
@jwt_required()
def waste_report():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = WasteRecord.query
        
        if start_date:
            query = query.filter(WasteRecord.waste_date >= datetime.fromisoformat(start_date))
        if end_date:
            query = query.filter(WasteRecord.waste_date <= datetime.fromisoformat(end_date))
        
        records = query.all()
        
        total_quantity = sum(float(r.quantity) for r in records if r.quantity)
        total_weight = sum(float(r.weight_kg) for r in records if r.weight_kg)
        total_value = sum(float(r.estimated_value) for r in records if r.estimated_value)
        
        # Group by hazard level
        hazard_stats = {}
        for record in records:
            level = record.hazard_level or 'unknown'
            if level not in hazard_stats:
                hazard_stats[level] = {'count': 0, 'quantity': 0}
            hazard_stats[level]['count'] += 1
            hazard_stats[level]['quantity'] += float(record.quantity or 0)
        
        return jsonify({
            'total_records': len(records),
            'total_quantity': total_quantity,
            'total_weight_kg': total_weight,
            'total_estimated_value': total_value,
            'hazard_level_breakdown': hazard_stats,
            'records': [{
                'record_number': r.record_number,
                'waste_date': r.waste_date.isoformat() if r.waste_date else None,
                'category': r.category.name if r.category else None,
                'quantity': float(r.quantity) if r.quantity else 0,
                'hazard_level': r.hazard_level,
                'disposal_method': r.disposal_method
            } for r in records]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/maintenance', methods=['GET'])
@jwt_required()
def maintenance_report():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = MaintenanceRecord.query
        
        if start_date:
            query = query.filter(MaintenanceRecord.maintenance_date >= datetime.fromisoformat(start_date))
        if end_date:
            query = query.filter(MaintenanceRecord.maintenance_date <= datetime.fromisoformat(end_date))
        
        records = query.all()
        
        total_cost = sum(float(r.cost) for r in records if r.cost)
        completed_count = len([r for r in records if r.status == 'completed'])
        
        # Group by maintenance type
        type_stats = {}
        for record in records:
            mtype = record.maintenance_type or 'unknown'
            if mtype not in type_stats:
                type_stats[mtype] = {'count': 0, 'cost': 0}
            type_stats[mtype]['count'] += 1
            type_stats[mtype]['cost'] += float(record.cost or 0)
        
        return jsonify({
            'total_records': len(records),
            'completed_count': completed_count,
            'total_cost': total_cost,
            'completion_rate': (completed_count / len(records) * 100) if records else 0,
            'maintenance_type_breakdown': type_stats,
            'records': [{
                'record_number': r.record_number,
                'machine_name': r.machine.name if r.machine else None,
                'maintenance_date': r.maintenance_date.isoformat() if r.maintenance_date else None,
                'maintenance_type': r.maintenance_type,
                'status': r.status,
                'cost': float(r.cost) if r.cost else 0
            } for r in records]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/quality', methods=['GET'])
@jwt_required()
def quality_report():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = QualityTest.query
        
        if start_date:
            query = query.filter(QualityTest.test_date >= datetime.fromisoformat(start_date))
        if end_date:
            query = query.filter(QualityTest.test_date <= datetime.fromisoformat(end_date))
        
        tests = query.all()
        
        passed_count = len([t for t in tests if t.result == 'pass'])
        failed_count = len([t for t in tests if t.result == 'fail'])
        
        return jsonify({
            'total_tests': len(tests),
            'passed_count': passed_count,
            'failed_count': failed_count,
            'pass_rate': (passed_count / len(tests) * 100) if tests else 0,
            'tests': [{
                'test_number': t.test_number,
                'product_name': t.product.name if t.product else None,
                'test_date': t.test_date.isoformat() if t.test_date else None,
                'result': t.result,
                'defect_count': t.defect_count or 0
            } for t in tests]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/hr', methods=['GET'])
@jwt_required()
def hr_report():
    try:
        # Employee statistics
        total_employees = Employee.query.filter_by(is_active=True).count()
        
        # Get department breakdown
        dept_stats = db.session.query(
            Employee.department,
            func.count(Employee.id).label('count')
        ).filter(Employee.is_active == True).group_by(Employee.department).all()
        
        # Get attendance for current month
        today = get_local_now()
        first_day = today.replace(day=1)
        
        attendance_records = Attendance.query.filter(
            Attendance.date >= first_day,
            Attendance.date <= today
        ).all()
        
        present_count = len([a for a in attendance_records if a.status == 'present'])
        absent_count = len([a for a in attendance_records if a.status == 'absent'])
        late_count = len([a for a in attendance_records if a.status == 'late'])
        
        return jsonify({
            'total_employees': total_employees,
            'department_breakdown': {d.department: d.count for d in dept_stats if d.department},
            'attendance_this_month': {
                'total_records': len(attendance_records),
                'present': present_count,
                'absent': absent_count,
                'late': late_count,
                'attendance_rate': (present_count / len(attendance_records) * 100) if attendance_records else 0
            }
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/financial', methods=['GET'])
@jwt_required()
def financial_report():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Default to current month
        if not start_date:
            start_date = get_local_now().replace(day=1).isoformat()
        if not end_date:
            end_date = get_local_now().isoformat()
        
        # Get invoices
        invoices = Invoice.query.filter(
            Invoice.invoice_date >= datetime.fromisoformat(start_date),
            Invoice.invoice_date <= datetime.fromisoformat(end_date)
        ).all()
        
        total_invoiced = sum(float(i.total_amount) for i in invoices if i.total_amount)
        total_paid = sum(float(i.paid_amount) for i in invoices if i.paid_amount)
        total_outstanding = total_invoiced - total_paid
        
        # Invoice status breakdown
        status_breakdown = {}
        for inv in invoices:
            status = inv.status or 'unknown'
            if status not in status_breakdown:
                status_breakdown[status] = {'count': 0, 'amount': 0}
            status_breakdown[status]['count'] += 1
            status_breakdown[status]['amount'] += float(inv.total_amount or 0)
        
        return jsonify({
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'summary': {
                'total_invoiced': total_invoiced,
                'total_paid': total_paid,
                'total_outstanding': total_outstanding,
                'collection_rate': (total_paid / total_invoiced * 100) if total_invoiced > 0 else 0
            },
            'status_breakdown': status_breakdown,
            'invoice_count': len(invoices)
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/generate/<report_type>', methods=['POST'])
@jwt_required()
def generate_report(report_type):
    try:
        data = request.get_json()
        filters = data.get('filters', {})
        fields = data.get('fields', [])
        
        # Route to appropriate report function based on type
        if report_type == 'sales-summary':
            return sales_report()
        elif report_type == 'production-efficiency':
            return production_report()
        elif report_type == 'inventory-status':
            return inventory_report()
        elif report_type == 'waste-management':
            return waste_report()
        elif report_type == 'maintenance-schedule':
            return maintenance_report()
        elif report_type == 'quality-metrics':
            return quality_report()
        elif report_type == 'hr-analytics':
            return hr_report()
        elif report_type == 'financial-summary':
            return financial_report()
        else:
            return jsonify(error_response('api.error', error_code=400)), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/dashboard-summary', methods=['GET'])
@jwt_required()
def dashboard_summary():
    try:
        # Get data for the last 30 days
        end_date = get_local_now()
        start_date = end_date - timedelta(days=30)
        
        # Sales summary
        sales_orders = SalesOrder.query.filter(
            SalesOrder.order_date >= start_date,
            SalesOrder.order_date <= end_date
        ).all()
        
        # Production summary
        work_orders = WorkOrder.query.filter(
            WorkOrder.actual_end_date >= start_date,
            WorkOrder.actual_end_date <= end_date,
            WorkOrder.status == 'completed'
        ).all()
        
        # Quality summary
        quality_tests = QualityTest.query.filter(
            QualityTest.test_date >= start_date,
            QualityTest.test_date <= end_date
        ).all()
        
        # Waste summary
        waste_records = WasteRecord.query.filter(
            WasteRecord.waste_date >= start_date,
            WasteRecord.waste_date <= end_date
        ).all()
        
        return jsonify({
            'period': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat()
            },
            'sales': {
                'total_orders': len(sales_orders),
                'total_amount': sum(float(o.total_amount) for o in sales_orders)
            },
            'production': {
                'total_work_orders': len(work_orders),
                'total_quantity': sum(float(wo.quantity_produced) for wo in work_orders)
            },
            'quality': {
                'total_tests': len(quality_tests),
                'pass_rate': (len([t for t in quality_tests if t.result == 'pass']) / len(quality_tests) * 100) if quality_tests else 0
            },
            'waste': {
                'total_records': len(waste_records),
                'total_quantity': sum(float(r.quantity) for r in waste_records if r.quantity)
            }
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/production-by-product', methods=['GET'])
@jwt_required()
def production_by_product_report():
    """Get production report grouped by product and period"""
    try:
        period_type = request.args.get('period_type', 'monthly')  # daily, weekly, monthly, yearly
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        product_id = request.args.get('product_id', type=int)
        
        # Parse dates
        if start_date_str:
            start_date = datetime.fromisoformat(start_date_str)
        else:
            start_date = get_local_now() - timedelta(days=30)
        
        if end_date_str:
            end_date = datetime.fromisoformat(end_date_str)
        else:
            end_date = get_local_now()
        
        # Ensure end_date includes the full day
        end_date = end_date.replace(hour=23, minute=59, second=59)
        
        # Build base query for production records
        # Use COALESCE to prefer ProductionRecord.product_id, fallback to WorkOrder.product_id
        query = db.session.query(
            ProductionRecord.production_date,
            ProductionRecord.work_order_id,
            func.coalesce(ProductionRecord.product_id, WorkOrder.product_id).label('product_id'),
            Product.code.label('product_code'),
            Product.name.label('product_name'),
            ProductionRecord.quantity_produced,
            ProductionRecord.quantity_good,
            ProductionRecord.quantity_reject,
            ProductionRecord.waste_kg
        ).join(
            WorkOrder, ProductionRecord.work_order_id == WorkOrder.id
        ).outerjoin(
            Product, func.coalesce(ProductionRecord.product_id, WorkOrder.product_id) == Product.id
        ).filter(
            ProductionRecord.production_date >= start_date.date(),
            ProductionRecord.production_date <= end_date.date()
        )
        
        if product_id:
            query = query.filter(
                db.or_(
                    ProductionRecord.product_id == product_id,
                    db.and_(ProductionRecord.product_id.is_(None), WorkOrder.product_id == product_id)
                )
            )
        
        records = query.all()
        
        # Group by period
        periods_data = {}
        
        for record in records:
            # Determine period key based on period_type
            prod_date = record.production_date
            if period_type == 'daily':
                period_key = prod_date.strftime('%Y-%m-%d')
                period_label = prod_date.strftime('%d %b %Y')
            elif period_type == 'weekly':
                # Get week start (Monday)
                week_start = prod_date - timedelta(days=prod_date.weekday())
                period_key = week_start.strftime('%Y-W%W')
                week_end = week_start + timedelta(days=6)
                period_label = f"{week_start.strftime('%d %b')} - {week_end.strftime('%d %b %Y')}"
            elif period_type == 'monthly':
                period_key = prod_date.strftime('%Y-%m')
                period_label = prod_date.strftime('%B %Y')
            else:  # yearly
                period_key = prod_date.strftime('%Y')
                period_label = f"Tahun {prod_date.strftime('%Y')}"
            
            # Initialize period if not exists
            if period_key not in periods_data:
                periods_data[period_key] = {
                    'period': period_key,
                    'period_label': period_label,
                    'total_produced': 0,
                    'total_good': 0,
                    'total_reject': 0,
                    'products': {}
                }
            
            period = periods_data[period_key]
            
            # Add to period totals
            qty_produced = float(record.quantity_produced or 0)
            qty_good = float(record.quantity_good or 0)
            qty_reject = float(record.quantity_reject or 0)
            waste = float(record.waste_kg or 0)
            
            period['total_produced'] += qty_produced
            period['total_good'] += qty_good
            period['total_reject'] += qty_reject
            
            # Group by product within period
            prod_id = record.product_id
            if prod_id not in period['products']:
                period['products'][prod_id] = {
                    'product_id': prod_id,
                    'product_code': record.product_code,
                    'product_name': record.product_name,
                    'total_produced': 0,
                    'total_good': 0,
                    'total_reject': 0,
                    'total_waste': 0,
                    'work_order_ids': set()
                }
            
            prod_data = period['products'][prod_id]
            prod_data['total_produced'] += qty_produced
            prod_data['total_good'] += qty_good
            prod_data['total_reject'] += qty_reject
            prod_data['total_waste'] += waste
            prod_data['work_order_ids'].add(record.work_order_id)
        
        # Convert to list and calculate rates
        result = []
        for period_key in sorted(periods_data.keys(), reverse=True):
            period = periods_data[period_key]
            
            # Convert products dict to list
            products_list = []
            for prod_data in period['products'].values():
                prod_data['work_order_count'] = len(prod_data['work_order_ids'])
                del prod_data['work_order_ids']  # Remove set before JSON serialization
                prod_data['reject_rate'] = (prod_data['total_reject'] / prod_data['total_produced'] * 100) if prod_data['total_produced'] > 0 else 0
                products_list.append(prod_data)
            
            # Sort products by total_produced descending
            products_list.sort(key=lambda x: x['total_produced'], reverse=True)
            period['products'] = products_list
            
            result.append(period)
        
        return jsonify({
            'periods': result,
            'period_type': period_type,
            'start_date': start_date.date().isoformat(),
            'end_date': end_date.date().isoformat()
        }), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/production-by-product/export', methods=['GET'])
@jwt_required()
def export_production_by_product():
    """Export production by product report to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        period_type = request.args.get('period_type', 'monthly')
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        product_id = request.args.get('product_id', type=int)
        
        # Parse dates
        if start_date_str:
            start_date = datetime.fromisoformat(start_date_str)
        else:
            start_date = get_local_now() - timedelta(days=30)
        
        if end_date_str:
            end_date = datetime.fromisoformat(end_date_str)
        else:
            end_date = get_local_now()
        
        end_date = end_date.replace(hour=23, minute=59, second=59)
        
        # Get data - use COALESCE to prefer ProductionRecord.product_id
        query = db.session.query(
            ProductionRecord.production_date,
            func.coalesce(ProductionRecord.product_id, WorkOrder.product_id).label('product_id'),
            Product.code.label('product_code'),
            Product.name.label('product_name'),
            func.sum(ProductionRecord.quantity_produced).label('total_produced'),
            func.sum(ProductionRecord.quantity_good).label('total_good'),
            func.sum(ProductionRecord.quantity_reject).label('total_reject'),
            func.sum(ProductionRecord.waste_kg).label('total_waste'),
            func.count(func.distinct(ProductionRecord.work_order_id)).label('wo_count')
        ).join(
            WorkOrder, ProductionRecord.work_order_id == WorkOrder.id
        ).outerjoin(
            Product, func.coalesce(ProductionRecord.product_id, WorkOrder.product_id) == Product.id
        ).filter(
            ProductionRecord.production_date >= start_date.date(),
            ProductionRecord.production_date <= end_date.date()
        )
        
        if product_id:
            query = query.filter(
                db.or_(
                    ProductionRecord.product_id == product_id,
                    db.and_(ProductionRecord.product_id.is_(None), WorkOrder.product_id == product_id)
                )
            )
        
        # Group by product
        query = query.group_by(
            func.coalesce(ProductionRecord.product_id, WorkOrder.product_id),
            Product.code,
            Product.name
        ).order_by(func.sum(ProductionRecord.quantity_produced).desc())
        
        records = query.all()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produksi per Produk"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Laporan Produksi per Produk ({period_type.title()})"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Periode: {start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['No', 'Kode Produk', 'Nama Produk', 'Total Produksi', 'Baik', 'Reject', 'Waste (kg)', 'Reject Rate (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        total_produced = 0
        total_good = 0
        total_reject = 0
        total_waste = 0
        
        for idx, record in enumerate(records, 1):
            row = idx + 4
            produced = float(record.total_produced or 0)
            good = float(record.total_good or 0)
            reject = float(record.total_reject or 0)
            waste = float(record.total_waste or 0)
            reject_rate = (reject / produced * 100) if produced > 0 else 0
            
            total_produced += produced
            total_good += good
            total_reject += reject
            total_waste += waste
            
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=record.product_code).border = border
            ws.cell(row=row, column=3, value=record.product_name).border = border
            ws.cell(row=row, column=4, value=produced).border = border
            ws.cell(row=row, column=5, value=good).border = border
            ws.cell(row=row, column=6, value=reject).border = border
            ws.cell(row=row, column=7, value=waste).border = border
            ws.cell(row=row, column=8, value=round(reject_rate, 2)).border = border
        
        # Total row
        total_row = len(records) + 5
        ws.cell(row=total_row, column=1, value='').border = border
        ws.cell(row=total_row, column=2, value='TOTAL').font = Font(bold=True)
        ws.cell(row=total_row, column=2).border = border
        ws.cell(row=total_row, column=3, value='').border = border
        ws.cell(row=total_row, column=4, value=total_produced).font = Font(bold=True)
        ws.cell(row=total_row, column=4).border = border
        ws.cell(row=total_row, column=5, value=total_good).font = Font(bold=True)
        ws.cell(row=total_row, column=5).border = border
        ws.cell(row=total_row, column=6, value=total_reject).font = Font(bold=True)
        ws.cell(row=total_row, column=6).border = border
        ws.cell(row=total_row, column=7, value=total_waste).font = Font(bold=True)
        ws.cell(row=total_row, column=7).border = border
        total_reject_rate = (total_reject / total_produced * 100) if total_produced > 0 else 0
        ws.cell(row=total_row, column=8, value=round(total_reject_rate, 2)).font = Font(bold=True)
        ws.cell(row=total_row, column=8).border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"production_report_{period_type}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
