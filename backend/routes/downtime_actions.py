"""
Routes for Downtime Action Items - Root Cause & Follow Up Tracking
"""
from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from models.production import DowntimeActionItem, Machine, ShiftProduction
from models.product import Product
from models.user import User
from models import db
from datetime import datetime, timedelta
from sqlalchemy import func, and_, or_
import calendar
import logging
import traceback

logger = logging.getLogger(__name__)

downtime_actions_bp = Blueprint('downtime_actions', __name__)

# Kategori downtime yang dianggap UNPLANNED
UNPLANNED_CATEGORIES = ['mesin', 'operator', 'material']
# Kategori PLANNED yang tidak perlu action item
PLANNED_CATEGORIES = ['idle', 'design']


def get_week_number(date):
    """Get week number in month (1-5)"""
    day = date.day
    if day <= 7:
        return 1
    elif day <= 14:
        return 2
    elif day <= 21:
        return 3
    elif day <= 28:
        return 4
    else:
        return 5


@downtime_actions_bp.route('/action-items', methods=['GET'])
@jwt_required()
def get_action_items():
    """Get action items with filters"""
    try:
        # Filters
        machine_id = request.args.get('machine_id', type=int)
        product_id = request.args.get('product_id', type=int)
        week_number = request.args.get('week_number', type=int)
        month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)
        status = request.args.get('status')
        
        query = DowntimeActionItem.query
        
        if machine_id:
            query = query.filter_by(machine_id=machine_id)
        if product_id:
            query = query.filter_by(product_id=product_id)
        if week_number:
            query = query.filter_by(week_number=week_number)
        if month:
            query = query.filter_by(month=month)
        if year:
            query = query.filter_by(year=year)
        if status:
            query = query.filter_by(status=status)
        
        # Order by duration desc (most critical first)
        action_items = query.order_by(DowntimeActionItem.total_duration.desc()).all()
        
        return jsonify({
            'action_items': [item.to_dict() for item in action_items]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@downtime_actions_bp.route('/action-items/<int:item_id>', methods=['GET'])
@jwt_required()
def get_action_item(item_id):
    """Get single action item"""
    try:
        item = DowntimeActionItem.query.get_or_404(item_id)
        return jsonify(item.to_dict()), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@downtime_actions_bp.route('/action-items', methods=['POST'])
@jwt_required()
def create_action_item():
    """Create new action item"""
    try:
        data = request.get_json()
        current_user_id = get_jwt_identity()
        
        # Validate required fields
        required = ['downtime_reason', 'machine_id', 'week_number', 'year', 'month', 'total_duration']
        for field in required:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Create action item
        action_item = DowntimeActionItem(
            downtime_reason=data['downtime_reason'],
            machine_id=data['machine_id'],
            product_id=data.get('product_id'),
            week_number=data['week_number'],
            year=data['year'],
            month=data['month'],
            total_duration=data['total_duration'],
            root_cause=data.get('root_cause'),
            follow_up=data.get('follow_up'),
            status=data.get('status', 'pending'),
            pic=data.get('pic'),
            created_by=current_user_id,
            updated_by=current_user_id
        )
        
        db.session.add(action_item)
        db.session.commit()
        
        return jsonify({
            'message': 'Action item created successfully',
            'action_item': action_item.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@downtime_actions_bp.route('/action-items/<int:item_id>', methods=['PUT'])
@jwt_required()
def update_action_item(item_id):
    """Update action item (root cause, follow up, status, PIC)"""
    try:
        data = request.get_json()
        current_user_id = get_jwt_identity()
        
        action_item = DowntimeActionItem.query.get_or_404(item_id)
        
        # Update fields
        if 'root_cause' in data:
            action_item.root_cause = data['root_cause']
        if 'follow_up' in data:
            action_item.follow_up = data['follow_up']
        if 'status' in data:
            action_item.status = data['status']
        if 'pic' in data:
            action_item.pic = data['pic']
        
        action_item.updated_by = current_user_id
        action_item.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'message': 'Action item updated successfully',
            'action_item': action_item.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@downtime_actions_bp.route('/action-items/<int:item_id>', methods=['DELETE'])
@jwt_required()
def delete_action_item(item_id):
    """Delete action item"""
    try:
        action_item = DowntimeActionItem.query.get_or_404(item_id)
        db.session.delete(action_item)
        db.session.commit()
        
        return jsonify({'message': 'Action item deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@downtime_actions_bp.route('/generate-action-items', methods=['POST'])
@jwt_required()
def generate_action_items():
    """
    Generate action items from top 3 unplanned downtime per machine per product per week
    This should be run weekly (e.g., every Monday morning)
    """
    try:
        data = request.get_json()
        current_user_id = get_jwt_identity()
        
        # Get date range (default: last week)
        if 'start_date' in data and 'end_date' in data:
            start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
            end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
        else:
            # Default: last week
            today = datetime.now().date()
            start_date = today - timedelta(days=today.weekday() + 7)  # Last Monday
            end_date = start_date + timedelta(days=6)  # Last Sunday
        
        week_number = get_week_number(start_date)
        year = start_date.year
        month = start_date.month
        
        # Get all shift productions in date range
        shift_productions = ShiftProduction.query.filter(
            and_(
                ShiftProduction.production_date >= start_date,
                ShiftProduction.production_date <= end_date
            )
        ).all()
        
        # Aggregate downtime by machine, product, and reason
        downtime_agg = {}
        
        for sp in shift_productions:
            if not sp.issues:
                continue
            
            # Parse issues string: "60 menit - Description [category]"
            issues_list = sp.issues.split('\n')
            
            for issue in issues_list:
                if not issue.strip():
                    continue
                
                try:
                    # Extract duration, reason, and category
                    parts = issue.split(' - ', 1)
                    if len(parts) < 2:
                        continue
                    
                    duration_str = parts[0].replace('menit', '').strip()
                    duration = int(duration_str)
                    
                    rest = parts[1]
                    if '[' in rest and ']' in rest:
                        reason = rest[:rest.rfind('[')].strip()
                        category = rest[rest.rfind('[')+1:rest.rfind(']')].strip()
                    else:
                        reason = rest.strip()
                        category = 'unknown'
                    
                    # Only process UNPLANNED downtime
                    if category not in UNPLANNED_CATEGORIES:
                        continue
                    
                    # Create key for aggregation
                    key = (sp.machine_id, sp.product_id, reason)
                    
                    if key not in downtime_agg:
                        downtime_agg[key] = {
                            'machine_id': sp.machine_id,
                            'product_id': sp.product_id,
                            'reason': reason,
                            'total_duration': 0
                        }
                    
                    downtime_agg[key]['total_duration'] += duration
                    
                except Exception as e:
                    print(f"Error parsing issue: {issue} - {str(e)}")
                    continue
        
        # Group by machine and product, get top 3 per group
        grouped = {}
        for key, data in downtime_agg.items():
            machine_id = data['machine_id']
            product_id = data['product_id']
            group_key = (machine_id, product_id)
            
            if group_key not in grouped:
                grouped[group_key] = []
            
            grouped[group_key].append(data)
        
        # Create action items for top 3 in each group
        created_count = 0
        
        for group_key, items in grouped.items():
            # Sort by duration desc, take top 3
            top_3 = sorted(items, key=lambda x: x['total_duration'], reverse=True)[:3]
            
            for item in top_3:
                # Check if action item already exists
                existing = DowntimeActionItem.query.filter_by(
                    downtime_reason=item['reason'],
                    machine_id=item['machine_id'],
                    product_id=item['product_id'],
                    week_number=week_number,
                    year=year,
                    month=month
                ).first()
                
                if existing:
                    # Update duration if different
                    if existing.total_duration != item['total_duration']:
                        existing.total_duration = item['total_duration']
                        existing.updated_by = current_user_id
                        existing.updated_at = datetime.utcnow()
                    continue
                
                # Create new action item
                action_item = DowntimeActionItem(
                    downtime_reason=item['reason'],
                    machine_id=item['machine_id'],
                    product_id=item['product_id'],
                    week_number=week_number,
                    year=year,
                    month=month,
                    total_duration=item['total_duration'],
                    status='pending',
                    created_by=current_user_id,
                    updated_by=current_user_id
                )
                
                db.session.add(action_item)
                created_count += 1
        
        db.session.commit()
        
        return jsonify({
            'message': f'Generated {created_count} action items',
            'week_number': week_number,
            'year': year,
            'month': month,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@downtime_actions_bp.route('/export-excel', methods=['GET'])
@jwt_required(optional=True)
def export_downtime_excel():
    """Export downtime action items report to Excel"""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from utils.timezone import get_local_now
        import re
        from utils import detect_downtime_category
        
        def clean_product_name(name):
            if not name:
                return name
            if name.startswith('@'):
                name = name[1:].strip()
            return re.sub(r'\s*@\S+', '', name).strip()
        
        year = request.args.get('year', get_local_now().year, type=int)
        month = request.args.get('month', get_local_now().month, type=int)
        view_mode = request.args.get('view', 'monthly')  # 'weekly' or 'monthly'
        week_number = request.args.get('week', 0, type=int)
        
        # Calculate date range
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
            
        month_end_original = end_date
        
        # Calculate first Monday of the month for proper week calculation
        first_day_of_month = datetime(year, month, 1).date()
        days_until_monday = (7 - first_day_of_month.weekday()) % 7
        if first_day_of_month.weekday() != 0:
            first_monday = first_day_of_month + timedelta(days=days_until_monday)
        else:
            first_monday = first_day_of_month

        if view_mode == 'weekly' and week_number > 0:
            week_start = first_monday + timedelta(days=(week_number - 1) * 7)
            week_end = min(week_start + timedelta(days=6), month_end_original)
            start_date = week_start
            end_date = week_end
            
        # Get shift productions in date range
        shift_productions = ShiftProduction.query.filter(
            ShiftProduction.production_date >= start_date,
            ShiftProduction.production_date <= end_date
        ).all()
        
        # Aggregate unplanned downtime by product and reason (mesin and idle, to match dashboard)
        product_downtime_agg = {}
        for sp in shift_productions:
            if not sp.issues or not sp.product:
                continue
            
            product_name = clean_product_name(sp.product.name)
            machine_name = sp.machine.name if sp.machine else f"Machine {sp.machine_id}"
            
            issue_parts = sp.issues.split(';')
            for idx, part in enumerate(issue_parts):
                part = part.strip()
                if not part:
                    continue
                
                match = re.match(r'(\d+)\s*menit\s*-\s*(.+?)(?:\s*\[([^\]]+)\])?\s*$', part, re.IGNORECASE)
                if match:
                    duration = int(match.group(1))
                    reason = match.group(2).strip()
                    reason = re.sub(r'\s*\[.+\]\s*$', '', reason).strip()
                    
                    excluded = ['istirahat', 'sholat', 'solat', 'toilet', 'makan', 'minum']
                    if any(kw in reason.lower() for kw in excluded):
                        continue
                    
                    try:
                        is_first = (idx == 0)
                        category = detect_downtime_category(reason, is_first)
                    except TypeError:
                        category = detect_downtime_category(reason)
                    
                    # Dashboard filters only mesin and idle
                    UNPLANNED_CATEGORIES = ['mesin', 'idle']
                    if category not in UNPLANNED_CATEGORIES:
                        continue
                    
                    key = (product_name, reason)
                    if key not in product_downtime_agg:
                        product_downtime_agg[key] = {
                            'product_name': product_name,
                            'reason': reason,
                            'category': category,
                            'total_duration': 0,
                            'machines': set()
                        }
                    
                    product_downtime_agg[key]['total_duration'] += duration
                    product_downtime_agg[key]['machines'].add(machine_name)
                    
        # Group and take top 3 per product
        top_unplanned_by_product = {}
        for (product_name, reason), data in product_downtime_agg.items():
            if product_name not in top_unplanned_by_product:
                top_unplanned_by_product[product_name] = []
            
            top_unplanned_by_product[product_name].append({
                'reason': reason,
                'category': data['category'],
                'total_duration': data['total_duration'],
                'machines': ', '.join(sorted(data['machines']))
            })
            
        for product_name in top_unplanned_by_product:
            top_unplanned_by_product[product_name].sort(key=lambda x: x['total_duration'], reverse=True)
            top_unplanned_by_product[product_name] = top_unplanned_by_product[product_name][:3]
            
        # Get all saved action items for this month/week
        action_query = DowntimeActionItem.query.filter(
            DowntimeActionItem.year == year,
            DowntimeActionItem.month == month
        )
        if view_mode == 'weekly' and week_number > 0:
            action_query = action_query.filter(DowntimeActionItem.week_number == week_number)
            
        action_items = action_query.all()
        saved_action_items = {}
        for item in action_items:
            # We map using both exact name and cleaned product name
            key = f"{item.product.name if item.product else ''}__{item.downtime_reason}"
            cleaned_key = f"{clean_product_name(item.product.name if item.product else '')}__{item.downtime_reason}"
            saved_action_items[key] = item
            saved_action_items[cleaned_key] = item
            
        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Unplanned Downtime"
        
        # Styles
        title_font = Font(name="Calibri", size=14, bold=True, color="1F497D")
        subtitle_font = Font(name="Calibri", size=10, italic=True, color="595959")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        bold_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='medium', color='1F497D')
        )
        
        # Status styling helper
        status_colors = {
            'resolved': {'fill': "E2EFDA", 'font': "375623"}, # Green
            'in_progress': {'fill': "DDEBF7", 'font': "1F4E79"}, # Blue
            'pending': {'fill': "FFF2CC", 'font': "7F6000"} # Yellow
        }
        
        # Header title banner
        month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
            'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        period_label = f"Bulan {month_names[month]} {year}"
        if view_mode == 'weekly' and week_number > 0:
            period_label = f"Minggu {week_number} ({month_names[month]} {year})"
            
        ws.cell(row=1, column=1, value="LAPORAN UNPLANNED DOWNTIME & TINDAK LANJUT").font = title_font
        ws.cell(row=2, column=1, value=f"Periode: {period_label} | Dibuat pada: {get_local_now().strftime('%d-%m-%Y %H:%M:%S')}").font = subtitle_font
        
        # Headers on Row 4
        headers = [
            "Rank", "Produk", "Penyebab Downtime", "Kategori", "Mesin", 
            "Durasi (menit)", "Root Cause (Akar Masalah)", "Follow Up / Solusi", "PIC", "Status"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = bold_border
            
        ws.row_dimensions[4].height = 25
        
        # Populate data
        row_idx = 5
        
        # Flatten and sort
        flat_items = []
        for product_name, items in top_unplanned_by_product.items():
            for idx, item in enumerate(items):
                flat_items.append({
                    'product_name': product_name,
                    'rank': idx + 1,
                    'reason': item['reason'],
                    'category': item['category'],
                    'machines': item['machines'],
                    'total_duration': item['total_duration']
                })
                
        # Sort flat items by product_name, then by rank
        flat_items.sort(key=lambda x: (x['product_name'], x['rank']))
        
        for item in flat_items:
            p_name = item['product_name']
            reason = item['reason']
            lookup_key = f"{p_name}__{reason}"
            saved = saved_action_items.get(lookup_key)
            
            # Map values
            status_val = saved.status if saved else 'pending'
            root_cause_val = saved.root_cause if saved else ''
            follow_up_val = saved.follow_up if saved else ''
            pic_val = saved.pic if saved else ''
            
            # Write to row
            ws.cell(row=row_idx, column=1, value=item['rank']).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value=p_name)
            ws.cell(row=row_idx, column=3, value=reason)
            ws.cell(row=row_idx, column=4, value=item['category']).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=5, value=item['machines'])
            
            dur_cell = ws.cell(row=row_idx, column=6, value=item['total_duration'])
            dur_cell.alignment = Alignment(horizontal="right")
            dur_cell.font = Font(name="Calibri", size=11, bold=True, color="C00000") # Dark red for duration
            
            ws.cell(row=row_idx, column=7, value=root_cause_val)
            ws.cell(row=row_idx, column=8, value=follow_up_val)
            ws.cell(row=row_idx, column=9, value=pic_val)
            
            # Status styling
            status_cell = ws.cell(row=row_idx, column=10, value=status_val.upper())
            status_cell.alignment = Alignment(horizontal="center")
            if status_val in status_colors:
                color_info = status_colors[status_val]
                status_cell.fill = PatternFill(start_color=color_info['fill'], end_color=color_info['fill'], fill_type="solid")
                status_cell.font = Font(name="Calibri", size=10, bold=True, color=color_info['font'])
                
            # Apply borders to all columns in this row
            for col_idx in range(1, 11):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                # Vertical center alignment for text
                if col_idx not in [1, 4, 6, 10]: # Except rank, category, duration, status
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center" if col_idx != 6 else "right", vertical="center")
                    
            ws.row_dimensions[row_idx].height = 20
            row_idx += 1
            
        # Auto-adjust column widths
        column_widths = {
            1: 6,   # Rank
            2: 25,  # Produk
            3: 30,  # Penyebab Downtime
            4: 12,  # Kategori
            5: 15,  # Mesin
            6: 15,  # Durasi (menit)
            7: 35,  # Root Cause
            8: 35,  # Follow Up
            9: 15,  # PIC
            10: 15  # Status
        }
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            
        # Freeze panes below header
        ws.freeze_panes = 'A5'
        
        # Save to byte stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Filename
        mode_str = "bulanan" if view_mode == "monthly" else f"mingguan_w{week_number}"
        filename = f"downtime_action_items_{mode_str}_{year}_{month}.xlsx"
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Export Excel error: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@downtime_actions_bp.route('/export-pdf', methods=['GET'])
@jwt_required(optional=True)
def export_downtime_pdf():
    """Export downtime action items report to vector PDF"""
    try:
        import io
        import re
        from flask import send_file
        from utils.timezone import get_local_now
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        year = request.args.get('year', get_local_now().year, type=int)
        month = request.args.get('month', get_local_now().month, type=int)
        view_mode = request.args.get('view', 'monthly')  # 'weekly' or 'monthly'
        week_number = request.args.get('week', 0, type=int)
        
        # Calculate date range
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
            
        month_end_original = end_date
        
        # Calculate first Monday of the month for proper week calculation
        first_day_of_month = datetime(year, month, 1).date()
        days_until_monday = (7 - first_day_of_month.weekday()) % 7
        if first_day_of_month.weekday() != 0:
            first_monday = first_day_of_month + timedelta(days=days_until_monday)
        else:
            first_monday = first_day_of_month

        if view_mode == 'weekly' and week_number > 0:
            week_start = first_monday + timedelta(days=(week_number - 1) * 7)
            week_end = min(week_start + timedelta(days=6), month_end_original)
            start_date = week_start
            end_date = week_end
            
        # Get shift productions in date range
        shift_productions = ShiftProduction.query.filter(
            ShiftProduction.production_date >= start_date,
            ShiftProduction.production_date <= end_date
        ).all()
        
        # Define clean_product_name locally
        def clean_product_name(name):
            if not name:
                return name
            if name.startswith('@'):
                name = name[1:].strip()
            return re.sub(r'\s*@\S+', '', name).strip()
            
        # Aggregate unplanned downtime by product and reason (mesin and idle, to match dashboard)
        product_downtime_agg = {}
        for sp in shift_productions:
            if not sp.issues or not sp.product:
                continue
            
            product_name = clean_product_name(sp.product.name)
            machine_name = sp.machine.name if sp.machine else f"Machine {sp.machine_id}"
            
            issue_parts = sp.issues.split(';')
            for idx, part in enumerate(issue_parts):
                part = part.strip()
                if not part:
                    continue
                
                match = re.match(r'(\d+)\s*menit\s*-\s*(.+?)(?:\s*\[([^\]]+)\])?\s*$', part, re.IGNORECASE)
                if match:
                    duration = int(match.group(1))
                    reason = match.group(2).strip()
                    reason = re.sub(r'\s*\[.+\]\s*$', '', reason).strip()
                    
                    excluded = ['istirahat', 'sholat', 'solat', 'toilet', 'makan', 'minum']
                    if any(kw in reason.lower() for kw in excluded):
                        continue
                    
                    try:
                        from utils import detect_downtime_category
                        is_first = (idx == 0)
                        category = detect_downtime_category(reason, is_first)
                    except TypeError:
                        category = detect_downtime_category(reason)
                    
                    # Dashboard filters only mesin and idle
                    UNPLANNED_CATEGORIES = ['mesin', 'idle']
                    if category not in UNPLANNED_CATEGORIES:
                        continue
                    
                    key = (product_name, reason)
                    if key not in product_downtime_agg:
                        product_downtime_agg[key] = {
                            'product_name': product_name,
                            'reason': reason,
                            'category': category,
                            'total_duration': 0,
                            'machines': set()
                        }
                    
                    product_downtime_agg[key]['total_duration'] += duration
                    product_downtime_agg[key]['machines'].add(machine_name)
                    
        # Group and take top 3 per product
        top_unplanned_by_product = {}
        for (product_name, reason), data in product_downtime_agg.items():
            if product_name not in top_unplanned_by_product:
                top_unplanned_by_product[product_name] = []
            
            top_unplanned_by_product[product_name].append({
                'reason': reason,
                'category': data['category'],
                'total_duration': data['total_duration'],
                'machines': ', '.join(sorted(data['machines']))
            })
            
        for product_name in top_unplanned_by_product:
            top_unplanned_by_product[product_name].sort(key=lambda x: x['total_duration'], reverse=True)
            top_unplanned_by_product[product_name] = top_unplanned_by_product[product_name][:3]
            
        # Get all saved action items for this month/week
        action_query = DowntimeActionItem.query.filter(
            DowntimeActionItem.year == year,
            DowntimeActionItem.month == month
        )
        if view_mode == 'weekly' and week_number > 0:
            action_query = action_query.filter(DowntimeActionItem.week_number == week_number)
            
        action_items = action_query.all()
        saved_action_items = {}
        for item in action_items:
            # We map using both exact name and cleaned product name
            key = f"{item.product.name if item.product else ''}__{item.downtime_reason}"
            cleaned_key = f"{clean_product_name(item.product.name if item.product else '')}__{item.downtime_reason}"
            saved_action_items[key] = item
            saved_action_items[cleaned_key] = item
            
        # Flatten and sort
        flat_items = []
        for product_name, items in top_unplanned_by_product.items():
            for idx, item in enumerate(items):
                flat_items.append({
                    'product_name': product_name,
                    'rank': idx + 1,
                    'reason': item['reason'],
                    'category': item['category'],
                    'machines': item['machines'],
                    'total_duration': item['total_duration']
                })
                
        # Sort flat items by product_name, then by rank
        flat_items.sort(key=lambda x: (x['product_name'], x['rank']))
        
        # Create PDF BytesIO buffer
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20
        )
        
        # Styles - use unique suffix to avoid ReportLab global style name collisions across requests
        import time
        _sfx = str(int(time.time() * 1000))
        
        styles = getSampleStyleSheet()
        normal_style = ParagraphStyle(
            f'CellNormal_{_sfx}',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#333333')
        )
        bold_style = ParagraphStyle(
            f'CellBold_{_sfx}',
            parent=normal_style,
            fontName='Helvetica-Bold'
        )
        header_style = ParagraphStyle(
            f'HeaderStyle_{_sfx}',
            parent=normal_style,
            fontName='Helvetica-Bold',
            fontSize=9,
            leading=11,
            textColor=colors.white,
            alignment=1 # Center
        )
        title_style = ParagraphStyle(
            f'TitleStyle_{_sfx}',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=16,
            leading=18,
            textColor=colors.HexColor('#1F497D'),
            spaceAfter=6
        )
        subtitle_style = ParagraphStyle(
            f'SubtitleStyle_{_sfx}',
            parent=styles['Normal'],
            fontName='Helvetica-Oblique',
            fontSize=10,
            leading=12,
            textColor=colors.HexColor('#595959'),
            spaceAfter=15
        )
        
        # Header title banner
        month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
            'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        period_label = f"Bulan {month_names[month]} {year}"
        if view_mode == 'weekly' and week_number > 0:
            period_label = f"Minggu {week_number} ({month_names[month]} {year})"
            
        # Header row
        table_data = [[
            Paragraph("Rank", header_style),
            Paragraph("Produk", header_style),
            Paragraph("Penyebab Downtime", header_style),
            Paragraph("Kategori", header_style),
            Paragraph("Mesin", header_style),
            Paragraph("Durasi", header_style),
            Paragraph("Root Cause (Akar Masalah)", header_style),
            Paragraph("Follow Up / Solusi", header_style),
            Paragraph("PIC", header_style),
            Paragraph("Status", header_style)
        ]]
        
        # Pre-define reusable styles OUTSIDE the loop (ReportLab rejects duplicate style names)
        center_bold_style = ParagraphStyle(
            f'CenterBold_{_sfx}', parent=bold_style, alignment=1
        )
        center_style = ParagraphStyle(
            f'CenterNormal_{_sfx}', parent=normal_style, alignment=1
        )
        dur_style_cell = ParagraphStyle(
            f'DurationStyle_{_sfx}',
            parent=normal_style,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#C00000'),
            alignment=2  # Right
        )
        status_styles = {
            'PENDING': ParagraphStyle(
                f'Status_PENDING_{_sfx}', parent=normal_style,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#7F6000'),
                alignment=1
            ),
            'RESOLVED': ParagraphStyle(
                f'Status_RESOLVED_{_sfx}', parent=normal_style,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#375623'),
                alignment=1
            ),
            'IN_PROGRESS': ParagraphStyle(
                f'Status_IN_PROGRESS_{_sfx}', parent=normal_style,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor('#1F4E79'),
                alignment=1
            ),
        }
        
        # Populate rows
        for item in flat_items:
            p_name = item['product_name']
            reason = item['reason']
            lookup_key = f"{p_name}__{reason}"
            saved = saved_action_items.get(lookup_key)
            
            status_val = (saved.status if saved else 'pending').upper()
            root_cause_val = saved.root_cause if saved else ''
            follow_up_val = saved.follow_up if saved else ''
            pic_val = saved.pic if saved else ''
            
            # Get pre-defined status style (fallback to PENDING)
            status_style_cell = status_styles.get(status_val, status_styles['PENDING'])
            
            table_data.append([
                Paragraph(str(item['rank']), center_bold_style),
                Paragraph(p_name, normal_style),
                Paragraph(reason, normal_style),
                Paragraph(item['category'], center_style),
                Paragraph(item['machines'], normal_style),
                Paragraph(f"{item['total_duration']} m", dur_style_cell),
                Paragraph(root_cause_val or '', normal_style),
                Paragraph(follow_up_val or '', normal_style),
                Paragraph(pic_val or '', normal_style),
                Paragraph(status_val, status_style_cell)
            ])
            
        # Define table column widths (total 800)
        col_widths = [30, 90, 110, 60, 60, 60, 130, 130, 60, 70]
        
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Styles
        t_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F497D')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1F497D')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]
        
        # Alternating row colors
        for i in range(1, len(table_data)):
            bg_color = colors.HexColor('#F2F5F8') if i % 2 == 0 else colors.white
            t_style.append(('BACKGROUND', (0, i), (-1, i), bg_color))
            
        t.setStyle(TableStyle(t_style))
        
        story = [
            Paragraph("LAPORAN UNPLANNED DOWNTIME & TINDAK LANJUT", title_style),
            Paragraph(f"Periode: {period_label} | Dibuat pada: {get_local_now().strftime('%d-%m-%Y %H:%M:%S')}", subtitle_style),
            t
        ]
        
        doc.build(story)
        output.seek(0)
        
        mode_str = "bulanan" if view_mode == "monthly" else f"mingguan_w{week_number}"
        filename = f"downtime_action_items_{mode_str}_{year}_{month}.pdf"
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Export PDF error: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@downtime_actions_bp.route('/action-items/summary', methods=['GET'])
@jwt_required()
def get_action_items_summary():
    """Get summary of action items by status"""
    try:
        # Count by status
        summary = db.session.query(
            DowntimeActionItem.status,
            func.count(DowntimeActionItem.id).label('count'),
            func.sum(DowntimeActionItem.total_duration).label('total_duration')
        ).group_by(DowntimeActionItem.status).all()
        
        result = {
            'by_status': [
                {
                    'status': row.status,
                    'count': row.count,
                    'total_duration': int(row.total_duration) if row.total_duration else 0
                }
                for row in summary
            ],
            'total_items': sum(row.count for row in summary),
            'total_duration': sum(int(row.total_duration) if row.total_duration else 0 for row in summary)
        }
        
        return jsonify(result), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
