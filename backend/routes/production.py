from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db, Machine, WorkOrder, ProductionRecord, BillOfMaterials, BOMItem, ProductionSchedule, Product, Employee
from models.production import RemainingStock
from models.work_order_bom import WorkOrderBOMItem
from models.product import Material
from models.product_excel_schema import ProductNew
from models.notification import Notification
from models.user import User
from utils.i18n import success_response, error_response, get_message
from utils import generate_number
from utils.timezone import get_local_now, get_local_today, utc_to_local
from datetime import datetime, timedelta
from sqlalchemy import func, and_, or_
from io import BytesIO

def create_production_notification(user_ids, notification_type, category, title, message, reference_type=None, reference_id=None, priority='normal', action_url=None):
    """Create notifications for production events"""
    try:
        for user_id in user_ids:
            notification = Notification(
                user_id=user_id,
                notification_type=notification_type,
                category=category,
                title=title,
                message=message,
                reference_type=reference_type,
                reference_id=reference_id,
                priority=priority,
                action_url=action_url
            )
            db.session.add(notification)
        db.session.commit()
    except Exception as e:
        print(f"Error creating notification: {e}")

def get_supervisor_user_ids():
    """Get user IDs of all active users for notifications"""
    try:
        # Get all active users
        users = User.query.filter(User.is_active == True).all()
        return [u.id for u in users]
    except:
        return []

def get_product_name_from_new(product_code):
    """Get updated product name from ProductNew model"""
    if not product_code:
        return None
    product_new = ProductNew.query.filter_by(code=product_code).first()
    if product_new:
        return product_new.name
    return None

production_bp = Blueprint('production', __name__)

# ============= MACHINES =============
@production_bp.route('/machines', methods=['GET'])
@jwt_required()
def get_machines():
    """
    Get all active machines
    ---
    tags:
      - Production
    summary: Get all machines
    description: Retrieve all active machines in the production system
    security:
      - BearerAuth: []
    responses:
      200:
        description: Machines retrieved successfully
        schema:
          type: object
          properties:
            machines:
              type: array
              items:
                type: object
                properties:
                  id:
                    type: integer
                  code:
                    type: string
                  name:
                    type: string
                  machine_type:
                    type: string
                  status:
                    type: string
                  location:
                    type: string
                  department:
                    type: string
      401:
        description: Unauthorized
      500:
        description: Server error
    """
    try:
        machines = Machine.query.filter_by(is_active=True).all()
        return jsonify({
            'machines': [{
                'id': m.id,
                'code': m.code,
                'name': m.name,
                'machine_type': m.machine_type,
                'manufacturer': m.manufacturer,
                'model': m.model,
                'serial_number': m.serial_number,
                'status': m.status,
                'location': m.location,
                'department': m.department,
                'capacity_per_hour': float(m.capacity_per_hour) if m.capacity_per_hour else None,
                'capacity_uom': m.capacity_uom,
                'efficiency': float(m.efficiency) if m.efficiency else 100,
                'availability': float(m.availability) if m.availability else 100,
                'last_maintenance': m.last_maintenance.isoformat() if m.last_maintenance else None,
                'next_maintenance': m.next_maintenance.isoformat() if m.next_maintenance else None,
                'installation_date': m.installation_date.isoformat() if m.installation_date else None,
                'notes': m.notes,
                'created_at': m.created_at.isoformat() if m.created_at else None,
                'updated_at': m.updated_at.isoformat() if m.updated_at else None
            } for m in machines]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@production_bp.route('/machines/<int:id>', methods=['GET', 'PUT'])
@jwt_required()
def get_or_update_machine(id):
    """
    Get or update machine by ID
    ---
    tags:
      - Production
    summary: Get or update machine
    description: Retrieve a specific machine by ID or update it
    security:
      - BearerAuth: []
    parameters:
      - in: path
        name: id
        required: true
        type: integer
        description: Machine ID
      - in: body
        name: body
        required: false
        schema:
          type: object
          properties:
            name:
              type: string
            status:
              type: string
            location:
              type: string
            department:
              type: string
    responses:
      200:
        description: Machine retrieved/updated successfully
        schema:
          type: object
          properties:
            id:
              type: integer
            code:
              type: string
            name:
              type: string
            machine_type:
              type: string
            status:
              type: string
      401:
        description: Unauthorized
      404:
        description: Machine not found
      500:
        description: Server error
    """
    try:
        machine = db.session.get(Machine, id)
        if not machine:
            return jsonify(error_response('api.error', error_code=404)), 404
        
        # Handle PUT request - update machine
        if request.method == 'PUT':
            data = request.get_json()
            print(f"=== UPDATE MACHINE {id} ===")
            print(f"Received data: {data}")
            
            if 'name' in data:
                machine.name = data['name']
            if 'code' in data:
                machine.code = data['code']
            if 'machine_type' in data:
                machine.machine_type = data['machine_type']
            if 'manufacturer' in data:
                machine.manufacturer = data['manufacturer']
            if 'model' in data:
                machine.model = data['model']
            if 'serial_number' in data:
                machine.serial_number = data['serial_number']
            if 'status' in data:
                machine.status = data['status']
            if 'location' in data:
                machine.location = data['location']
            if 'department' in data:
                machine.department = data['department']
            # Handle both capacity and capacity_per_hour - prefer 'capacity' from form
            if 'capacity' in data and data['capacity']:
                machine.capacity_per_hour = data['capacity']
            elif 'capacity_per_hour' in data and data['capacity_per_hour']:
                machine.capacity_per_hour = data['capacity_per_hour']
            # Handle both uom and capacity_uom - prefer 'uom' from form
            if 'uom' in data and data['uom']:
                machine.capacity_uom = data['uom']
            elif 'capacity_uom' in data and data['capacity_uom']:
                machine.capacity_uom = data['capacity_uom']
            if 'notes' in data:
                machine.notes = data['notes']
            if 'specifications' in data:
                machine.specifications = data['specifications']
            if 'is_active' in data:
                machine.is_active = data['is_active']
            if 'maintenance_schedule' in data:
                machine.maintenance_schedule = data['maintenance_schedule']
            
            if data.get('installation_date'):
                try:
                    date_str = data['installation_date']
                    if isinstance(date_str, str):
                        if 'T' in date_str:
                            machine.installation_date = datetime.fromisoformat(date_str.replace('Z', '+00:00')).date()
                        else:
                            machine.installation_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                except:
                    pass
            
            db.session.commit()
            
            # Debug: show what was saved
            print(f"=== SAVED MACHINE {id} ===")
            print(f"name: {machine.name}")
            print(f"code: {machine.code}")
            print(f"machine_type: {machine.machine_type}")
            print(f"capacity_per_hour: {machine.capacity_per_hour}")
            print(f"location: {machine.location}")
            print(f"notes: {machine.notes}")
            
            return jsonify({'message': 'Machine updated successfully', 'machine_id': machine.id}), 200
            
        # Handle GET request - return machine details
        return jsonify({
            'machine': {
                'id': machine.id,
                'code': machine.code,
                'name': machine.name,
                'machine_type': machine.machine_type,
                'manufacturer': machine.manufacturer,
                'model': machine.model,
                'serial_number': machine.serial_number,
                'status': machine.status,
                'location': machine.location,
                'department': machine.department,
                # Include both field names for compatibility
                'capacity_per_hour': float(machine.capacity_per_hour) if machine.capacity_per_hour else 0,
                'capacity': float(machine.capacity_per_hour) if machine.capacity_per_hour else 0,
                'capacity_uom': machine.capacity_uom,
                'uom': machine.capacity_uom,
                'specifications': machine.specifications,
                'efficiency': float(machine.efficiency) if machine.efficiency else 100,
                'availability': float(machine.availability) if machine.availability else 100,
                'last_maintenance': machine.last_maintenance.isoformat() if machine.last_maintenance else None,
                'next_maintenance': machine.next_maintenance.isoformat() if machine.next_maintenance else None,
                'installation_date': machine.installation_date.isoformat() if machine.installation_date else None,
                'maintenance_schedule': machine.maintenance_schedule,
                'notes': machine.notes,
                'is_active': machine.is_active if machine.is_active is not None else True,
                'created_at': machine.created_at.isoformat() if machine.created_at else None,
                'updated_at': machine.updated_at.isoformat() if machine.updated_at else None
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/machines', methods=['POST'])
@jwt_required()
def create_machine():
    """
    Create a new machine
    ---
    tags:
      - Production
    summary: Create new machine
    description: Create a new machine in the production system
    security:
      - BearerAuth: []
    parameters:
      - in: body
        name: body
        required: true
        schema:
          type: object
          required:
            - code
            - name
            - machine_type
          properties:
            code:
              type: string
            name:
              type: string
            machine_type:
              type: string
            manufacturer:
              type: string
            model:
              type: string
            status:
              type: string
            location:
              type: string
            department:
              type: string
    responses:
      201:
        description: Machine created successfully
        schema:
          type: object
          properties:
            message:
              type: string
            machine:
              type: object
      401:
        description: Unauthorized
      500:
        description: Server error
    """
    try:
        data = request.get_json()
        machine = Machine(
            code=data['code'],
            name=data['name'],
            machine_type=data['machine_type'],
            manufacturer=data.get('manufacturer'),
            model=data.get('model'),
            serial_number=data.get('serial_number'),
            status='idle',
            location=data.get('location'),
            capacity_per_hour=data.get('capacity_per_hour')
        )
        db.session.add(machine)
        db.session.commit()
        return jsonify({'message': 'Machine created', 'machine_id': machine.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/machines/<int:id>/update', methods=['PUT'])
@jwt_required()
def update_machine(id):
    try:
        machine = db.session.get(Machine, id)
        if not machine:
            return jsonify(error_response('api.error', error_code=404)), 404
            
        data = request.get_json()
        
        # Update machine fields
        if 'name' in data:
            machine.name = data['name']
        if 'machine_type' in data:
            machine.machine_type = data['machine_type']
        if 'manufacturer' in data:
            machine.manufacturer = data['manufacturer']
        if 'model' in data:
            machine.model = data['model']
        if 'serial_number' in data:
            machine.serial_number = data['serial_number']
        if 'status' in data:
            machine.status = data['status']
        if 'location' in data:
            machine.location = data['location']
        if 'department' in data:
            machine.department = data['department']
        if 'capacity_per_hour' in data:
            machine.capacity_per_hour = data['capacity_per_hour']
        if 'capacity_uom' in data:
            machine.capacity_uom = data['capacity_uom']
        if 'efficiency' in data:
            machine.efficiency = data['efficiency']
        if 'availability' in data:
            machine.availability = data['availability']
        if 'last_maintenance' in data:
            machine.last_maintenance = datetime.fromisoformat(data['last_maintenance']).date() if data['last_maintenance'] else None
        if 'next_maintenance' in data:
            machine.next_maintenance = datetime.fromisoformat(data['next_maintenance']).date() if data['next_maintenance'] else None
        if 'notes' in data:
            machine.notes = data['notes']
        if 'is_active' in data:
            machine.is_active = data['is_active']
            
        machine.updated_at = datetime.utcnow()
        
        db.session.commit()
        return jsonify(success_response('api.success')), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/machines/<int:id>/efficiency', methods=['GET'])
@jwt_required()
def get_machine_efficiency(id):
    """Get machine efficiency data for specific time period"""
    try:
        machine = db.session.get(Machine, id)
        if not machine:
            return jsonify(error_response('api.error', error_code=404)), 404
            
        start_date = request.args.get('start_date', (get_local_now() - timedelta(days=30)).isoformat())
        end_date = request.args.get('end_date', get_local_now().isoformat())
        
        start_dt = datetime.fromisoformat(start_date)
        end_dt = datetime.fromisoformat(end_date)
        
        # Get production records for the period
        records = ProductionRecord.query.filter(
            ProductionRecord.machine_id == id,
            ProductionRecord.production_date.between(start_dt, end_dt)
        ).all()
        
        # Calculate efficiency metrics
        total_produced = sum(float(r.quantity_produced) for r in records)
        total_good = sum(float(r.quantity_good) for r in records)
        total_scrap = sum(float(r.quantity_scrap) for r in records)
        total_downtime = sum(r.downtime_minutes for r in records)
        
        # Calculate efficiency percentages
        quality_rate = (total_good / total_produced * 100) if total_produced > 0 else 0
        scrap_rate = (total_scrap / total_produced * 100) if total_produced > 0 else 0
        
        # Theoretical capacity calculation (assume 8 hours per day)
        days_in_period = (end_dt - start_dt).days + 1
        theoretical_hours = days_in_period * 8 * 60  # in minutes
        actual_runtime = theoretical_hours - total_downtime
        availability_rate = (actual_runtime / theoretical_hours * 100) if theoretical_hours > 0 else 0
        
        # Overall Equipment Effectiveness (OEE)
        oee = (availability_rate * quality_rate * float(machine.efficiency or 100)) / 10000
        
        return jsonify({
            'machine_id': id,
            'machine_name': machine.name,
            'period': {
                'start': start_date,
                'end': end_date,
                'days': days_in_period
            },
            'production': {
                'total_produced': total_produced,
                'total_good': total_good,
                'total_scrap': total_scrap,
                'quality_rate': round(quality_rate, 2),
                'scrap_rate': round(scrap_rate, 2)
            },
            'availability': {
                'theoretical_hours': theoretical_hours / 60,  # convert to hours
                'total_downtime_hours': total_downtime / 60,
                'actual_runtime_hours': actual_runtime / 60,
                'availability_rate': round(availability_rate, 2)
            },
            'efficiency': {
                'performance_rate': float(machine.efficiency or 100),
                'oee': round(oee, 2)
            }
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============= WORK ORDERS =============
@production_bp.route('/work-orders', methods=['GET'])
@jwt_required()
def get_work_orders():
    try:
        from models.production import ShiftProduction
        
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        status = request.args.get('status')
        
        query = WorkOrder.query
        if status:
            query = query.filter_by(status=status)
        
        wos = query.order_by(WorkOrder.created_at.desc()).paginate(page=page, per_page=per_page)
        
        result = []
        for wo in wos.items:
            # Get last input date from ShiftProduction
            last_shift = ShiftProduction.query.filter_by(work_order_id=wo.id).order_by(
                ShiftProduction.created_at.desc()
            ).first()
            
            result.append({
                'id': wo.id,
                'wo_number': wo.wo_number,
                'product_name': get_product_name_from_new(wo.product.code if wo.product else None) or (wo.product.name if wo.product else 'Unknown Product'),
                'quantity': float(wo.quantity) if wo.quantity else 0,
                'quantity_produced': float(wo.quantity_produced) if wo.quantity_produced else 0,
                'quantity_good': float(wo.quantity_good) if wo.quantity_good else 0,
                'status': wo.status,
                'priority': wo.priority or 'normal',
                'machine': wo.machine.name if wo.machine else None,
                'machine_name': wo.machine.name if wo.machine else None,
                'scheduled_start_date': wo.scheduled_start_date.isoformat() if wo.scheduled_start_date else None,
                'scheduled_end_date': wo.scheduled_end_date.isoformat() if wo.scheduled_end_date else None,
                'start_date': wo.scheduled_start_date.isoformat() if wo.scheduled_start_date else wo.created_at.isoformat() if wo.created_at else None
            })
        
        # Summary counts across ALL work orders (not just current page)
        from sqlalchemy import func as sqla_func
        summary_query = db.session.query(
            sqla_func.count(WorkOrder.id).label('total'),
            sqla_func.count(sqla_func.nullif(WorkOrder.status != 'in_progress', True)).label('in_progress'),
            sqla_func.count(sqla_func.nullif(WorkOrder.status != 'completed', True)).label('completed'),
            sqla_func.coalesce(sqla_func.sum(WorkOrder.quantity_produced), 0).label('total_produced')
        ).first()
        
        return jsonify({
            'work_orders': result,
            'total': wos.total,
            'page': page,
            'per_page': per_page,
            'pages': wos.pages,
            'summary': {
                'total': summary_query.total if summary_query else 0,
                'in_progress': summary_query.in_progress if summary_query else 0,
                'completed': summary_query.completed if summary_query else 0,
                'total_produced': float(summary_query.total_produced) if summary_query else 0
            }
        }), 200
    except Exception as e:
        print(f"Work orders error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/work-orders/status-tracking', methods=['GET'])
@jwt_required()
def get_work_orders_status_tracking():
    """Get work orders with input status tracking"""
    try:
        from models.production import ShiftProduction
        
        # Get active work orders (not cancelled)
        wos = WorkOrder.query.filter(
            WorkOrder.status.in_(['draft', 'planned', 'in_progress', 'completed'])
        ).order_by(WorkOrder.created_at.desc()).all()
        
        result = []
        for wo in wos:
            # Get shift productions for this work order
            shift_prods = ShiftProduction.query.filter_by(work_order_id=wo.id).order_by(
                ShiftProduction.production_date.desc(),
                ShiftProduction.created_at.desc()
            ).all()
            
            total_shifts = len(shift_prods)
            last_input = shift_prods[0] if shift_prods else None
            
            # Determine input status
            if total_shifts == 0:
                input_status = 'not_started'
            elif wo.status == 'completed':
                input_status = 'completed'
            else:
                input_status = 'in_progress'
            
            # Calculate progress
            quantity = float(wo.quantity) if wo.quantity else 0
            quantity_good = float(wo.quantity_good) if wo.quantity_good else 0
            progress_percent = (quantity_good / quantity * 100) if quantity > 0 else 0
            
            result.append({
                'id': wo.id,
                'wo_number': wo.wo_number,
                'product_name': wo.product.name if wo.product else 'Unknown',
                'machine_name': wo.machine.name if wo.machine else None,
                'quantity': quantity,
                'quantity_produced': float(wo.quantity_produced) if wo.quantity_produced else 0,
                'quantity_good': quantity_good,
                'status': wo.status,
                'input_status': input_status,
                'progress_percent': round(progress_percent, 1),
                'total_shifts': total_shifts,
                'last_input_date': utc_to_local(last_input.created_at).isoformat() if last_input and last_input.created_at else None,
                'last_input_by': last_input.created_by_user.full_name if last_input and last_input.created_by_user else None,
                'created_at': utc_to_local(wo.created_at).isoformat() if wo.created_at else None,
                'start_date': wo.scheduled_start_date.isoformat() if wo.scheduled_start_date else None,
                'end_date': wo.scheduled_end_date.isoformat() if wo.scheduled_end_date else None,
                'approval_status': None  # Can be extended for approval tracking
            })
        
        return jsonify({'work_orders': result}), 200
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/work-orders/<int:id>', methods=['GET'])
@jwt_required()
def get_work_order(id):
    """Get single work order detail"""
    try:
        wo = db.session.get(WorkOrder, id)
        if not wo:
            return jsonify({'error': 'Work order not found'}), 404
        
        # Get pack_per_carton and consumption data from Product
        pack_per_carton = 0
        consumption_data = {
            'berat_kering_per_pack': 0,  # kg - for kain consumption (per pack)
            'volume_per_pack': 0,  # liter - for ingredient consumption (per pack)
            'berat_akhir_per_pack': 0,  # kg - for packaging & stiker consumption (per pack)
        }
        if wo.product_id:
            from models.production import BillOfMaterials
            from models.product import ProductPackaging
            
            active_bom = BillOfMaterials.query.filter_by(
                product_id=wo.product_id, 
                is_active=True
            ).first()
            # Only use BOM pack_per_carton if it's greater than 1 (meaningful value)
            if active_bom and active_bom.pack_per_carton and active_bom.pack_per_carton > 1:
                pack_per_carton = active_bom.pack_per_carton
            
            # Get consumption data from products_new table (has berat_kering, ingredient, pack_per_karton)
            # Try matching by product code first, then by product name
            product_code = wo.product.code if wo.product else None
            product_name = wo.product.name if wo.product else None
            product_new_data = None
            
            if product_code:
                product_new_data = db.session.execute(
                    db.text('SELECT berat_kering, ratio, ingredient, pack_per_karton FROM products WHERE code = :code'),
                    {'code': product_code}
                ).fetchone()
            
            # If not found by code, try matching by name (partial match)
            # Remove common prefixes like "WIP " for better matching
            if not product_new_data and product_name:
                search_name = product_name
                # Remove WIP prefix if present
                if search_name.upper().startswith('WIP '):
                    search_name = search_name[4:]
                
                # Prioritize @24 packaging (standard) over other variants like @6
                product_new_data = db.session.execute(
                    db.text("SELECT berat_kering, ratio, ingredient, pack_per_karton FROM products WHERE name LIKE :name ORDER BY CASE WHEN name LIKE '%@24' THEN 0 WHEN name LIKE '%@24%' THEN 1 ELSE 2 END, pack_per_karton DESC LIMIT 1"),
                    {'name': f'%{search_name}%'}
                ).fetchone()
            
            # Process product_new_data if found
            if product_new_data:
                berat_kering = float(product_new_data[0]) if product_new_data[0] else 0  # kg per karton
                ratio = float(product_new_data[1]) if product_new_data[1] else 0
                ingredient = float(product_new_data[2]) if product_new_data[2] else 0  # liter per karton
                packs = int(product_new_data[3]) if product_new_data[3] else 1
                
                # Always use packs from products_new as it's the most accurate source
                pack_per_carton = packs
                
                # Convert per karton to per pack
                consumption_data['berat_kering_per_pack'] = berat_kering / packs if packs > 0 else 0  # kg per pack
                consumption_data['volume_per_pack'] = ingredient / packs if packs > 0 else 0  # liter per pack
                consumption_data['ratio'] = ratio
                consumption_data['packs_per_karton'] = packs
                
                # Debug logging
                print(f"DEBUG: WorkOrder {wo.wo_number} - products_new data:")
                print(f"  - product_code: {product_code}")
                print(f"  - product_name: {product_name}")
                print(f"  - packs_per_karton: {packs}")
                print(f"  - berat_kering_per_karton: {berat_kering}")
                print(f"  - ingredient_per_karton: {ingredient}")
                print(f"  - berat_kering_per_pack: {consumption_data['berat_kering_per_pack']}")
                print(f"  - volume_per_pack: {consumption_data['volume_per_pack']}")
            
            # Fallback to ProductPackaging if products_new doesn't have data
            if not consumption_data['berat_kering_per_pack'] and not consumption_data['volume_per_pack']:
                packaging = ProductPackaging.query.filter_by(product_id=wo.product_id).first()
                if packaging:
                    packs = packaging.packs_per_karton or 1
                    if not pack_per_carton:
                        pack_per_carton = packs
                    
                    if packaging.berat_kering_per_karton:
                        consumption_data['berat_kering_per_pack'] = float(packaging.berat_kering_per_karton) / 1000 / packs  # gram to kg per pack
                    
                    if packaging.volume_per_pack:
                        consumption_data['volume_per_pack'] = float(packaging.volume_per_pack) / 1000  # ml to liter
                    
                    if packaging.berat_akhir_per_karton:
                        consumption_data['berat_akhir_per_pack'] = float(packaging.berat_akhir_per_karton) / 1000 / packs  # gram to kg per pack
                else:
                    # Final fallback - use default values if no data found
                    # This ensures waste calculation always works
                    consumption_data['berat_kering_per_pack'] = 0.8  # Default 800g per pack
                    consumption_data['volume_per_pack'] = 0.5  # Default 500ml per pack
                    consumption_data['berat_akhir_per_pack'] = 1.2  # Default 1.2kg per pack
        
        # Build response safely
        response_data = {
            'id': wo.id,
            'wo_number': wo.wo_number,
            'product_id': wo.product_id,
            'product_name': get_product_name_from_new(wo.product.code if wo.product else None) or (wo.product.name if wo.product else 'Unknown Product'),
            'product_code': wo.product.code if wo.product else None,
            'bom_id': wo.bom_id,
            'bom_number': wo.bom.bom_number if wo.bom else None,
            'quantity': float(wo.quantity) if wo.quantity else 0,
            'quantity_produced': float(wo.quantity_produced) if wo.quantity_produced else 0,
            'quantity_good': float(wo.quantity_good) if wo.quantity_good else 0,
            'quantity_scrap': float(wo.quantity_scrap) if wo.quantity_scrap else 0,
            'uom': wo.uom,
            'status': wo.status,
            'priority': wo.priority,
            'source_type': getattr(wo, 'source_type', 'manual'),  # manual, from_bom, from_schedule
            'schedule_grid_id': getattr(wo, 'schedule_grid_id', None),
            'schedule_days': {},  # Will be populated if from_schedule
            'batch_number': getattr(wo, 'batch_number', None),
            'machine_id': wo.machine_id,
            'machine_name': wo.machine.name if wo.machine else None,
            'machine_default_speed': int(wo.machine.default_speed) if wo.machine and wo.machine.default_speed else 0,
            'pack_per_carton': pack_per_carton,
            'consumption_data': consumption_data,
            'scheduled_start_date': wo.scheduled_start_date.isoformat() if wo.scheduled_start_date else None,
            'scheduled_end_date': wo.scheduled_end_date.isoformat() if wo.scheduled_end_date else None,
            'actual_start_date': wo.actual_start_date.isoformat() if wo.actual_start_date else None,
            'actual_end_date': wo.actual_end_date.isoformat() if wo.actual_end_date else None,
            'notes': wo.notes,
            'created_at': wo.created_at.isoformat() if wo.created_at else None
        }
        
        # Add supervisor info safely
        if hasattr(wo, 'supervisor_id'):
            response_data['supervisor_id'] = wo.supervisor_id
            response_data['supervisor_name'] = wo.supervisor.full_name if wo.supervisor else None
        
        # Add BOM materials used for production
        # If bom_id is not set, try to find active BOM for this product
        bom_materials = []
        bom_to_use = wo.bom
        if not wo.bom_id and wo.product_id:
            from models.production import BillOfMaterials
            active_bom = BillOfMaterials.query.filter_by(
                product_id=wo.product_id,
                is_active=True
            ).first()
            if active_bom:
                bom_to_use = active_bom
                response_data['bom_id'] = active_bom.id
                response_data['bom_number'] = active_bom.bom_number
        
        if bom_to_use:
            # Get pack_per_karton from products_new first (by code or name), then BOM, then ProductPackaging
            packs_per_karton = 1
            product_code = wo.product.code if wo.product else None
            product_name = wo.product.name if wo.product else None
            
            # Try by code first
            if product_code:
                product_new_packs = db.session.execute(
                    db.text('SELECT pack_per_karton FROM products WHERE code = :code'),
                    {'code': product_code}
                ).fetchone()
                if product_new_packs and product_new_packs[0]:
                    packs_per_karton = int(product_new_packs[0])
            
            # Try by name if code didn't work
            if packs_per_karton == 1 and product_name:
                search_name = product_name
                # Remove WIP prefix if present
                if search_name.upper().startswith('WIP '):
                    search_name = search_name[4:]
                
                product_new_packs = db.session.execute(
                    db.text('SELECT pack_per_karton FROM products WHERE name LIKE :name ORDER BY LENGTH(name) LIMIT 1'),
                    {'name': f'%{search_name}%'}
                ).fetchone()
                if product_new_packs and product_new_packs[0]:
                    packs_per_karton = int(product_new_packs[0])
            
            if packs_per_karton == 1:
                if bom_to_use.pack_per_carton and bom_to_use.pack_per_carton > 1:
                    packs_per_karton = bom_to_use.pack_per_carton
                elif wo.product_id:
                    packaging = ProductPackaging.query.filter_by(product_id=wo.product_id).first()
                    if packaging and packaging.packs_per_karton:
                        packs_per_karton = packaging.packs_per_karton
            
            response_data['packs_per_karton'] = packs_per_karton
            
            for item in bom_to_use.items:
                # BOM quantity is per karton, convert to per pack
                qty_per_batch_per_pack = (float(item.quantity) / packs_per_karton) if item.quantity else 0
                
                material_info = {
                    'id': item.id,
                    'line_number': item.line_number,
                    'item_code': item.item_code,
                    'item_name': item.item_name,
                    'item_type': item.item_type,
                    'quantity_per_batch': round(qty_per_batch_per_pack, 6),  # per pack now
                    'quantity_per_karton': float(item.quantity) if item.quantity else 0,  # original per karton
                    'uom': item.uom,
                    'scrap_percent': float(item.scrap_percent) if item.scrap_percent else 0,
                    'is_critical': item.is_critical,
                    'unit_cost': float(item.unit_cost) if item.unit_cost else 0,
                }
                
                # Calculate required quantity based on WO quantity (in packs) and BOM per pack
                # WO quantity is in packs, BOM is now converted to per pack
                wo_quantity = float(wo.quantity) if wo.quantity else 0
                required_qty = qty_per_batch_per_pack * wo_quantity
                effective_qty = required_qty * (1 + (float(item.scrap_percent) / 100)) if item.scrap_percent else required_qty
                
                material_info['required_quantity'] = round(required_qty, 4)
                material_info['effective_quantity'] = round(effective_qty, 4)
                material_info['total_cost'] = round(effective_qty * (float(item.unit_cost) if item.unit_cost else 0), 2)
                
                bom_materials.append(material_info)
        
        response_data['bom_materials'] = bom_materials
        response_data['batch_size'] = float(bom_to_use.batch_size) if bom_to_use and bom_to_use.batch_size else None
        
        # Get schedule_days if WO is from schedule
        if response_data.get('source_type') == 'from_schedule' and response_data.get('schedule_grid_id'):
            from routes.schedule_grid import ScheduleGridItem
            import json
            schedule_item = db.session.get(ScheduleGridItem, response_data['schedule_grid_id'])
            if schedule_item and schedule_item.schedule_days:
                response_data['schedule_days'] = json.loads(schedule_item.schedule_days)
        
        return jsonify({'work_order': response_data}), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/work-orders', methods=['POST'])
@jwt_required()
def create_work_order():
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        
        wo_number = generate_number('WO', WorkOrder, 'wo_number')
        
        # Get UOM from data or from product
        uom = data.get('uom')
        if not uom:
            from models.product import Product
            product = db.session.get(Product, data['product_id'])
            uom = product.primary_uom if product else 'pcs'
        
        # Determine source_type: manual, from_bom, from_schedule
        source_type = data.get('source_type', 'manual')
        bom_id = data.get('bom_id')
        
        wo = WorkOrder(
            wo_number=wo_number,
            product_id=data['product_id'],
            bom_id=bom_id,
            quantity=data['quantity'],
            uom=uom,
            pack_per_carton=data.get('pack_per_carton'),
            status='planned',
            priority=data.get('priority', 'normal'),
            source_type=source_type,  # manual, from_bom, from_schedule
            machine_id=data.get('machine_id'),
            scheduled_start_date=datetime.fromisoformat(data['scheduled_start_date']) if data.get('scheduled_start_date') else None,
            scheduled_end_date=datetime.fromisoformat(data['scheduled_end_date']) if data.get('scheduled_end_date') else None,
            notes=data.get('notes'),
            created_by=user_id
        )
        
        db.session.add(wo)
        db.session.commit()
        
        return jsonify({'message': 'Work order created', 'wo_id': wo.id, 'wo_number': wo_number}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/work-orders/<int:id>', methods=['PUT'])
@jwt_required()
def update_work_order(id):
    """Update work order"""
    try:
        wo = db.session.get(WorkOrder, id)
        if not wo:
            return jsonify({'error': 'Work order not found'}), 404
        
        data = request.get_json()
        force = data.get('force', False)
        
        # Only allow editing if status is planned or released, unless force
        if wo.status not in ['planned', 'released', 'in_progress'] and not force:
            return jsonify({'error': 'Cannot edit work order in current status'}), 400
        
        if 'product_id' in data:
            wo.product_id = data['product_id']
        if 'quantity' in data:
            wo.quantity = data['quantity']
        if 'priority' in data:
            wo.priority = data['priority']
        if 'machine_id' in data:
            wo.machine_id = data['machine_id']
        if 'uom' in data:
            wo.uom = data['uom']
        if 'pack_per_carton' in data:
            wo.pack_per_carton = data['pack_per_carton']
        if 'scheduled_start_date' in data:
            wo.scheduled_start_date = datetime.fromisoformat(data['scheduled_start_date']) if data['scheduled_start_date'] else None
        if 'scheduled_end_date' in data:
            wo.scheduled_end_date = datetime.fromisoformat(data['scheduled_end_date']) if data['scheduled_end_date'] else None
        if 'notes' in data:
            wo.notes = data['notes']
        if 'status' in data:
            wo.status = data['status']
        
        # Reset production counts if requested
        if data.get('reset_production'):
            wo.quantity_produced = 0
            wo.quantity_good = 0
            wo.quantity_scrap = 0
        
        db.session.commit()
        return jsonify({'message': 'Work order updated'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/work-orders/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_work_order(id):
    """Delete work order"""
    try:
        wo = db.session.get(WorkOrder, id)
        if not wo:
            return jsonify({'error': 'Work order not found'}), 404
        
        force = request.args.get('force', 'false').lower() == 'true'
        
        # Import all related models
        from models.production import ShiftProduction, PackingList, PackingListItem, ProductionSchedule, ProductionApproval, ProductChangeover, WeeklyProductionPlanItem
        from models.wip_job_costing import WIPBatch, WIPStageMovement, JobCostEntry
        
        # Check if there are production records
        has_production = ProductionRecord.query.filter_by(work_order_id=id).first() is not None
        has_shift_production = ShiftProduction.query.filter_by(work_order_id=id).first() is not None
        
        if (has_production or has_shift_production) and not force:
            return jsonify({
                'error': 'Work order has production records. Use force=true to delete anyway.',
                'has_production': True
            }), 400
        
        # Always delete related records to avoid FK constraint errors
        # 1. Delete packing lists and their items FIRST
        packing_lists = PackingList.query.filter_by(work_order_id=id).all()
        for pl in packing_lists:
            PackingListItem.query.filter_by(packing_list_id=pl.id).delete()
            db.session.delete(pl)
        
        # 2. Delete WIP batches and their related records
        wip_batches = WIPBatch.query.filter_by(work_order_id=id).all()
        for wb in wip_batches:
            WIPStageMovement.query.filter_by(wip_batch_id=wb.id).delete()
            JobCostEntry.query.filter_by(wip_batch_id=wb.id).delete()
            db.session.delete(wb)
        
        # 3. Delete production approvals
        ProductionApproval.query.filter_by(work_order_id=id).delete()
        
        # 4. Delete production schedules
        ProductionSchedule.query.filter_by(work_order_id=id).delete()
        
        # 5. Handle product changeovers
        ProductChangeover.query.filter_by(from_work_order_id=id).delete()
        ProductChangeover.query.filter_by(to_work_order_id=id).update({'to_work_order_id': None})
        
        # 6. Update weekly plan items to remove WO reference
        WeeklyProductionPlanItem.query.filter_by(work_order_id=id).update({'work_order_id': None})
        
        # 7. Delete production records if force
        if force:
            ProductionRecord.query.filter_by(work_order_id=id).delete()
            ShiftProduction.query.filter_by(work_order_id=id).delete()
        
        # 8. Recalculate machine efficiency after deleting ShiftProduction
        if wo.machine_id:
            machine = wo.machine
            if machine:
                # Get remaining ShiftProduction for this machine (last 30 days)
                from datetime import timedelta
                thirty_days_ago = datetime.utcnow().date() - timedelta(days=30)
                remaining_shifts = ShiftProduction.query.filter(
                    ShiftProduction.machine_id == machine.id,
                    ShiftProduction.production_date >= thirty_days_ago
                ).all()
                
                if remaining_shifts:
                    # Recalculate average efficiency and availability
                    total_efficiency = sum(float(sp.efficiency_rate or 0) for sp in remaining_shifts)
                    total_availability = sum(
                        (float(sp.actual_runtime or 0) / float(sp.planned_runtime) * 100) 
                        if sp.planned_runtime and sp.planned_runtime > 0 else 100 
                        for sp in remaining_shifts
                    )
                    machine.efficiency = round(total_efficiency / len(remaining_shifts), 2)
                    machine.availability = round(total_availability / len(remaining_shifts), 2)
                else:
                    # No remaining production data, reset to default
                    machine.efficiency = 100
                    machine.availability = 100
                machine.status = 'idle'
        
        db.session.delete(wo)
        db.session.commit()
        return jsonify({'message': 'Work order deleted successfully'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/work-orders/<int:id>/status', methods=['PUT'])
@jwt_required()
def update_work_order_status(id):
    """Update work order status with auto warehouse integration"""
    try:
        from routes.production_integration import auto_deduct_materials, auto_receive_finished_goods
        
        wo = db.session.get(WorkOrder, id)
        if not wo:
            return jsonify({'error': 'Work order not found'}), 404
        
        data = request.get_json()
        new_status = data.get('status')
        auto_deduct = data.get('auto_deduct', True)  # Default true for auto-deduction
        
        valid_statuses = ['planned', 'released', 'in_progress', 'completed', 'cancelled']
        if new_status not in valid_statuses:
            return jsonify({'error': f'Invalid status. Must be one of: {valid_statuses}'}), 400
        
        old_status = wo.status
        integration_results = {}
        
        # Update status
        wo.status = new_status
        
        # Set timestamps based on status
        if new_status == 'in_progress' and not wo.actual_start_date:
            wo.actual_start_date = datetime.utcnow()
            if wo.machine:
                wo.machine.status = 'running'
            
            # AUTO-DEDUCT MATERIALS when starting production
            # Guard: skip if materials already issued (e.g. via /start endpoint)
            if auto_deduct and wo.bom_id:
                from models.material_issue import MaterialIssue
                existing_mi = MaterialIssue.query.filter_by(
                    work_order_id=id,
                    status='issued'
                ).first()
                
                if existing_mi:
                    integration_results['material_deduction'] = {
                        'success': True,
                        'message': f'Materials already issued via {existing_mi.issue_number}',
                        'transactions': [],
                        'skipped': True
                    }
                else:
                    user_id = get_jwt_identity()
                    success, message, data_result = auto_deduct_materials(id, user_id)
                    integration_results['material_deduction'] = {
                        'success': success,
                        'message': message,
                        'transactions': data_result if success else [],
                        'insufficient_materials': data_result if not success else []
                    }
                    
                    if not success:
                        # Rollback status change if material deduction fails
                        wo.status = old_status
                        wo.actual_start_date = None
                        db.session.rollback()
                        return jsonify({
                            'error': 'Cannot start production: Material shortage',
                            'details': integration_results['material_deduction']
                        }), 400
        
        elif new_status == 'completed' and not wo.actual_end_date:
            wo.actual_end_date = datetime.utcnow()
            if wo.machine:
                wo.machine.status = 'idle'
            
            # AUTO-RECEIVE FINISHED GOODS when completing production
            if auto_deduct and wo.quantity_produced > 0:
                user_id = get_jwt_identity()
                qty_good = float(wo.quantity_good) if wo.quantity_good else float(wo.quantity_produced)
                success, message, inventory_id = auto_receive_finished_goods(id, qty_good, user_id)
                integration_results['finished_goods_receipt'] = {
                    'success': success,
                    'message': message,
                    'inventory_id': inventory_id,
                    'quantity_received': qty_good
                }
        
        db.session.commit()
        
        response = {
            'message': f'Work order status updated to {new_status}',
            'wo_number': wo.wo_number,
            'old_status': old_status,
            'new_status': new_status
        }
        
        if integration_results:
            response['integration_results'] = integration_results
        
        return jsonify(response), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/work-orders/<int:id>/production-records', methods=['GET'])
@jwt_required()
def get_work_order_production_records(id):
    """Get production records for a specific work order, plus same-machine records for shift usage"""
    try:
        from models.production import ShiftProduction
        wo = db.session.get(WorkOrder, id)
        if not wo:
            return jsonify({'error': 'Work order not found'}), 404
        
        # Get this WO's records
        records = ProductionRecord.query.filter_by(work_order_id=id).order_by(ProductionRecord.production_date.desc()).all()
        
        # Also get ShiftProduction records from the same machine (for shift usage calculation)
        # ShiftProduction has planned_runtime, actual_runtime, idle_time that we need
        machine_shift_records = []
        if wo.machine_id:
            machine_shift_records = ShiftProduction.query.filter(
                ShiftProduction.machine_id == wo.machine_id,
            ).order_by(ShiftProduction.production_date.desc()).limit(100).all()
        
        def format_record(r):
            # Get product name from record's own product or WO's product
            product_name = None
            product_id = r.product_id
            if r.product:
                product_name = r.product.name
            elif r.work_order and r.work_order.product:
                product_name = r.work_order.product.name
                product_id = r.work_order.product_id
            
            return {
                'id': r.id,
                'work_order_id': r.work_order_id,
                'product_id': product_id,
                'product_name': product_name,
                'production_date': r.production_date.isoformat() if r.production_date else None,
                'shift': r.shift,
                'quantity_produced': float(r.quantity_produced) if r.quantity_produced else 0,
                'quantity_good': float(r.quantity_good) if r.quantity_good else 0,
                'quantity_reject': float(r.quantity_scrap) if r.quantity_scrap else 0,
                'setting_sticker': float(r.setting_sticker) if hasattr(r, 'setting_sticker') and r.setting_sticker else 0,
                'setting_packaging': float(r.setting_packaging) if hasattr(r, 'setting_packaging') and r.setting_packaging else 0,
                'downtime_minutes': r.downtime_minutes or 0,
                'operator_name': r.operator.full_name if r.operator else None,
                'notes': r.notes
            }
        
        def format_shift_record(sp):
            product_name = sp.product.name if sp.product else None
            return {
                'id': sp.id,
                'work_order_id': sp.work_order_id,
                'production_date': sp.production_date.isoformat() if sp.production_date else None,
                'shift': sp.shift,
                'quantity_good': float(sp.good_quantity) if sp.good_quantity else 0,
                'downtime_minutes': sp.downtime_minutes or 0,
                'runtime': sp.actual_runtime or 0,
                'average_time': sp.planned_runtime or 0,
                'idle_time': sp.idle_time or 0,
                'product_name': product_name,
            }
        
        return jsonify({
            'records': [format_record(r) for r in records],
            'machine_records': [format_shift_record(sp) for sp in machine_shift_records]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@production_bp.route('/work-orders/<int:id>/production-records', methods=['POST'])
@jwt_required()
def create_work_order_production_record(id):
    """Create production record for a work order and ShiftProduction for OEE tracking"""
    try:
        from models.production import ShiftProduction
        from models.wip_job_costing import WIPBatch, JobCostEntry, WIPWorkflowIntegration
        
        user_id = get_jwt_identity()
        wo = db.session.get(WorkOrder, id)
        if not wo:
            return jsonify({'error': 'Work order not found'}), 404
        
        data = request.get_json()
        
        # Debug logging
        print(f"[DEBUG] Production record creation - WO ID: {id}")
        print(f"[DEBUG] Data keys: {list(data.keys()) if data else 'None'}")
        print(f"[DEBUG] Work Order: {wo.wo_number}, Status: {wo.status}")
        
        # Build downtime notes from entries
        downtime_entries = data.get('downtime_entries', [])
        downtime_notes = ''
        if downtime_entries:
            downtime_notes = '; '.join([
                f"{e.get('total_minutes', e.get('duration_minutes', 0))} menit - {e.get('reason', '')} [{e.get('category', 'others')}]" 
                for e in downtime_entries
            ])
        
        # Parse production date
        production_date = datetime.fromisoformat(data['production_date']).date() if data.get('production_date') else datetime.utcnow().date()
        
        # Get machine_id from work order or request data
        machine_id = wo.machine_id or data.get('machine_id')
        if not machine_id:
            return jsonify({'error': 'Machine ID is required for production record'}), 400
        
        # Create production record
        # Use product_id from request if provided (for multi-product per shift)
        record_product_id = data.get('product_id') or wo.product_id
        
        record = ProductionRecord(
            work_order_id=id,
            machine_id=machine_id,
            product_id=record_product_id,  # Store product_id for multi-product per shift
            production_date=datetime.fromisoformat(data['production_date']) if data.get('production_date') else datetime.utcnow(),
            shift=data.get('shift', '1'),
            quantity_produced=data.get('quantity_produced', 0),
            quantity_good=data.get('quantity_good', 0),
            quantity_scrap=data.get('quantity_reject', 0),  # Model uses quantity_scrap
            downtime_minutes=data.get('downtime_minutes', 0),
            operator_id=data.get('operator_id'),
            notes=data.get('notes'),
            uom=data.get('uom', 'pcs')
        )
        
        # Store downtime entries as JSON in notes if needed
        if downtime_notes:
            record.notes = f"{data.get('notes', '')}\n\n[Downtime Details]\n{downtime_notes}".strip()
        
        db.session.add(record)
        
        # Get downtime by category from request
        downtime_mesin = int(data.get('downtime_mesin', 0))
        downtime_operator = int(data.get('downtime_operator', 0))
        downtime_material = int(data.get('downtime_material', 0))
        downtime_design = int(data.get('downtime_design', 0))
        downtime_others = int(data.get('downtime_others', 0))
        total_downtime = int(data.get('downtime_minutes', 0))
        
        # Get quantities first (needed for runtime calculation)
        actual_qty = float(data.get('quantity_produced', 0))
        good_qty = float(data.get('quantity_good', 0))
        
        # Get average_time from request (manual input, default 510)
        average_time = int(data.get('average_time', 510))
        
        # Calculate runtime = Grade A / speed mesin
        machine_speed = int(data.get('machine_speed', 0))
        runtime = int(data.get('runtime', 0))
        if runtime == 0 and machine_speed > 0:
            runtime = int(round(good_qty / machine_speed))
        
        # Waktu tercatat = runtime + downtime
        waktu_tercatat = int(data.get('waktu_tercatat', runtime + total_downtime))
        
        # Waktu tidak tercatat = average_time - waktu_tercatat
        waktu_tidak_tercatat = int(data.get('waktu_tidak_tercatat', max(0, average_time - waktu_tercatat)))
        
        # Keep planned_runtime for backward compatibility
        shift_num = data.get('shift', '1')
        is_friday = production_date.weekday() == 4  # 4 = Friday
        default_runtime = 510
        if shift_num in ['1', 'shift_1']:
            default_runtime = 540 if is_friday else 510
        elif shift_num in ['2', 'shift_2']:
            default_runtime = 480
        elif shift_num in ['3', 'shift_3']:
            default_runtime = 450
        
        planned_runtime = average_time  # Use average_time as planned_runtime
        actual_runtime = runtime  # Use calculated runtime as actual_runtime
        
        # Calculate quality rate
        quality_rate = (good_qty / actual_qty * 100) if actual_qty > 0 else 100
        
        # Calculate efficiency rate (100% - total loss from downtime)
        efficiency_rate = float(data.get('efficiency_rate', 100))
        
        # Calculate OEE score
        oee_score = (efficiency_rate * quality_rate) / 100
        
        # Calculate loss percentages
        def calc_loss(downtime_min, planned_min):
            return round((downtime_min / planned_min * 100), 2) if planned_min > 0 else 0
        
        loss_mesin = calc_loss(downtime_mesin, planned_runtime)
        loss_operator = calc_loss(downtime_operator, planned_runtime)
        loss_material = calc_loss(downtime_material, planned_runtime)
        loss_design = calc_loss(downtime_design, planned_runtime)
        loss_others = calc_loss(downtime_others, planned_runtime)
        
        # Shift times based on shift number and day
        # Shift 1: 06:30-15:00 (normal), 06:30-15:30 (Friday)
        # Shift 2: 15:00-23:00
        # Shift 3: 23:00-06:30
        shift_key = f"shift_{data.get('shift', '1')}"
        
        # Determine shift times based on day and shift
        if shift_num in ['1', 'shift_1']:
            shift_start_str = '06:30'
            shift_end_str = '15:30' if is_friday else '15:00'
        elif shift_num in ['2', 'shift_2']:
            shift_start_str = '15:00'
            shift_end_str = '23:00'
        elif shift_num in ['3', 'shift_3']:
            shift_start_str = '23:00'
            shift_end_str = '06:30'
        else:
            # Default to shift 1
            shift_start_str = '06:30'
            shift_end_str = '15:30' if is_friday else '15:00'
        
        shift_start = datetime.strptime(shift_start_str, '%H:%M').time()
        shift_end = datetime.strptime(shift_end_str, '%H:%M').time()
        
        # Create or update ShiftProduction for OEE tracking
        # Use product_id from request if provided (for multi-product per shift), otherwise use WO's product
        product_id = data.get('product_id') or wo.product_id
        
        shift_key_value = shift_key if shift_key.startswith('shift_') else f"shift_{data.get('shift', '1')}"
        
        # Double-click guard: skip if exact same record created within last 60 seconds
        from sqlalchemy import and_
        recent_dup = ShiftProduction.query.filter(
            and_(
                ShiftProduction.machine_id == machine_id,
                ShiftProduction.production_date == production_date,
                ShiftProduction.shift == shift_key_value,
                ShiftProduction.work_order_id == id,
                ShiftProduction.product_id == product_id,
                ShiftProduction.actual_quantity == actual_qty,
                ShiftProduction.good_quantity == good_qty,
                ShiftProduction.created_at >= datetime.utcnow() - timedelta(seconds=60)
            )
        ).first()
        
        if recent_dup:
            print(f"[DEBUG] Skipped duplicate ShiftProduction (same data within 60s) ID={recent_dup.id}")
            shift_production = recent_dup
        else:
            # CREATE new ShiftProduction (always new - allows same product multiple times per shift)
            # Get sub_shift from request (manual selection) - NULL means legacy/auto-detect
            sub_shift = data.get('sub_shift')  # 'a', 'b', 'c' or None
            
            shift_production = ShiftProduction(
                production_date=production_date,
                shift=shift_key_value,
                sub_shift=sub_shift,
                shift_start=shift_start,
                shift_end=shift_end,
                machine_id=machine_id,
                product_id=product_id,
                work_order_id=id,
                target_quantity=wo.quantity,
                actual_quantity=actual_qty,
                good_quantity=good_qty,
                reject_quantity=float(data.get('quantity_reject', 0)),
                rework_quantity=float(data.get('quantity_rework', 0)),
                setting_sticker=float(data.get('setting_sticker', 0)),
                setting_packaging=float(data.get('setting_packaging', 0)),
                uom=data.get('uom', 'pcs'),
                planned_runtime=planned_runtime,
                actual_runtime=actual_runtime,
                downtime_minutes=total_downtime,
                machine_speed=int(data.get('machine_speed', 0)),
                downtime_mesin=downtime_mesin,
                downtime_operator=downtime_operator,
                downtime_material=downtime_material,
                downtime_design=downtime_design,
                downtime_others=downtime_others,
                idle_time=int(data.get('idle_time', 0)),
                waktu_tidak_tercatat=int(data.get('waktu_tidak_tercatat', 0)),
                loss_mesin=loss_mesin,
                loss_operator=loss_operator,
                loss_material=loss_material,
                loss_design=loss_design,
                loss_others=loss_others,
                quality_rate=round(quality_rate, 2),
                efficiency_rate=round(efficiency_rate, 2),
                oee_score=round(oee_score, 2),
                operator_id=data.get('operator_id'),
                notes=data.get('notes'),
                issues=downtime_notes if downtime_notes else None,
                status='completed',
                created_by=user_id,
                early_stop=data.get('early_stop', False),
                early_stop_time=datetime.strptime(data.get('early_stop_time'), '%H:%M').time() if data.get('early_stop_time') else None,
                early_stop_reason=data.get('early_stop_reason'),
                early_stop_notes=data.get('early_stop_notes'),
                operator_reassigned=data.get('operator_reassigned', False),
                reassignment_task=data.get('reassignment_task'),
                reassignment_notes=data.get('reassignment_notes'),
                pack_per_carton=int(data.get('pack_per_carton', 0)) or (wo.pack_per_carton if wo.pack_per_carton else 0)
            )
            db.session.add(shift_production)
            print(f"[DEBUG] Created new ShiftProduction")
        
        # Update work order quantities - convert to float to avoid Decimal conflicts
        wo.quantity_produced = float(wo.quantity_produced or 0) + float(data.get('quantity_produced', 0))
        wo.quantity_good = float(wo.quantity_good or 0) + float(data.get('quantity_good', 0))
        wo.quantity_scrap = float(wo.quantity_scrap or 0) + float(data.get('quantity_reject', 0))
        
        # Calculate buffer stock (production exceeding target)
        buffer_stock_qty = 0
        if wo.quantity_good > wo.quantity:
            buffer_stock_qty = float(wo.quantity_good) - float(wo.quantity)
        
        # Add buffer stock to inventory as finished goods
        if buffer_stock_qty > 0:
            from models.warehouse import Inventory, InventoryMovement, WarehouseLocation
            
            # Find or create inventory for this product
            # First try to find existing inventory for this product
            inventory = Inventory.query.filter_by(
                product_id=wo.product_id,
                stock_status='released'
            ).first()
            
            if not inventory:
                # Find a suitable location for finished goods
                fg_location = WarehouseLocation.query.filter(
                    WarehouseLocation.material_type.in_(['finished_goods', 'all'])
                ).first()
                
                if fg_location:
                    inventory = Inventory(
                        product_id=wo.product_id,
                        location_id=fg_location.id,
                        quantity_on_hand=0,
                        quantity_reserved=0,
                        quantity_available=0,
                        stock_status='released',
                        is_active=True,
                        created_by=user_id
                    )
                    db.session.add(inventory)
                    db.session.flush()
            
            if inventory:
                # Add buffer stock to inventory - convert to float to avoid Decimal conflicts
                inventory.quantity_on_hand = float(inventory.quantity_on_hand or 0) + float(buffer_stock_qty)
                inventory.quantity_available = float(inventory.quantity_available or 0) + float(buffer_stock_qty)
                inventory.updated_at = datetime.utcnow()
                
                # Create inventory movement for buffer stock
                buffer_movement = InventoryMovement(
                    inventory_id=inventory.id,
                    movement_type='stock_in',
                    movement_date=production_date,
                    quantity=buffer_stock_qty,
                    reference_number=wo.wo_number,
                    reference_type='buffer_stock',
                    notes=f'Buffer stock dari WO {wo.wo_number} - produksi melebihi target',
                    created_by=user_id
                )
                db.session.add(buffer_movement)
        
        # Auto-complete if target reached
        if wo.quantity_produced >= wo.quantity:
            wo.status = 'completed'
            wo.actual_end_date = datetime.utcnow()
        
        # Update machine efficiency and availability based on latest production
        machine = db.session.get(Machine, machine_id)
        if machine:
            machine.efficiency = round(efficiency_rate, 2)
            machine.availability = round((actual_runtime / planned_runtime * 100) if planned_runtime > 0 else 100, 2)
            machine.status = 'running' if wo.status == 'in_progress' else machine.status
        
        # ============= JOB COSTING INTEGRATION =============
        # Get or create WIP batch for this work order
        wip_batch = WIPBatch.query.filter_by(work_order_id=id).first()
        if not wip_batch:
            wip_batch = WIPBatch(
                wip_batch_no=f"WIP-{wo.wo_number}",
                work_order_id=id,
                product_id=wo.product_id,
                current_stage='production',
                qty_started=float(wo.quantity),
                qty_in_process=float(wo.quantity),
                status='in_progress',
                machine_id=machine_id,
                shift=data.get('shift', '1'),
                created_by=int(user_id)
            )
            db.session.add(wip_batch)
            db.session.flush()
        
        # Update WIP batch quantities
        wip_batch.qty_completed = float(wo.quantity_good or 0)
        wip_batch.qty_rejected = float(wo.quantity_scrap or 0)
        wip_batch.qty_in_process = float(wo.quantity) - wip_batch.qty_completed - wip_batch.qty_rejected
        
        # Calculate labor cost (based on actual runtime and labor rate)
        labor_rate_per_hour = float(data.get('labor_rate', 25000))  # Default Rp 25.000/jam
        labor_hours = actual_runtime / 60
        labor_cost = labor_hours * labor_rate_per_hour
        
        # Calculate overhead cost (machine cost per hour)
        overhead_rate_per_hour = float(data.get('overhead_rate', 50000))  # Default Rp 50.000/jam
        overhead_cost = labor_hours * overhead_rate_per_hour
        
        # Create job cost entry for labor
        if labor_cost > 0:
            labor_job_cost = JobCostEntry(
                job_cost_no=generate_number('JC', JobCostEntry, 'job_cost_no'),
                wip_batch_id=wip_batch.id,
                work_order_id=id,
                cost_type='labor',
                cost_category='direct_labor',
                description=f'Labor cost - Shift {data.get("shift", "1")} - {production_date}',
                quantity=labor_hours,
                unit_cost=labor_rate_per_hour,
                total_cost=labor_cost,
                allocation_basis='per_hour',
                allocation_rate=labor_rate_per_hour,
                cost_date=datetime.utcnow(),
                production_stage='production',
                shift=f'shift_{data.get("shift", "1")}',
                created_by=int(user_id)
            )
            db.session.add(labor_job_cost)
            # Convert to float to avoid Decimal conflicts
            wip_batch.labor_cost = float(wip_batch.labor_cost or 0) + float(labor_cost)
        
        # Create job cost entry for overhead
        if overhead_cost > 0:
            overhead_job_cost = JobCostEntry(
                job_cost_no=generate_number('JC', JobCostEntry, 'job_cost_no'),
                wip_batch_id=wip_batch.id,
                work_order_id=id,
                cost_type='overhead',
                cost_category='machine_overhead',
                description=f'Machine overhead - {wo.machine.name if wo.machine else "N/A"} - {production_date}',
                quantity=labor_hours,
                unit_cost=overhead_rate_per_hour,
                total_cost=overhead_cost,
                allocation_basis='per_hour',
                allocation_rate=overhead_rate_per_hour,
                cost_date=datetime.utcnow(),
                production_stage='production',
                shift=f'shift_{data.get("shift", "1")}',
                created_by=int(user_id)
            )
            db.session.add(overhead_job_cost)
            # Convert to float to avoid Decimal conflicts
            wip_batch.overhead_cost = float(wip_batch.overhead_cost or 0) + float(overhead_cost)
        
        # Update total WIP value
        wip_batch.update_wip_value()
        
        # Mark WIP as completed if work order is completed
        if wo.status == 'completed':
            wip_batch.status = 'completed'
            wip_batch.completed_at = datetime.utcnow()
        
        # ============= WIP STOCK FOR PACKING LIST =============
        # Add good quantity to WIP Stock ONLY for WIP products
        # FG products don't need WIP stock - their availability is calculated from WIP components via BOM
        from models.production import WIPStock, WIPStockMovement
        
        good_qty_produced = float(data.get('quantity_good', 0))
        is_wip_product = (wo.product and (
            wo.product.material_type == 'wip' or 
            (wo.product.name or '').startswith('WIP ')
        ))
        if good_qty_produced > 0 and is_wip_product:
            # Get pack_per_carton from products_new table (same as WorkOrder detail)
            pack_per_carton = 1
            product_code = wo.product.code if wo.product else None
            product_name = wo.product.name if wo.product else None
            
            if product_code:
                product_new_data = db.session.execute(
                    db.text('SELECT pack_per_karton FROM products WHERE code = :code'),
                    {'code': product_code}
                ).fetchone()
                if product_new_data and product_new_data[0]:
                    pack_per_carton = int(product_new_data[0])
            
            # Fallback: try by product name - prioritize standard @24 packaging
            if pack_per_carton == 1 and product_name:
                search_name = product_name
                if search_name.upper().startswith('WIP '):
                    search_name = search_name[4:]
                # First try exact match with @24 suffix (standard packaging)
                product_new_data = db.session.execute(
                    db.text("SELECT pack_per_karton FROM products WHERE name LIKE :name ORDER BY CASE WHEN name LIKE '%@24' THEN 0 WHEN name LIKE '%@24%' THEN 1 ELSE 2 END, pack_per_karton DESC LIMIT 1"),
                    {'name': f'%{search_name}%'}
                ).fetchone()
                if product_new_data and product_new_data[0]:
                    pack_per_carton = int(product_new_data[0])
            
            # Calculate cartons (full cartons only for WIP tracking)
            carton_qty = int(good_qty_produced / pack_per_carton) if pack_per_carton > 0 else 0
            pcs_qty = int(good_qty_produced)
            
            if carton_qty > 0 or pcs_qty > 0:
                # Get or create WIP stock for this product
                wip_stock = WIPStock.query.filter_by(product_id=wo.product_id).first()
                
                if not wip_stock:
                    wip_stock = WIPStock(
                        product_id=wo.product_id,
                        quantity_pcs=0,
                        quantity_carton=0,
                        pack_per_carton=pack_per_carton,
                        last_wo_number=wo.wo_number
                    )
                    db.session.add(wip_stock)
                    db.session.flush()
                
                # Update WIP stock quantities
                old_balance_pcs = wip_stock.quantity_pcs
                old_balance_carton = wip_stock.quantity_carton
                
                # Convert to float to avoid Decimal conflicts
                wip_stock.quantity_pcs = float(wip_stock.quantity_pcs or 0) + float(pcs_qty)
                wip_stock.quantity_carton = float(wip_stock.quantity_carton or 0) + float(carton_qty)
                wip_stock.last_wo_number = wo.wo_number
                wip_stock.last_updated_at = datetime.utcnow()
                
                # Create movement record
                wip_movement = WIPStockMovement(
                    wip_stock_id=wip_stock.id,
                    product_id=wo.product_id,
                    movement_type='in',
                    quantity_pcs=pcs_qty,
                    quantity_carton=carton_qty,
                    balance_pcs=wip_stock.quantity_pcs,
                    balance_carton=wip_stock.quantity_carton,
                    reference_type='work_order',
                    reference_id=wo.id,
                    reference_number=wo.wo_number,
                    notes=f'Hasil produksi shift {data.get("shift", "1")} - {production_date}',
                    created_by=user_id
                )
                db.session.add(wip_movement)
        
        db.session.commit()
        
        # ============= NOTIFICATIONS =============
        try:
            supervisor_ids = get_supervisor_user_ids()
            product_name = wo.product.name if wo.product else 'Unknown'
            
            # Notification for production input
            create_production_notification(
                user_ids=supervisor_ids,
                notification_type='info',
                category='production',
                title='Input Produksi Baru',
                message=f'Input produksi untuk {wo.wo_number} - {product_name}. Grade A: {good_qty:,.0f} pcs, Efisiensi: {efficiency_rate:.1f}%',
                reference_type='work_order',
                reference_id=wo.id,
                priority='normal',
                action_url=f'/app/production/work-orders/{wo.id}'
            )
            
            # Notification if work order completed
            if wo.status == 'completed':
                create_production_notification(
                    user_ids=supervisor_ids,
                    notification_type='success',
                    category='production',
                    title='Work Order Selesai',
                    message=f'Work Order {wo.wo_number} - {product_name} telah selesai. Total produksi: {wo.quantity_good:,.0f} pcs',
                    reference_type='work_order',
                    reference_id=wo.id,
                    priority='high',
                    action_url=f'/app/production/work-orders/{wo.id}'
                )
        except Exception as notif_err:
            print(f"Notification error (non-critical): {notif_err}")
        
        return jsonify({
            'message': 'Production record created',
            'record_id': record.id,
            'shift_production_id': shift_production.id,
            'work_order_progress': {
                'quantity_produced': float(wo.quantity_produced),
                'quantity_good': float(wo.quantity_good),
                'quantity_scrap': float(wo.quantity_scrap),
                'target_quantity': float(wo.quantity),
                'status': wo.status
            },
            'buffer_stock': {
                'quantity': buffer_stock_qty,
                'added_to_inventory': buffer_stock_qty > 0,
                'message': f'{buffer_stock_qty} unit ditambahkan ke stok barang jadi' if buffer_stock_qty > 0 else None
            },
            'oee_data': {
                'efficiency_rate': round(efficiency_rate, 2),
                'quality_rate': round(quality_rate, 2),
                'oee_score': round(oee_score, 2),
                'downtime_breakdown': {
                    'mesin': {'minutes': downtime_mesin, 'loss_pct': loss_mesin},
                    'operator': {'minutes': downtime_operator, 'loss_pct': loss_operator},
                    'material': {'minutes': downtime_material, 'loss_pct': loss_material},
                    'design': {'minutes': downtime_design, 'loss_pct': loss_design},
                    'others': {'minutes': downtime_others, 'loss_pct': loss_others}
                }
            },
            'job_costing': {
                'wip_batch_id': wip_batch.id,
                'wip_batch_no': wip_batch.wip_batch_no,
                'labor_cost': round(labor_cost, 2),
                'overhead_cost': round(overhead_cost, 2),
                'total_wip_value': round(wip_batch.total_wip_value, 2),
                'material_cost': round(wip_batch.material_cost, 2)
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        # Detailed error logging for debugging
        import traceback
        error_detail = {
            'error': str(e),
            'type': type(e).__name__,
            'traceback': traceback.format_exc(),
            'data_keys': list(data.keys()) if 'data' in locals() else []
        }
        print(f"[ERROR] Production record creation failed: {error_detail}")
        return jsonify({'error': str(e), 'debug_info': error_detail}), 500

@production_bp.route('/work-orders/<int:id>/start', methods=['PUT'])
@jwt_required()
def start_work_order(id):
    try:
        from models.material_issue import MaterialIssue, MaterialIssueItem
        from models.warehouse import Inventory, InventoryMovement
        from models.production import BillOfMaterials
        from models.wip_job_costing import WIPBatch, JobCostEntry
        from models.product import Material
        
        user_id = int(get_jwt_identity())
        total_material_cost = 0  # Track total material cost
        
        wo = db.session.get(WorkOrder, id)
        if not wo:
            return jsonify(error_response('api.error', error_code=404)), 404
        
        # Auto-issue materials from BOM
        bom = BillOfMaterials.query.filter_by(product_id=wo.product_id, is_active=True).first()
        
        if bom:
            # Check opname lock for all BOM materials before issuing
            from utils.opname_lock import check_opname_lock
            for bom_item in bom.items:
                lock = check_opname_lock(material_id=bom_item.material_id)
                if lock['locked']:
                    return jsonify({'error': lock['message']}), 423
            
            # Check if material issue already exists (prevent double deduct)
            existing_mi = MaterialIssue.query.filter_by(
                work_order_id=id,
                status='issued'
            ).first()
            
            if not existing_mi:
                # Create and auto-issue materials
                issue_number = generate_number('MI', MaterialIssue, 'issue_number')
                
                mi = MaterialIssue(
                    issue_number=issue_number,
                    work_order_id=id,
                    issue_date=datetime.utcnow(),
                    requested_by=user_id,
                    approved_by=user_id,
                    issued_by=user_id,
                    status='issued',
                    priority='normal',
                    issue_type='production',
                    approved_date=datetime.utcnow(),
                    issued_date=datetime.utcnow(),
                    notes=f'Auto-issued when starting Work Order {wo.wo_number}'
                )
                db.session.add(mi)
                db.session.flush()
                
                # Issue materials from BOM
                for idx, bom_item in enumerate(bom.items, 1):
                    required_qty = float(bom_item.quantity) * float(wo.quantity)
                    
                    # Find inventory with stock
                    inventories = Inventory.query.filter(
                        Inventory.material_id == bom_item.material_id,
                        Inventory.quantity_on_hand > 0,
                        Inventory.is_active == True
                    ).order_by(Inventory.expiry_date.asc().nullslast()).all()
                    
                    remaining_to_issue = required_qty
                    
                    # Find production zone for transfer (if exists)
                    from models.warehouse import WarehouseZone, WarehouseLocation
                    production_zone = WarehouseZone.query.filter_by(
                        zone_type='production',
                        is_active=True
                    ).first()
                    production_location = None
                    if production_zone:
                        production_location = WarehouseLocation.query.filter_by(
                            zone_id=production_zone.id,
                            is_active=True
                        ).first()
                    
                    for inv in inventories:
                        if remaining_to_issue <= 0:
                            break
                        
                        issue_qty = min(float(inv.quantity_on_hand), remaining_to_issue)
                        source_location_id = inv.location_id
                        
                        # Deduct from source inventory
                        inv.quantity_on_hand -= issue_qty
                        if inv.quantity_available >= issue_qty:
                            inv.quantity_available -= issue_qty
                        inv.updated_at = datetime.utcnow()
                        
                        # Create stock_out movement from source
                        movement_out = InventoryMovement(
                            inventory_id=inv.id,
                            movement_type='transfer_out',
                            movement_date=datetime.utcnow().date(),
                            quantity=issue_qty,
                            reference_number=issue_number,
                            reference_type='material_issue',
                            batch_number=inv.batch_number,
                            notes=f'Transfer to production for WO {wo.wo_number}',
                            created_by=user_id
                        )
                        db.session.add(movement_out)
                        
                        # If production location exists, create inventory there
                        if production_location:
                            # Check if inventory already exists in production location
                            prod_inv = Inventory.query.filter_by(
                                material_id=bom_item.material_id,
                                location_id=production_location.id,
                                batch_number=inv.batch_number
                            ).first()
                            
                            if prod_inv:
                                # Convert to float to avoid Decimal conflicts
                                prod_inv.quantity_on_hand = float(prod_inv.quantity_on_hand or 0) + float(issue_qty)
                                prod_inv.quantity_available = float(prod_inv.quantity_available or 0) + float(issue_qty)
                            else:
                                prod_inv = Inventory(
                                    material_id=bom_item.material_id,
                                    location_id=production_location.id,
                                    quantity_on_hand=issue_qty,
                                    quantity_available=issue_qty,
                                    batch_number=inv.batch_number,
                                    lot_number=inv.lot_number,
                                    work_order_id=id,
                                    stock_status='in_production',
                                    is_active=True
                                )
                                db.session.add(prod_inv)
                            
                            db.session.flush()
                            
                            # Create transfer_in movement to production
                            movement_in = InventoryMovement(
                                inventory_id=prod_inv.id,
                                movement_type='transfer_in',
                                movement_date=datetime.utcnow().date(),
                                quantity=issue_qty,
                                reference_number=issue_number,
                                reference_type='material_issue',
                                batch_number=inv.batch_number,
                                notes=f'Received from storage for WO {wo.wo_number}',
                                created_by=user_id
                            )
                            db.session.add(movement_in)
                        
                        remaining_to_issue -= issue_qty
                    
                    # Calculate material cost
                    issued_qty = required_qty - remaining_to_issue
                    material = db.session.get(Material, bom_item.material_id)
                    unit_cost = float(material.unit_cost) if material and material.unit_cost else 0
                    item_cost = issued_qty * unit_cost
                    total_material_cost += item_cost
                    
                    # Create material issue item
                    item = MaterialIssueItem(
                        material_issue_id=mi.id,
                        line_number=idx,
                        material_id=bom_item.material_id,
                        description=bom_item.material.name if bom_item.material else '',
                        required_quantity=required_qty,
                        issued_quantity=issued_qty,
                        uom=bom_item.uom,
                        unit_cost=unit_cost,
                        total_cost=item_cost,
                        status='issued' if remaining_to_issue <= 0 else 'partial'
                    )
                    db.session.add(item)
        
        wo.status = 'in_progress'
        wo.actual_start_date = datetime.utcnow()
        
        if wo.machine:
            wo.machine.status = 'running'
        
        # ============= CREATE WIP BATCH WITH MATERIAL COST =============
        wip_batch = WIPBatch.query.filter_by(work_order_id=id).first()
        if not wip_batch:
            wip_batch = WIPBatch(
                wip_batch_no=f"WIP-{wo.wo_number}",
                work_order_id=id,
                product_id=wo.product_id,
                current_stage='production',
                qty_started=float(wo.quantity),
                qty_in_process=float(wo.quantity),
                material_cost=total_material_cost,
                status='in_progress',
                machine_id=wo.machine_id,
                created_by=user_id
            )
            wip_batch.update_wip_value()
            db.session.add(wip_batch)
            db.session.flush()
            
            # Create job cost entry for material
            if total_material_cost > 0:
                material_job_cost = JobCostEntry(
                    job_cost_no=generate_number('JC', JobCostEntry, 'job_cost_no'),
                    wip_batch_id=wip_batch.id,
                    work_order_id=id,
                    cost_type='material',
                    cost_category='raw_material',
                    description=f'Material cost for WO {wo.wo_number}',
                    quantity=float(wo.quantity),
                    unit_cost=total_material_cost / float(wo.quantity) if wo.quantity > 0 else 0,
                    total_cost=total_material_cost,
                    allocation_basis='per_batch',
                    cost_date=datetime.utcnow(),
                    production_stage='material_issue',
                    created_by=user_id
                )
                db.session.add(material_job_cost)
        
        db.session.commit()
        return jsonify({
            'success': True,
            'message': 'Work order started and materials issued',
            'material_cost': round(total_material_cost, 2),
            'wip_batch_no': wip_batch.wip_batch_no if wip_batch else None
        }), 200
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/work-orders/<int:id>/complete', methods=['PUT'])
@jwt_required()
def complete_work_order(id):
    try:
        from routes.production_integration import auto_receive_finished_goods
        
        wo = db.session.get(WorkOrder, id)
        if not wo:
            return jsonify(error_response('api.error', error_code=404)), 404
        
        if wo.status == 'completed':
            return jsonify({'error': 'Work order already completed'}), 400
        
        data = request.get_json() or {}
        user_id = int(get_jwt_identity())
        
        wo.status = 'completed'
        wo.actual_end_date = datetime.utcnow()
        
        if wo.machine:
            wo.machine.status = 'idle'
        
        # Auto-receive finished goods to warehouse
        integration_results = {}
        qty_good = float(data.get('quantity_good', 0)) or (float(wo.quantity_good) if wo.quantity_good else 0) or (float(wo.quantity_produced) if wo.quantity_produced else 0)
        
        if qty_good > 0:
            success, message, inventory_id = auto_receive_finished_goods(id, qty_good, user_id)
            integration_results['finished_goods_receipt'] = {
                'success': success,
                'message': message,
                'inventory_id': inventory_id,
                'quantity_received': qty_good
            }
        
        db.session.commit()
        
        response = {
            'success': True,
            'message': f'Work order completed. {qty_good} units received to warehouse.' if qty_good > 0 else 'Work order completed.',
            'wo_number': wo.wo_number,
            'quantity_received': qty_good
        }
        if integration_results:
            response['integration_results'] = integration_results
        
        return jsonify(response), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/work-orders/bulk-complete', methods=['PUT'])
@jwt_required()
def bulk_complete_work_orders():
    """Complete all in_progress work orders at once"""
    try:
        from routes.production_integration import auto_receive_finished_goods
        
        user_id = int(get_jwt_identity())
        
        # Get all in_progress work orders
        in_progress_wos = WorkOrder.query.filter_by(status='in_progress').all()
        
        if not in_progress_wos:
            return jsonify({'message': 'Tidak ada Work Order in_progress', 'completed': 0}), 200
        
        completed_list = []
        errors = []
        
        for wo in in_progress_wos:
            try:
                wo.status = 'completed'
                wo.actual_end_date = datetime.utcnow()
                
                if wo.machine:
                    wo.machine.status = 'idle'
                
                # Auto-receive finished goods to warehouse
                qty_good = (float(wo.quantity_good) if wo.quantity_good else 0) or (float(wo.quantity_produced) if wo.quantity_produced else 0)
                
                if qty_good > 0:
                    success, message, inventory_id = auto_receive_finished_goods(wo.id, qty_good, user_id)
                
                completed_list.append({
                    'id': wo.id,
                    'wo_number': wo.wo_number,
                    'quantity_good': qty_good
                })
            except Exception as item_err:
                errors.append({
                    'id': wo.id,
                    'wo_number': wo.wo_number,
                    'error': str(item_err)
                })
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{len(completed_list)} Work Order berhasil diselesaikan',
            'completed': len(completed_list),
            'completed_list': completed_list,
            'errors': errors
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@production_bp.route('/production-records', methods=['GET'])
@jwt_required()
def get_production_records():
    """Get production records"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        
        records = ProductionRecord.query.order_by(ProductionRecord.production_date.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        return jsonify({
            'records': [{
                'id': r.id,
                'work_order_id': r.work_order_id,
                'work_order_number': r.work_order.wo_number if r.work_order else '',
                'machine_id': r.machine_id,
                'machine_name': r.machine.name if r.machine else '',
                'production_date': r.production_date.isoformat(),
                'shift': r.shift,
                'quantity_produced': r.quantity_produced,
                'quantity_good': r.quantity_good,
                'quantity_scrap': r.quantity_scrap,
                'uom': r.uom,
                'downtime_minutes': r.downtime_minutes,
                'efficiency': (r.quantity_good / r.quantity_produced * 100) if r.quantity_produced > 0 else 0
            } for r in records.items],
            'total': records.total,
            'pages': records.pages
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@production_bp.route('/production-records', methods=['POST'])
@jwt_required()
def create_production_record():
    try:
        data = request.get_json()
        
        # Get product_id - use provided or fallback to WO's product
        product_id = data.get('product_id')
        if not product_id:
            wo = db.session.get(WorkOrder, data['work_order_id'])
            product_id = wo.product_id if wo else None
        
        record = ProductionRecord(
            work_order_id=data['work_order_id'],
            product_id=product_id,
            machine_id=data.get('machine_id'),
            operator_id=data.get('operator_id'),
            production_date=datetime.utcnow(),
            shift=data.get('shift'),
            quantity_produced=data['quantity_produced'],
            quantity_good=data['quantity_good'],
            quantity_scrap=data.get('quantity_scrap', 0),
            uom=data['uom'],
            downtime_minutes=data.get('downtime_minutes', 0),
            notes=data.get('notes')
        )
        
        db.session.add(record)
        
        # Update work order quantities - convert to float to avoid Decimal conflicts
        wo = db.session.get(WorkOrder, data['work_order_id'])
        wo.quantity_produced = float(wo.quantity_produced or 0) + float(data['quantity_produced'])
        wo.quantity_good = float(wo.quantity_good or 0) + float(data['quantity_good'])
        wo.quantity_scrap = float(wo.quantity_scrap or 0) + float(data.get('quantity_scrap', 0))
        
        db.session.commit()
        return jsonify({'message': 'Production record created', 'record_id': record.id}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/production-records/<int:record_id>', methods=['GET'])
@jwt_required()
def get_production_record(record_id):
    """Get single production record by ID"""
    try:
        record = db.session.get(ProductionRecord, record_id)
        if not record:
            return jsonify({'error': 'Production record not found'}), 404
        
        return jsonify({
            'record': {
                'id': record.id,
                'work_order_id': record.work_order_id,
                'work_order_number': record.work_order.wo_number if record.work_order else '',
                'product_id': record.product_id,
                'product': {
                    'id': record.product.id,
                    'code': record.product.code,
                    'name': record.product.name
                } if record.product else None,
                'product_name': record.product.name if record.product else (record.work_order.product.name if record.work_order and record.work_order.product else None),
                'machine_id': record.machine_id,
                'production_date': record.production_date.isoformat() if record.production_date else None,
                'shift': record.shift,
                'quantity_produced': float(record.quantity_produced) if record.quantity_produced else 0,
                'quantity_good': float(record.quantity_good) if record.quantity_good else 0,
                'quantity_scrap': float(record.quantity_scrap) if record.quantity_scrap else 0,
                'quantity_rework': float(record.quantity_rework) if hasattr(record, 'quantity_rework') and record.quantity_rework else 0,
                'setting_sticker': float(record.setting_sticker) if hasattr(record, 'setting_sticker') and record.setting_sticker else 0,
                'setting_packaging': float(record.setting_packaging) if hasattr(record, 'setting_packaging') and record.setting_packaging else 0,
                'average_time': record.average_time if hasattr(record, 'average_time') and record.average_time else None,
                'machine_speed': record.machine_speed if hasattr(record, 'machine_speed') and record.machine_speed else None,
                'uom': record.uom,
                'downtime_minutes': record.downtime_minutes or 0,
                'operator_id': record.operator_id,
                'operator_name': record.operator.full_name if record.operator else None,
                'notes': record.notes,
                'early_stop': record.early_stop if hasattr(record, 'early_stop') else False,
                'early_stop_time': str(record.early_stop_time) if hasattr(record, 'early_stop_time') and record.early_stop_time else None,
                'early_stop_reason': record.early_stop_reason if hasattr(record, 'early_stop_reason') else None,
                'early_stop_notes': record.early_stop_notes if hasattr(record, 'early_stop_notes') else None,
            }
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@production_bp.route('/production-records/<int:record_id>', methods=['PUT'])
@jwt_required()
def update_production_record(record_id):
    """Update production record and adjust work order totals and ShiftProduction"""
    try:
        from models.production import ShiftProduction
        
        record = db.session.get(ProductionRecord, record_id)
        if not record:
            return jsonify({'error': 'Production record not found'}), 404
        
        data = request.get_json()
        wo = record.work_order
        
        # Store old values to calculate difference
        old_produced = float(record.quantity_produced) if record.quantity_produced else 0
        old_good = float(record.quantity_good) if record.quantity_good else 0
        old_scrap = float(record.quantity_scrap) if record.quantity_scrap else 0
        old_shift = record.shift
        old_date = record.production_date.date() if record.production_date else None
        
        # Update record fields
        if 'production_date' in data:
            record.production_date = datetime.fromisoformat(data['production_date']) if data['production_date'] else record.production_date
        if 'shift' in data:
            record.shift = data['shift']
        if 'quantity_produced' in data:
            record.quantity_produced = data['quantity_produced']
        if 'quantity_good' in data:
            record.quantity_good = data['quantity_good']
        if 'quantity_scrap' in data:
            record.quantity_scrap = data['quantity_scrap']
        if 'quantity_rework' in data:
            record.quantity_rework = data['quantity_rework']
        if 'setting_sticker' in data:
            record.setting_sticker = data['setting_sticker']
        if 'setting_packaging' in data:
            record.setting_packaging = data['setting_packaging']
        if 'downtime_minutes' in data:
            record.downtime_minutes = data['downtime_minutes']
        if 'operator_id' in data:
            record.operator_id = data['operator_id']
        if 'notes' in data:
            record.notes = data['notes']
        
        # Handle product_id change (for multi-product per shift feature)
        new_product_id = data.get('product_id')
        old_product_id = record.product_id  # Save old product_id BEFORE updating
        if new_product_id:
            record.product_id = new_product_id
        
        # Calculate new values
        new_produced = float(record.quantity_produced) if record.quantity_produced else 0
        new_good = float(record.quantity_good) if record.quantity_good else 0
        new_scrap = float(record.quantity_scrap) if record.quantity_scrap else 0
        new_rework = float(data.get('quantity_rework', 0))
        
        # Update work order totals with difference
        if wo:
            wo.quantity_produced = float(wo.quantity_produced or 0) - old_produced + new_produced
            wo.quantity_good = float(wo.quantity_good or 0) - old_good + new_good
            wo.quantity_scrap = float(wo.quantity_scrap or 0) - old_scrap + new_scrap
        
        # Find and update corresponding ShiftProduction record
        shift_key = f"shift_{record.shift}" if not str(record.shift).startswith('shift_') else record.shift
        production_date = record.production_date.date() if record.production_date else old_date
        
        # Use OLD product_id for lookup (the DB still has the old value)
        lookup_product_id = old_product_id or (wo.product_id if wo else None)
        
        print(f"[UPDATE] Looking for ShiftProduction: wo={record.work_order_id}, date={production_date}, shift={shift_key}, product_id={lookup_product_id} (old_product_id={old_product_id}, new_product_id={new_product_id})")
        
        # Try to find ShiftProduction by work_order_id, date, shift, AND product_id (for multi-product per shift)
        shift_production = ShiftProduction.query.filter(
            ShiftProduction.work_order_id == record.work_order_id,
            ShiftProduction.production_date == production_date,
            ShiftProduction.shift == shift_key,
            ShiftProduction.product_id == lookup_product_id
        ).first()
        
        # If not found with product_id, ONLY fallback if there's exactly ONE match (avoid wrong row in multi-product shifts)
        if not shift_production:
            candidates = ShiftProduction.query.filter(
                ShiftProduction.work_order_id == record.work_order_id,
                ShiftProduction.production_date == production_date,
                ShiftProduction.shift == shift_key
            ).all()
            if len(candidates) == 1:
                shift_production = candidates[0]
                print(f"[UPDATE] Found single ShiftProduction fallback: id={shift_production.id}, product_id={shift_production.product_id}")
            elif len(candidates) > 1:
                print(f"[UPDATE] WARNING: {len(candidates)} ShiftProduction candidates found, cannot determine which to update")
        
        # If still not found, try with old shift and date
        if not shift_production and old_shift:
            old_shift_key = f"shift_{old_shift}" if not str(old_shift).startswith('shift_') else old_shift
            shift_production = ShiftProduction.query.filter(
                ShiftProduction.work_order_id == record.work_order_id,
                ShiftProduction.production_date == old_date,
                ShiftProduction.shift == old_shift_key,
                ShiftProduction.product_id == lookup_product_id
            ).first()
        
        # Update ShiftProduction if found
        if shift_production:
            # Update product_id if changed (multi-product per shift feature)
            if new_product_id:
                shift_production.product_id = new_product_id
            
            # Convert to float to avoid Decimal conflicts
            shift_production.actual_quantity = float(new_produced)
            shift_production.good_quantity = float(new_good)
            shift_production.reject_quantity = float(new_scrap)
            shift_production.rework_quantity = float(new_rework)
            if 'setting_sticker' in data:
                shift_production.setting_sticker = float(data.get('setting_sticker', 0))
            if 'setting_packaging' in data:
                shift_production.setting_packaging = float(data.get('setting_packaging', 0))
            shift_production.downtime_minutes = float(data.get('downtime_minutes', float(shift_production.downtime_minutes or 0)))
            
            # Update Early Stop fields
            if 'early_stop' in data:
                shift_production.early_stop = data.get('early_stop', False)
            if 'early_stop_time' in data:
                from datetime import datetime as dt
                shift_production.early_stop_time = dt.strptime(data['early_stop_time'], '%H:%M').time() if data['early_stop_time'] else None
            if 'early_stop_reason' in data:
                shift_production.early_stop_reason = data.get('early_stop_reason')
            if 'early_stop_notes' in data:
                shift_production.early_stop_notes = data.get('early_stop_notes')
            
            # Update Operator Reassignment fields
            if 'operator_reassigned' in data:
                shift_production.operator_reassigned = data.get('operator_reassigned', False)
            if 'reassignment_task' in data:
                shift_production.reassignment_task = data.get('reassignment_task')
            
            # Recalculate quality rate and OEE
            if new_produced > 0:
                shift_production.quality_rate = round((new_good / new_produced) * 100, 2)
            
            # Update downtime by category if provided
            if 'downtime_entries' in data:
                downtime_entries = data['downtime_entries']
                downtime_mesin = sum(e.get('total_minutes', 0) for e in downtime_entries if e.get('category') == 'mesin')
                downtime_operator = sum(e.get('total_minutes', 0) for e in downtime_entries if e.get('category') == 'operator')
                downtime_material = sum(e.get('total_minutes', 0) for e in downtime_entries if e.get('category') == 'material')
                downtime_design = sum(e.get('total_minutes', 0) for e in downtime_entries if e.get('category') == 'design')
                downtime_others = sum(e.get('total_minutes', 0) for e in downtime_entries if e.get('category') == 'others')
                
                shift_production.downtime_mesin = downtime_mesin
                shift_production.downtime_operator = downtime_operator
                shift_production.downtime_material = downtime_material
                shift_production.downtime_design = downtime_design
                shift_production.downtime_others = downtime_others
                
                # Update issues/notes
                downtime_notes = '; '.join([
                    f"{e.get('total_minutes', 0)} menit - {e.get('reason', '')} [{e.get('category', 'others')}]" 
                    for e in downtime_entries if e.get('reason')
                ])
                if downtime_notes:
                    shift_production.issues = downtime_notes
            
            # Recalculate efficiency and OEE
            planned_runtime = shift_production.planned_runtime or 480
            total_downtime = float(shift_production.downtime_minutes or 0)
            actual_runtime = planned_runtime - total_downtime
            shift_production.actual_runtime = actual_runtime
            
            efficiency_rate = (actual_runtime / planned_runtime * 100) if planned_runtime > 0 else 100
            shift_production.efficiency_rate = round(efficiency_rate, 2)
            
            quality_rate = shift_production.quality_rate or 100
            shift_production.oee_score = round((efficiency_rate * quality_rate) / 100, 2)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Production record updated successfully',
            'record': {
                'id': record.id,
                'quantity_produced': float(record.quantity_produced),
                'quantity_good': float(record.quantity_good),
                'quantity_scrap': float(record.quantity_scrap)
            },
            'shift_production_updated': shift_production is not None
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/bom', methods=['GET'])
@jwt_required()
def get_boms():
    """
    Get all active Bill of Materials
    ---
    tags:
      - Production
    summary: Get all BOMs
    description: Retrieve all active Bill of Materials
    security:
      - BearerAuth: []
    responses:
      200:
        description: BOMs retrieved successfully
        schema:
          type: object
          properties:
            boms:
              type: array
              items:
                type: object
                properties:
                  id:
                    type: integer
                  bom_number:
                    type: string
                  product_name:
                    type: string
                  version:
                    type: string
                  batch_size:
                    type: number
                  item_count:
                    type: integer
      401:
        description: Unauthorized
      500:
        description: Server error
    """
    try:
        boms = BillOfMaterials.query.filter_by(is_active=True).all()
        return jsonify({
            'boms': [{
                'id': b.id,
                'bom_number': b.bom_number,
                'product_name': get_product_name_from_new(b.product.code if b.product else None) or (b.product.name if b.product else 'Unknown'),
                'version': b.version,
                'batch_size': float(b.batch_size),
                'item_count': len(b.items)
            } for b in boms]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@production_bp.route('/bom', methods=['POST'])
@jwt_required()
def create_bom():
    """
    Create a new Bill of Materials
    ---
    tags:
      - Production
    summary: Create new BOM
    description: Create a new Bill of Materials with items
    security:
      - BearerAuth: []
    parameters:
      - in: body
        name: body
        required: true
        schema:
          type: object
          required:
            - product_id
            - batch_size
            - batch_uom
            - items
          properties:
            product_id:
              type: integer
            version:
              type: string
              default: 1.0
            batch_size:
              type: number
            batch_uom:
              type: string
            notes:
              type: string
            items:
              type: array
              items:
                type: object
                properties:
                  material_id:
                    type: integer
                  quantity:
                    type: number
                  uom:
                    type: string
                  scrap_percent:
                    type: number
    responses:
      201:
        description: BOM created successfully
        schema:
          type: object
          properties:
            message:
              type: string
            bom_id:
              type: integer
            bom_number:
              type: string
      401:
        description: Unauthorized
      500:
        description: Server error
    """
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        
        bom_number = generate_number('BOM', BillOfMaterials, 'bom_number')
        
        bom = BillOfMaterials(
            bom_number=bom_number,
            product_id=data['product_id'],
            version=data.get('version', '1.0'),
            batch_size=data['batch_size'],
            batch_uom=data['batch_uom'],
            notes=data.get('notes'),
            created_by=user_id
        )
        
        db.session.add(bom)
        db.session.flush()
        
        for idx, item_data in enumerate(data.get('items', []), 1):
            bom_item = BOMItem(
                bom_id=bom.id,
                line_number=idx,
                material_id=item_data['material_id'],
                quantity=item_data['quantity'],
                uom=item_data['uom'],
                scrap_percent=item_data.get('scrap_percent', 0)
            )
            db.session.add(bom_item)
        
        db.session.commit()
        return jsonify({'message': 'BOM created', 'bom_id': bom.id, 'bom_number': bom_number}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/schedules', methods=['GET'])
@jwt_required()
def get_schedules():
    try:
        schedules = ProductionSchedule.query.order_by(ProductionSchedule.scheduled_start).all()
        return jsonify({
            'schedules': [{
                'id': s.id,
                'schedule_number': s.schedule_number,
                'wo_number': s.work_order.wo_number,
                'machine_name': s.machine.name,
                'scheduled_start': s.scheduled_start.isoformat(),
                'scheduled_end': s.scheduled_end.isoformat(),
                'status': s.status
            } for s in schedules]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============= ADVANCED SCHEDULING =============
@production_bp.route('/schedules', methods=['POST'])
@jwt_required()
def create_schedule():
    try:
        data = request.get_json()
        user_id = get_jwt_identity()
        
        schedule_number = generate_number('SCH', ProductionSchedule, 'schedule_number')
        
        schedule = ProductionSchedule(
            schedule_number=schedule_number,
            work_order_id=data['work_order_id'],
            machine_id=data['machine_id'],
            scheduled_start=datetime.fromisoformat(data['scheduled_start']),
            scheduled_end=datetime.fromisoformat(data['scheduled_end']),
            status='scheduled',
            shift=data.get('shift'),
            notes=data.get('notes'),
            created_by=user_id
        )
        
        db.session.add(schedule)
        db.session.commit()
        
        return jsonify({'message': 'Schedule created', 'schedule_id': schedule.id, 'schedule_number': schedule_number}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/traceability/<search_term>', methods=['GET'])
@jwt_required()
def get_traceability(search_term):
    """Get complete traceability information for a batch or work order"""
    try:
        # Find work order by batch number or WO number
        work_order = WorkOrder.query.filter_by(batch_number=search_term).first()
        if not work_order:
            # Try searching by WO number
            work_order = WorkOrder.query.filter_by(wo_number=search_term).first()
        if not work_order:
            # Try partial match on WO number
            work_order = WorkOrder.query.filter(WorkOrder.wo_number.ilike(f'%{search_term}%')).first()
        if not work_order:
            return jsonify(error_response('api.error', error_code=404)), 404
        
        # Get production records
        production_records = ProductionRecord.query.filter_by(work_order_id=work_order.id).all()
        
        # Build production records safely
        records_list = []
        for record in production_records:
            try:
                prod_date = record.production_date
                if hasattr(prod_date, 'isoformat'):
                    prod_date_str = prod_date.isoformat()
                elif hasattr(prod_date, 'strftime'):
                    prod_date_str = prod_date.strftime('%Y-%m-%d')
                else:
                    prod_date_str = str(prod_date) if prod_date else None
                    
                records_list.append({
                    'id': record.id,
                    'production_date': prod_date_str,
                    'shift': record.shift,
                    'machine_name': record.machine.name if record.machine else None,
                    'operator_name': record.operator.full_name if record.operator else None,
                    'quantity_produced': float(record.quantity_produced or 0),
                    'quantity_good': float(record.quantity_good or 0),
                    'quantity_scrap': float(record.quantity_scrap or 0),
                    'downtime_minutes': int(record.downtime_minutes or 0),
                    'notes': record.notes
                })
            except Exception as rec_err:
                print(f"Error processing record {record.id}: {rec_err}")
                continue
        
        return jsonify({
            'batch_number': work_order.batch_number or work_order.wo_number,
            'work_order': {
                'id': work_order.id,
                'wo_number': work_order.wo_number,
                'product_name': get_product_name_from_new(work_order.product.code if work_order.product else None) or (work_order.product.name if work_order.product else 'Unknown'),
                'quantity': float(work_order.quantity or 0),
                'quantity_produced': float(work_order.quantity_produced or 0),
                'status': work_order.status,
                'machine_name': work_order.machine.name if work_order.machine else None,
                'scheduled_start_date': work_order.scheduled_start_date.isoformat() if work_order.scheduled_start_date else None,
                'scheduled_end_date': work_order.scheduled_end_date.isoformat() if work_order.scheduled_end_date else None,
                'actual_start_date': work_order.actual_start_date.isoformat() if work_order.actual_start_date else None,
                'actual_end_date': work_order.actual_end_date.isoformat() if work_order.actual_end_date else None
            },
            'production_records': records_list
        }), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@production_bp.route('/dashboard/summary', methods=['GET'])
@jwt_required()
def get_production_dashboard():
    """Get production dashboard summary"""
    try:
        # Work Orders Summary
        total_wos = WorkOrder.query.count()
        active_wos = WorkOrder.query.filter(WorkOrder.status.in_(['planned', 'released', 'in_progress'])).count()
        completed_wos = WorkOrder.query.filter_by(status='completed').count()
        
        # Machine Status
        machines = Machine.query.filter_by(is_active=True).all()
        machine_status = {}
        for machine in machines:
            status = machine.status
            machine_status[status] = machine_status.get(status, 0) + 1
        
        return jsonify({
            'work_orders': {
                'total': total_wos,
                'active': active_wos,
                'completed': completed_wos
            },
            'machines': {
                'total_active': len(machines),
                'status_breakdown': machine_status
            }
        }), 200
    except Exception as e:
        print(f"Production dashboard error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============= PACKING LIST =============
from models.production import PackingList, PackingListItem

@production_bp.route('/work-orders/<int:wo_id>/packing-list', methods=['GET'])
@jwt_required()
def get_packing_list(wo_id):
    """Get packing list for a work order"""
    try:
        work_order = db.session.get(WorkOrder, wo_id)
        if not work_order:
            return jsonify({'error': 'Work order not found'}), 404
        
        # Get or create packing list
        packing_list = PackingList.query.filter_by(work_order_id=wo_id).first()
        
        if not packing_list:
            # Create new packing list
            packing_list = PackingList(
                work_order_id=wo_id,
                product_name=get_product_name_from_new(work_order.product.code if work_order.product else None) or (work_order.product.name if work_order.product else 'Unknown'),
                total_karton=0,
                last_carton_number=0
            )
            db.session.add(packing_list)
            db.session.commit()
        
        # Get items with pagination
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        items_query = PackingListItem.query.filter_by(packing_list_id=packing_list.id)\
            .order_by(PackingListItem.carton_number)
        
        paginated = items_query.paginate(page=page, per_page=per_page, error_out=False)
        
        return jsonify({
            'packing_list': packing_list.to_dict(),
            'items': [item.to_dict() for item in paginated.items],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': paginated.total,
                'pages': paginated.pages
            }
        }), 200
    except Exception as e:
        print(f"Get packing list error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@production_bp.route('/work-orders/<int:wo_id>/packing-list/sync', methods=['POST'])
@jwt_required()
def sync_packing_list(wo_id):
    """Sync packing list items based on actual karton count with user-defined start number"""
    try:
        work_order = db.session.get(WorkOrder, wo_id)
        if not work_order:
            return jsonify({'error': 'Work order not found'}), 404
        
        data = request.get_json() or {}
        total_karton = data.get('total_karton', 0)
        start_carton_number = data.get('start_carton_number', 1)
        
        # Validate start_carton_number (1-10000)
        if start_carton_number < 1:
            start_carton_number = 1
        if start_carton_number > 10000:
            start_carton_number = ((start_carton_number - 1) % 10000) + 1
        
        # Get or create packing list
        packing_list = PackingList.query.filter_by(work_order_id=wo_id).first()
        
        if not packing_list:
            packing_list = PackingList(
                work_order_id=wo_id,
                product_name=get_product_name_from_new(work_order.product.code if work_order.product else None) or (work_order.product.name if work_order.product else 'Unknown'),
                total_karton=total_karton,
                start_carton_number=start_carton_number,
                last_carton_number=0
            )
            db.session.add(packing_list)
            db.session.flush()
        else:
            # Update start_carton_number if changed
            packing_list.start_carton_number = start_carton_number
        
        # Delete existing items and recreate with new numbering
        PackingListItem.query.filter_by(packing_list_id=packing_list.id).delete()
        
        # Create items with sequential carton numbers starting from start_carton_number
        for i in range(total_karton):
            # Calculate carton number (reset at 10000)
            carton_num = ((start_carton_number + i - 1) % 10000) + 1
            
            item = PackingListItem(
                packing_list_id=packing_list.id,
                carton_number=carton_num,
                batch_mixing=packing_list.current_batch_mixing,
                is_batch_start=(i == 0)  # First item is batch start
            )
            db.session.add(item)
        
        # Update last carton number
        if total_karton > 0:
            packing_list.last_carton_number = ((start_carton_number + total_karton - 2) % 10000) + 1
        
        packing_list.total_karton = total_karton
        db.session.commit()
        
        return jsonify({
            'success': True,
            'packing_list': packing_list.to_dict(),
            'items_created': total_karton
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"Sync packing list error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@production_bp.route('/work-orders/<int:wo_id>/packing-list/items', methods=['PUT'])
@jwt_required()
def update_packing_list_items(wo_id):
    """Update packing list items (weight, batch mixing)"""
    try:
        packing_list = PackingList.query.filter_by(work_order_id=wo_id).first()
        if not packing_list:
            return jsonify({'error': 'Packing list not found'}), 404
        
        data = request.get_json()
        items_data = data.get('items', [])
        
        for item_data in items_data:
            item_id = item_data.get('id')
            if item_id:
                item = db.session.get(PackingListItem, item_id)
                if item and item.packing_list_id == packing_list.id:
                    if 'weight_kg' in item_data:
                        item.weight_kg = item_data['weight_kg']
                    if 'batch_mixing' in item_data:
                        item.batch_mixing = item_data['batch_mixing']
                    if 'is_batch_start' in item_data:
                        item.is_batch_start = item_data['is_batch_start']
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Items updated successfully'
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"Update packing list items error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@production_bp.route('/work-orders/<int:wo_id>/packing-list/batch-mixing', methods=['POST'])
@jwt_required()
def set_batch_mixing(wo_id):
    """Set new batch mixing for subsequent cartons"""
    try:
        packing_list = PackingList.query.filter_by(work_order_id=wo_id).first()
        if not packing_list:
            return jsonify({'error': 'Packing list not found'}), 404
        
        data = request.get_json()
        batch_mixing = data.get('batch_mixing', '')
        start_from_carton = data.get('start_from_carton')  # Optional: which carton to start from
        
        packing_list.current_batch_mixing = batch_mixing
        
        # If start_from_carton specified, update that item as batch start
        if start_from_carton:
            item = PackingListItem.query.filter_by(
                packing_list_id=packing_list.id,
                carton_number=start_from_carton
            ).first()
            if item:
                item.is_batch_start = True
                item.batch_mixing = batch_mixing
                
                # Update all subsequent items with new batch
                subsequent_items = PackingListItem.query.filter(
                    PackingListItem.packing_list_id == packing_list.id,
                    PackingListItem.carton_number >= start_from_carton
                ).all()
                for sub_item in subsequent_items:
                    sub_item.batch_mixing = batch_mixing
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'current_batch_mixing': packing_list.current_batch_mixing
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"Set batch mixing error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============= REMAINING STOCK (SISA ORDER) =============
@production_bp.route('/remaining-stocks', methods=['GET'])
@jwt_required()
def get_remaining_stocks():
    """Get all remaining stocks (Sisa Order)"""
    try:
        stocks = RemainingStock.query.order_by(RemainingStock.product_name).all()
        return jsonify({
            'remaining_stocks': [s.to_dict() for s in stocks],
            'total': len(stocks)
        }), 200
    except Exception as e:
        print(f"Get remaining stocks error: {e}")
        return jsonify({'error': str(e)}), 500


@production_bp.route('/remaining-stocks', methods=['POST'])
@jwt_required()
def create_remaining_stock():
    """Create new remaining stock entry"""
    try:
        user_id = get_jwt_identity()
        data = request.get_json()
        
        product_name = data.get('product_name')
        if not product_name:
            return jsonify({'error': 'Nama produk harus diisi'}), 400
        
        qty_karton = data.get('qty_karton', 0)
        if qty_karton <= 0:
            return jsonify({'error': 'Qty karton harus lebih dari 0'}), 400
        
        # If product_id provided, get product info
        product_id = data.get('product_id')
        product_code = data.get('product_code')
        
        if product_id:
            product = db.session.get(Product, product_id)
            if product:
                product_name = product.name
                product_code = product.code
        
        stock = RemainingStock(
            product_id=product_id,
            product_name=product_name,
            product_code=product_code,
            qty_karton=qty_karton,
            qty_pcs=data.get('qty_pcs'),
            pack_per_carton=data.get('pack_per_carton'),
            notes=data.get('notes'),
            location=data.get('location'),
            created_by=user_id
        )
        
        db.session.add(stock)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Sisa order berhasil ditambahkan',
            'remaining_stock': stock.to_dict()
        }), 201
    except Exception as e:
        db.session.rollback()
        print(f"Create remaining stock error: {e}")
        return jsonify({'error': str(e)}), 500


@production_bp.route('/remaining-stocks/<int:id>', methods=['GET'])
@jwt_required()
def get_remaining_stock(id):
    """Get single remaining stock by ID"""
    try:
        stock = db.session.get(RemainingStock, id)
        if not stock:
            return jsonify({'error': 'Data tidak ditemukan'}), 404
        return jsonify({'remaining_stock': stock.to_dict()}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@production_bp.route('/remaining-stocks/<int:id>', methods=['PUT'])
@jwt_required()
def update_remaining_stock(id):
    """Update remaining stock entry"""
    try:
        user_id = get_jwt_identity()
        stock = db.session.get(RemainingStock, id)
        if not stock:
            return jsonify({'error': 'Data tidak ditemukan'}), 404
        
        data = request.get_json()
        
        # Update product info
        if 'product_id' in data:
            stock.product_id = data['product_id']
            if data['product_id']:
                product = db.session.get(Product, data['product_id'])
                if product:
                    stock.product_name = product.name
                    stock.product_code = product.code
        
        if 'product_name' in data:
            stock.product_name = data['product_name']
        if 'product_code' in data:
            stock.product_code = data['product_code']
        if 'qty_karton' in data:
            stock.qty_karton = data['qty_karton']
        if 'qty_pcs' in data:
            stock.qty_pcs = data['qty_pcs']
        if 'pack_per_carton' in data:
            stock.pack_per_carton = data['pack_per_carton']
        if 'notes' in data:
            stock.notes = data['notes']
        if 'location' in data:
            stock.location = data['location']
        
        stock.updated_by = user_id
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Sisa order berhasil diupdate',
            'remaining_stock': stock.to_dict()
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"Update remaining stock error: {e}")
        return jsonify({'error': str(e)}), 500


@production_bp.route('/remaining-stocks/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_remaining_stock(id):
    """Delete remaining stock entry"""
    try:
        stock = db.session.get(RemainingStock, id)
        if not stock:
            return jsonify({'error': 'Data tidak ditemukan'}), 404
        
        db.session.delete(stock)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Sisa order berhasil dihapus'
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"Delete remaining stock error: {e}")
        return jsonify({'error': str(e)}), 500


@production_bp.route('/remaining-stocks/export-excel', methods=['GET'])
@jwt_required()
def export_remaining_stocks_excel():
    """Export remaining stocks to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        stocks = RemainingStock.query.order_by(RemainingStock.product_name).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sisa Order"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="center")
        number_alignment = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = 'LAPORAN SISA ORDER'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Date
        ws.merge_cells('A2:D2')
        ws['A2'] = f'Tanggal: {get_local_now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ['No', 'Nama Produk', 'Kode Produk', 'Qty Karton']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        for row_idx, stock in enumerate(stocks, 5):
            ws.cell(row=row_idx, column=1, value=row_idx - 4).border = thin_border
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
            
            ws.cell(row=row_idx, column=2, value=stock.product_name).border = thin_border
            ws.cell(row=row_idx, column=2).alignment = cell_alignment
            
            ws.cell(row=row_idx, column=3, value=stock.product_code or '-').border = thin_border
            ws.cell(row=row_idx, column=3).alignment = cell_alignment
            
            ws.cell(row=row_idx, column=4, value=float(stock.qty_karton) if stock.qty_karton else 0).border = thin_border
            ws.cell(row=row_idx, column=4).alignment = number_alignment
        
        # Total row
        total_row = len(stocks) + 5
        ws.cell(row=total_row, column=1, value='').border = thin_border
        ws.merge_cells(f'B{total_row}:C{total_row}')
        ws.cell(row=total_row, column=2, value='TOTAL').border = thin_border
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=3).border = thin_border
        
        total_karton = sum(float(s.qty_karton) if s.qty_karton else 0 for s in stocks)
        ws.cell(row=total_row, column=4, value=total_karton).border = thin_border
        ws.cell(row=total_row, column=4).font = Font(bold=True)
        ws.cell(row=total_row, column=4).alignment = number_alignment
        
        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Sisa_Order_{get_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Export remaining stocks error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============= WORK ORDER BOM (Editable copy of BOM for WO) =============

@production_bp.route('/work-orders/<int:wo_id>/bom', methods=['GET'])
@jwt_required()
def get_work_order_bom(wo_id):
    """Get BOM items for a work order (WO-specific copy or from master BOM)"""
    try:
        wo = db.session.get(WorkOrder, wo_id)
        if not wo:
            return jsonify({'error': 'Work order not found'}), 404
        
        # Check if WO has its own BOM items
        wo_bom_items = WorkOrderBOMItem.query.filter_by(work_order_id=wo_id).order_by(WorkOrderBOMItem.line_number).all()
        
        if wo_bom_items:
            # Return WO-specific BOM
            return jsonify({
                'source': 'work_order',
                'work_order_id': wo_id,
                'wo_number': wo.wo_number,
                'bom_items': [{
                    'id': item.id,
                    'line_number': item.line_number,
                    'material_id': item.material_id,
                    'product_id': item.product_id,
                    'item_name': item.item_name,
                    'item_code': item.item_code,
                    'item_type': item.item_type,
                    'quantity_per_unit': float(item.quantity_per_unit) if item.quantity_per_unit else 0,
                    'uom': item.uom,
                    'scrap_percent': float(item.scrap_percent) if item.scrap_percent else 0,
                    'quantity_planned': float(item.quantity_planned) if item.quantity_planned else 0,
                    'quantity_actual': float(item.quantity_actual) if item.quantity_actual else 0,
                    'unit_cost': float(item.unit_cost) if item.unit_cost else 0,
                    'is_modified': item.is_modified,
                    'is_added': item.is_added,
                    'modification_reason': item.modification_reason,
                    'notes': item.notes
                } for item in wo_bom_items]
            }), 200
        
        # If no WO-specific BOM, find master BOM
        # First check if bom_id is set, otherwise find active BOM by product_id
        bom = None
        if wo.bom_id:
            bom = db.session.get(BillOfMaterials, wo.bom_id)
        
        if not bom and wo.product_id:
            # Find active BOM for this product
            bom = BillOfMaterials.query.filter_by(
                product_id=wo.product_id,
                is_active=True
            ).first()
        
        if bom:
            return jsonify({
                'source': 'master_bom',
                'work_order_id': wo_id,
                'wo_number': wo.wo_number,
                'bom_id': bom.id,
                'bom_number': bom.bom_number,
                'bom_items': [{
                    'id': item.id,
                    'line_number': item.line_number,
                    'material_id': item.material_id,
                    'product_id': item.product_id,
                    'item_name': item.item_name,
                    'item_code': item.item_code,
                    'item_type': item.item_type,
                    'quantity': float(item.quantity) if item.quantity else 0,
                    'uom': item.uom,
                    'scrap_percent': float(item.scrap_percent) if item.scrap_percent else 0,
                    'unit_cost': float(item.unit_cost) if item.unit_cost else 0,
                    'notes': item.notes
                } for item in bom.items]
            }), 200
        
        return jsonify({
            'source': 'none',
            'work_order_id': wo_id,
            'wo_number': wo.wo_number,
            'bom_items': [],
            'message': 'No BOM associated with this work order'
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@production_bp.route('/work-orders/<int:wo_id>/bom/copy-from-master', methods=['POST'])
@jwt_required()
def copy_bom_to_work_order(wo_id):
    """Copy BOM items from master BOM to work order for editing"""
    try:
        user_id = get_jwt_identity()
        wo = db.session.get(WorkOrder, wo_id)
        if not wo:
            return jsonify({'error': 'Work order not found'}), 404
        
        # Find BOM - first by bom_id, then by product_id
        bom = None
        if wo.bom_id:
            bom = db.session.get(BillOfMaterials, wo.bom_id)
        
        if not bom and wo.product_id:
            # Find active BOM for this product
            bom = BillOfMaterials.query.filter_by(
                product_id=wo.product_id,
                is_active=True
            ).first()
        
        if not bom:
            return jsonify({'error': 'No BOM found for this work order'}), 404
        
        # Check if WO already has BOM items
        existing = WorkOrderBOMItem.query.filter_by(work_order_id=wo_id).first()
        if existing:
            return jsonify({'error': 'Work order already has BOM items. Delete them first to re-copy.'}), 400
        
        # Copy BOM items to WO
        wo_quantity = float(wo.quantity) if wo.quantity else 0
        
        for bom_item in bom.items:
            qty_per_unit = float(bom_item.quantity) if bom_item.quantity else 0
            qty_planned = qty_per_unit * wo_quantity
            
            wo_bom_item = WorkOrderBOMItem(
                work_order_id=wo_id,
                original_bom_id=bom.id,
                original_bom_item_id=bom_item.id,
                line_number=bom_item.line_number,
                material_id=bom_item.material_id,
                product_id=bom_item.product_id,
                item_name=bom_item.item_name,
                item_code=bom_item.item_code,
                item_type=bom_item.item_type,
                quantity_per_unit=bom_item.quantity,
                uom=bom_item.uom,
                scrap_percent=bom_item.scrap_percent,
                quantity_planned=qty_planned,
                unit_cost=bom_item.unit_cost,
                is_modified=False,
                is_added=False,
                notes=bom_item.notes
            )
            db.session.add(wo_bom_item)
        
        db.session.commit()
        
        return jsonify({
            'message': f'BOM items copied from {bom.bom_number} to work order',
            'items_copied': len(bom.items)
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@production_bp.route('/work-orders/<int:wo_id>/bom/<int:item_id>', methods=['PUT'])
@jwt_required()
def update_work_order_bom_item(wo_id, item_id):
    """Update a WO BOM item (does NOT affect master BOM)"""
    try:
        user_id = get_jwt_identity()
        
        item = WorkOrderBOMItem.query.filter_by(id=item_id, work_order_id=wo_id).first()
        if not item:
            return jsonify({'error': 'BOM item not found'}), 404
        
        data = request.get_json()
        
        # Update fields
        if 'quantity_per_unit' in data:
            item.quantity_per_unit = data['quantity_per_unit']
        if 'scrap_percent' in data:
            item.scrap_percent = data['scrap_percent']
        if 'unit_cost' in data:
            item.unit_cost = data['unit_cost']
        if 'notes' in data:
            item.notes = data['notes']
        if 'modification_reason' in data:
            item.modification_reason = data['modification_reason']
        if 'quantity_actual' in data:
            item.quantity_actual = data['quantity_actual']
        
        # Recalculate planned quantity
        wo = db.session.get(WorkOrder, wo_id)
        if wo:
            wo_quantity = float(wo.quantity) if wo.quantity else 0
            qty_per_unit = float(item.quantity_per_unit) if item.quantity_per_unit else 0
            item.quantity_planned = qty_per_unit * wo_quantity
            
            # Calculate variance if actual is set
            if item.quantity_actual:
                item.quantity_variance = float(item.quantity_actual) - float(item.quantity_planned)
        
        # Mark as modified
        item.is_modified = True
        item.modified_by = user_id
        item.modified_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'message': 'BOM item updated (master BOM unchanged)',
            'item': {
                'id': item.id,
                'item_name': item.item_name,
                'quantity_per_unit': float(item.quantity_per_unit),
                'quantity_planned': float(item.quantity_planned) if item.quantity_planned else 0,
                'is_modified': item.is_modified
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@production_bp.route('/work-orders/<int:wo_id>/bom', methods=['POST'])
@jwt_required()
def add_work_order_bom_item(wo_id):
    """Add a new BOM item to work order (manual addition)"""
    try:
        user_id = get_jwt_identity()
        
        wo = db.session.get(WorkOrder, wo_id)
        if not wo:
            return jsonify({'error': 'Work order not found'}), 404
        
        data = request.get_json()
        
        # Get next line number
        max_line = db.session.query(func.max(WorkOrderBOMItem.line_number)).filter_by(work_order_id=wo_id).scalar()
        next_line = (max_line or 0) + 1
        
        # Get item details
        item_name = data.get('item_name', '')
        item_code = data.get('item_code', '')
        item_type = data.get('item_type', '')
        
        # If material_id provided, get details from material
        material_id = data.get('material_id')
        if material_id:
            material = db.session.get(Material, material_id)
            if material:
                item_name = material.name
                item_code = material.code
                item_type = material.material_type
        
        # Calculate planned quantity
        wo_quantity = float(wo.quantity) if wo.quantity else 0
        qty_per_unit = float(data.get('quantity_per_unit', 0))
        qty_planned = qty_per_unit * wo_quantity
        
        item = WorkOrderBOMItem(
            work_order_id=wo_id,
            line_number=next_line,
            material_id=material_id,
            product_id=data.get('product_id'),
            item_name=item_name,
            item_code=item_code,
            item_type=item_type,
            quantity_per_unit=qty_per_unit,
            uom=data.get('uom', 'pcs'),
            scrap_percent=data.get('scrap_percent', 0),
            quantity_planned=qty_planned,
            unit_cost=data.get('unit_cost'),
            is_modified=False,
            is_added=True,  # Mark as manually added
            modified_by=user_id,
            modified_at=datetime.utcnow(),
            modification_reason=data.get('modification_reason', 'Manually added'),
            notes=data.get('notes')
        )
        
        db.session.add(item)
        db.session.commit()
        
        return jsonify({
            'message': 'BOM item added to work order',
            'item': {
                'id': item.id,
                'line_number': item.line_number,
                'item_name': item.item_name,
                'quantity_per_unit': float(item.quantity_per_unit),
                'quantity_planned': float(item.quantity_planned) if item.quantity_planned else 0
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@production_bp.route('/work-orders/<int:wo_id>/bom/<int:item_id>', methods=['DELETE'])
@jwt_required()
def delete_work_order_bom_item(wo_id, item_id):
    """Delete a WO BOM item (does NOT affect master BOM)"""
    try:
        item = WorkOrderBOMItem.query.filter_by(id=item_id, work_order_id=wo_id).first()
        if not item:
            return jsonify({'error': 'BOM item not found'}), 404
        
        db.session.delete(item)
        db.session.commit()
        
        return jsonify({'message': 'BOM item deleted from work order (master BOM unchanged)'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@production_bp.route('/work-orders/<int:wo_id>/bom/reset', methods=['POST'])
@jwt_required()
def reset_work_order_bom(wo_id):
    """Delete all WO BOM items and re-copy from master BOM"""
    try:
        wo = db.session.get(WorkOrder, wo_id)
        if not wo:
            return jsonify({'error': 'Work order not found'}), 404
        
        # Delete existing WO BOM items
        WorkOrderBOMItem.query.filter_by(work_order_id=wo_id).delete()
        db.session.commit()
        
        return jsonify({'message': 'Work order BOM items deleted. You can now copy from master BOM again.'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
